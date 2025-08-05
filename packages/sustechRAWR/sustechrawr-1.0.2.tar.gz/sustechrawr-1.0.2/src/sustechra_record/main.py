import argparse
import calendar
from concurrent.futures import ThreadPoolExecutor, as_completed
from get_holiday_cn.client import getHoliday
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side


def draw_excel(headers, rows, sheet_name, out_file):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # title
    ws.merge_cells("A1:F1")
    ws["A1"] = "工作情况记录表"
    ws["A1"].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws["A1"].font = Font(size=18, name="仿宋")

    # header
    ws.append(headers)

    # data
    for row in rows:
        ws.append(row)
    # merge teacher confirmation signature
    ws.merge_cells(f"F{3}:F{len(rows) + 2}")

    # center align
    alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
    ):
        for cell in row:
            cell.alignment = alignment

    # set width
    widths = [8, 10.33, 17.5, 40, 13.67, 11.5]
    for i, width in enumerate(widths, start=1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 54

    for i in range(3, len(rows) + 3):
        ws.row_dimensions[i].height = 26

    target_row_number = 2
    for cell in ws[target_row_number]:
        cell.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )

    # set border
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    min_row = 2  # 从标题行开始
    max_row = len(rows) + 2  # 1（标题）+1（表头）+数据行数
    min_col = 1
    max_col = len(headers)
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.border = border

    font1 = Font(size=14, name="仿宋")
    font2 = Font(size=12, name="仿体")

    for row in range(2, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if row == 2 or col == 1 or col == 5 or col == 6:
                cell.font = font1
            else:
                cell.font = font2

    wb.save(out_file)


def main(yy, mm, name, work_content, time_duration, out_file):
    _, dd = calendar.monthrange(yy, mm)

    client = getHoliday()
    is_holiday_mp = {}
    with ThreadPoolExecutor() as executor:
        future_to_date = {
            executor.submit(client.assemble_holiday_data, f"{yy}-{mm}-{i:02d}"): i
            for i in range(1, dd + 1)
        }
        for future in as_completed(future_to_date):
            date_str = future_to_date[future]
            try:
                data = future.result()
                day_type = data["type"][
                    "type"
                ]  # workday, weekend, holiday, compensatory work
                is_holiday_mp[int(date_str)] = not (day_type == 0 or day_type == 3)
            except Exception as exc:
                print(f"Date: {date_str} generated an exception: {exc}")

    headers = [
        "日期",
        "访问人员姓名",
        "在岗时间（按日签）",
        "工作内容",
        "访问人员确认（按日签）",
        "老师确认签名",
    ]
    holidays_list = [is_holiday_mp[i] for i in range(1, dd + 1)]
    rows = []

    for i in range(1, dd + 1):
        if holidays_list[i - 1]:
            rows.append([f"{mm}.{i}", "/", "/", "/", "/"])
        else:
            rows.append([f"{mm}.{i}", name, time_duration, work_content, ""])

    draw_excel(headers, rows, name, out_file)


def run():
    parser = argparse.ArgumentParser(
        description="Generate work record table Excel file"
    )
    parser.add_argument("-n", "--name", required=True, help="Visitor name")
    parser.add_argument("-w", "--work", required=True, help="Work content")
    parser.add_argument(
        "-y", "--year", required=True, type=int, help="Year (default: 2025)"
    )
    parser.add_argument(
        "-m", "--month", required=True, type=int, help="Month (default: 7)"
    )
    parser.add_argument(
        "-t", "--time", default="9:00-18:00", help="Working hours (default: 9:00-18:00)"
    )
    parser.add_argument(
        "-o", "--output", help="Output filename (default: {name}_{year}_{month}.xlsx)"
    )

    args = parser.parse_args()

    yy = args.year
    mm = args.month
    name = args.name
    work_content = args.work
    time_duration = args.time
    out_file = args.output if args.output else f"{name}_{yy}_{mm}.xlsx"

    main(yy, mm, name, work_content, time_duration, out_file)


if __name__ == "__main__":
    run()
