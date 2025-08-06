from copy import copy
import openpyxl
import sys
import os
# import pandas as pd
import csv
from openpyxl.worksheet.cell_range import CellRange

class WorkbookManager(object):
    def copy_sheet(self, src_ws, tgt_ws):
        """
        Полностью копирует лист Excel с сохранением:
        - данных и формул
        - стилей и форматирования
        - скрытия колонок/строк
        - комментариев и гиперссылок
        - объединенных ячеек
        - параметров страницы
        """
        # Копируем параметры колонок (включая скрытие)
        for col_letter, col_dim in src_ws.column_dimensions.items():
            tgt_col = tgt_ws.column_dimensions[col_letter]
            tgt_col.width = col_dim.width
            tgt_col.hidden = col_dim.hidden  # параметр скрытия колонки
            tgt_col.bestFit = col_dim.bestFit
            tgt_col.outline_level = col_dim.outline_level
        
        # Копируем параметры строк (включая скрытие)
        for row_num, row_dim in src_ws.row_dimensions.items():
            tgt_row = tgt_ws.row_dimensions[row_num]
            tgt_row.height = row_dim.height
            tgt_row.hidden = row_dim.hidden  # параметр скрытия строки
            tgt_row.outline_level = row_dim.outline_level
        
        # Копируем данные и стили ячеек
        for row in src_ws.iter_rows():
            for src_cell in row:
                tgt_cell = tgt_ws.cell(
                    row=src_cell.row,
                    column=src_cell.column,
                    value=src_cell.value
                )
                
                # Копируем стиль
                if src_cell.has_style:
                    tgt_cell.font = copy(src_cell.font)
                    tgt_cell.border = copy(src_cell.border)
                    tgt_cell.fill = copy(src_cell.fill)
                    tgt_cell.number_format = copy(src_cell.number_format)
                    tgt_cell.protection = copy(src_cell.protection)
                    tgt_cell.alignment = copy(src_cell.alignment)
                
                # Копируем комментарии и гиперссылки
                if src_cell.comment:
                    tgt_cell.comment = copy(src_cell.comment)
                if src_cell.hyperlink:
                    tgt_cell.hyperlink = copy(src_cell.hyperlink)
        
        # Копируем объединенные ячейки
        for merged_range in src_ws.merged_cells.ranges:
            tgt_ws.merge_cells(str(merged_range))
        
        # Копируем параметры страницы
        tgt_ws.page_margins = copy(src_ws.page_margins)
        tgt_ws.page_setup = copy(src_ws.page_setup)
        tgt_ws.print_options = copy(src_ws.print_options)
        tgt_ws.sheet_properties = copy(src_ws.sheet_properties)
        
        # Копируем фильтры
        if src_ws.auto_filter:
            tgt_ws.auto_filter.ref = src_ws.auto_filter.ref

        # Копируем параметры группировки (outline)
        tgt_ws.sheet_properties.outlinePr = copy(src_ws.sheet_properties.outlinePr)

    def write_to_range(self, sheet, excel_range, data):
        """
        Записать в указанный промежуток данные
        sheet - excel-лист
        excel_range - например A1:B2 - промежуток
        data - данные в виде list [[], []]
        """
        cells = sheet[excel_range]
        # Проверка размеров
        if len(data) != len(cells) or len(data[0]) != len(cells[0]):
            raise ValueError(f"Несоответствие размеров данных и диапазона rows: {len(data)}/{len(cells)} columns: {len(data[0])}/{len(cells[0])}")
        
        # Запись
        for row_cells, row_data in zip(cells, data):
            for cell, value in zip(row_cells, row_data):
                cell.value = value

    def transform(self, meta_data, separator, output_path):
        targetWb = openpyxl.Workbook()   
        targetWb.remove(targetWb.active)
        
        for row in meta_data:
            if os.path.exists(row['template_name']):
                templateWb = openpyxl.load_workbook(row['template_name'])
                templateSheet = templateWb[row['sheet_name']]

            if row['sheet_name'] not in targetWb.sheetnames:
                targetSheet = targetWb.create_sheet(row['sheet_name'])
                self.copy_sheet(templateSheet, targetSheet)
                
            if os.path.exists(row['csv_data_name']):
                with open(row['csv_data_name'], mode='r', encoding='utf-8') as file:
                    reader = csv.reader(file, delimiter=separator)  # автоматически использует первую строку как ключи
                    data = list(reader)[1:]  # преобразуем в список словарей

                    write_data_range = CellRange(row['mapping'])
                    write_data_range.expand(down=len(data) - 1)
                    self.write_to_range(targetSheet, write_data_range.coord, data)

        targetWb.save(output_path)
        return targetWb
