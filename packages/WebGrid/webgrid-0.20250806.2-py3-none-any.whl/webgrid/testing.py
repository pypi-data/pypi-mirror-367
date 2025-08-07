"""
A collection of utilities for testing webgrid functionality in client applications
"""

import re
from unittest import mock
import urllib


try:
    import openpyxl
except ImportError:
    openpyxl = None
from pyquery import PyQuery
import sqlalchemy


def compiler_instance_factory(compiler, dialect, statement):
    class LiteralCompiler(compiler.__class__):
        def render_literal_value(self, value, type_):
            import datetime

            """
            For date and datetime values, convert to a string
            format acceptable to the dialect. That seems to be the
            so-called ODBC canonical date format which looks
            like this:

                yyyy-mm-dd hh:mi:ss.mmm(24h)

            For other data types, call the base class implementation.
            """
            if isinstance(value, datetime.datetime):
                return "'" + value.strftime('%Y-%m-%d %H:%M:%S.%f') + "'"
            elif isinstance(value, datetime.date):
                return "'" + value.strftime('%Y-%m-%d') + "'"
            elif isinstance(value, datetime.time):
                return f"'{value:%H:%M:%S.%f}'"
            elif isinstance(value, datetime.timedelta):
                return str(value)
            elif isinstance(value, str):
                return f"'{value}'"
            elif isinstance(value, list) and isinstance(type_, sqlalchemy.ARRAY):
                elements = [
                    self.render_literal_value(list_val, type_.item_type) for list_val in value
                ]
                return f'({", ".join(elements)})'
            elif value is None:
                return 'NULL'
            else:
                # Turn off double percent escaping, since we don't run these strings and
                # it creates a large number of differences for test cases
                with mock.patch.object(dialect.identifier_preparer, '_double_percents', False):
                    return super().render_literal_value(value, type_)

        def visit_bindparam(
            self,
            bindparam,
            within_columns_clause=False,
            literal_binds=False,
            **kwargs,
        ):
            return super().render_literal_bindparam(
                bindparam,
                within_columns_clause=within_columns_clause,
                literal_binds=literal_binds,
                **kwargs,
            )

        def visit_table(
            self,
            table,
            asfrom=False,
            iscrud=False,
            ashint=False,
            fromhints=None,
            use_schema=True,
            **kwargs,
        ):
            """Strip the default schema from table names when it is not needed"""
            ret_val = super().visit_table(
                table,
                asfrom,
                iscrud,
                ashint,
                fromhints,
                use_schema,
                **kwargs,
            )
            if dialect.name == 'mssql' and ret_val.startswith('dbo.'):
                return ret_val[4:]
            return ret_val

        def visit_column(self, column, add_to_result_map=None, include_table=True, **kwargs):
            """Strip the default schema from table names when it is not needed"""
            ret_val = super().visit_column(column, add_to_result_map, include_table, **kwargs)
            if dialect.name == 'mssql' and ret_val.startswith('dbo.'):
                return ret_val[4:]
            return ret_val

    return LiteralCompiler(dialect, statement)


def query_to_str(statement, bind=None):
    """
    returns a string of a sqlalchemy.orm.Query with parameters bound

    WARNING: this is dangerous and ONLY for testing, executing the results
    of this function can result in an SQL Injection attack.
    """
    if isinstance(statement, sqlalchemy.orm.Query):
        if bind is None:
            bind = statement.session.get_bind()
        statement = statement.statement
    elif bind is None:
        bind = statement.bind

    if bind is None:
        raise Exception(
            'bind param (engine or connection object) required when using with an unbound statement',  # noqa: E501
        )

    dialect = bind.dialect
    compiler = statement._compiler(dialect)
    literal_compiler = compiler_instance_factory(compiler, dialect, statement)
    return 'TESTING ONLY BIND: ' + literal_compiler.process(statement)


def assert_in_query(obj, test_for):
    query = obj.build_query() if hasattr(obj, 'build_query') else obj
    query_str = query_to_str(query)
    assert test_for in query_str, query_str


def assert_not_in_query(obj, test_for):
    query = obj.build_query() if hasattr(obj, 'build_query') else obj
    query_str = query_to_str(query)
    assert test_for not in query_str, query_str


def assert_list_equal(list1, list2):
    """
    A list-specific equality assertion.

    This method is based on the Python `unittest.TestCase.assertListEqual` method.

    :param list1:
    :param list2:
    :return:
    """

    # resolve generators
    list1, list2 = map(list, (list1, list2))

    assert len(list1) == len(list2), f'Lists are different lengths: {len(list1)} != {len(list2)}'

    if list1 == list2:
        # the lists are the same, we're done
        return

    # the lists are different in at least one element; find it
    # and report it
    for index, (val1, val2) in enumerate(zip(list1, list2, strict=True)):
        assert val1 == val2, f'First differing element at index {index}: {val1!r} != {val2!r}'


def assert_rendered_xlsx_matches(rendered_xlsx, xlsx_headers, xlsx_rows):
    """
    Verifies that `rendered_xlsx` has a set of headers and values that match
    the given parameters.

    NOTE: This method does not perform in-depth analysis of complex workbooks!
          Assumes header rows and data rows are contiguous.
          Multiple worksheets or complex layouts *are not verified!*

    :param rendered_xlsx: binary data passed to openpyxl as file contents
    :param xlsx_headers: list of rows of column headers
    :param xlsx_rows: list of rows in order as they will appear in the worksheet
    """
    assert rendered_xlsx
    rendered_xlsx.filename.seek(0)

    if not openpyxl:
        raise Exception('openpyxl is required for webgrid testing helpers to read XLSX')

    book = openpyxl.load_workbook(rendered_xlsx.filename)
    assert len(book.sheetnames) >= 1
    sheet = book[book.sheetnames[0]]

    # # verify the shape of the sheet

    # ## shape of rows (1 row for the headers, 1 for each row of data)
    nrows = len(xlsx_rows)
    if xlsx_headers:
        nrows += len(xlsx_headers)
    assert max([nrows, 1]) == sheet.max_row, (
        f'Sheet max row mismatch, {max([nrows, 1])} != {sheet.max_row}'
    )

    # ## shape of columns
    ncols = max(
        max(len(values) for values in xlsx_headers) if xlsx_headers else 0,
        max(len(values) for values in xlsx_rows) if xlsx_rows else 0,
    )
    assert max([ncols, 1]) == sheet.max_column, (
        f'Sheet max column mismatch, {max([ncols, 1])} != {sheet.max_column}'
    )

    row_iter = sheet.iter_rows()

    expected_rows = (xlsx_headers or []) + (xlsx_rows or [])

    for row, expected_row in zip(row_iter, expected_rows, strict=True):
        assert_list_equal((cell.value for cell in row), expected_row)


class GridBase:
    """Base test class for Flask or Keg apps.

    Class Attributes:
        grid_cls: Application grid class to use during testing

        filters: Iterable of (name, op, value, expected) tuples to check for filter logic,
        or a callable returning such an iterable. `name` is the column key. `op` and `value`
        set the filter parameters. `expected` is either a SQL string or compiled regex to
        find when the filter is enabled.

        sort_tests: Iterable of (name, expected) tuples to check for sort logic. `name` is
        the column key. `expected` is a SQL string to find when the sort is enabled.
    """

    grid_cls = None
    filters = ()
    sort_tests = ()

    @classmethod
    def setup_class(cls):
        if hasattr(cls, 'init'):
            cls.init()

    def query_to_str(self, statement, bind=None):
        """Render a SQLAlchemy query to a string."""
        return query_to_str(statement, bind=bind)

    def assert_in_query(self, look_for, grid=None, _query_string=None, **kwargs):
        """Verify the given SQL string is in the grid's query.

        Args:
            look_for (str): SQL string to find.

            grid (BaseGrid, optional): Grid to use instead of `self.get_session_grid`.
            Defaults to None.

            kwargs (dict, optional): Additional args passed to `self.get_session_grid`.
        """
        grid = grid or self.get_session_grid(_query_string=_query_string, **kwargs)
        assert_in_query(grid, look_for)

    def assert_not_in_query(self, look_for, grid=None, _query_string=None, **kwargs):
        """Verify the given SQL string is not in the grid's query.

        Args:
            look_for (str): SQL string to find.

            grid (BaseGrid, optional): Grid to use instead of `self.get_session_grid`.
            Defaults to None.

            kwargs (dict, optional): Additional args passed to `self.get_session_grid`.
        """
        grid = grid or self.get_session_grid(_query_string=_query_string, **kwargs)
        assert_not_in_query(grid, look_for)

    def assert_regex_in_query(self, look_for, grid=None, _query_string=None, **kwargs):
        """Verify the given regex matches the grid's query.

        Args:
            look_for (str or regex): Regex to search (can be compiled or provided as string).

            grid (BaseGrid, optional): Grid to use instead of `self.get_session_grid`.
            Defaults to None.

            kwargs (dict, optional): Additional args passed to `self.get_session_grid`.
        """
        grid = grid or self.get_session_grid(_query_string=_query_string, **kwargs)
        query_str = self.query_to_str(grid.build_query())

        if hasattr(look_for, 'search'):
            assert look_for.search(query_str), f'"{look_for.pattern}" not found in: {query_str}'
        else:
            assert re.search(look_for, query_str), f'"{look_for}" not found in: {query_str}'

    def get_grid(self, grid_args, *args, **kwargs):
        """Construct grid from args and kwargs, and apply grid_args.

        Args:
            grid_args: grid query args

        Returns:
            grid instance
        """
        grid = self.grid_cls(*args, **kwargs)
        grid.apply_qs_args(add_user_warnings=False, grid_args=grid_args)
        return grid

    def get_session_grid(self, *args, _query_string=None, **kwargs):
        """Construct grid from args and kwargs, and apply query string.

        Args:
            _query_string: URL query string with grid query args

        Returns:
            grid instance
        """
        grid = self.grid_cls(*args, **kwargs)
        if grid.manager.request():
            # request context already exists
            grid.apply_qs_args()
        else:
            url = f'/?{_query_string}' if _query_string else '/'
            with grid.manager.test_request_context(url=url):
                grid.apply_qs_args()
        return grid

    def get_pyq(self, grid=None, _query_string=None, **kwargs):
        """Turn provided/constructed grid into a rendered PyQuery object.

        Args:
            grid (BaseGrid, optional): Grid to use instead of `self.get_session_grid`.
            Defaults to None.

            kwargs (dict, optional): Additional args passed to `self.get_session_grid`.

        Returns:
            PyQuery object
        """
        session_grid = grid or self.get_session_grid(_query_string=_query_string, **kwargs)
        if session_grid.manager.request():
            # request context already exists
            html = session_grid.html()
        else:
            url = f'/?{_query_string}' if _query_string else '/'
            with session_grid.manager.test_request_context(url=url):
                html = session_grid.html()
        return PyQuery(f'<html>{html}</html>')

    def check_filter(self, name, op, value, expected):
        """Assertions to perform on a filter test.

        Args:
            name (str): Column key to filter.
            op (str): Filter operator to enable.
            value (Any): Filter value to assign.
            expected (str or regex): SQL string or compiled regex to find.
        """
        qs_args = [(f'op({name})', op)]
        if isinstance(value, (list, tuple)):
            for v in value:
                qs_args.append((f'v1({name})', v))
        else:
            qs_args.append((f'v1({name})', value))

        def sub_func(ex):
            query_string = urllib.parse.urlencode(qs_args)
            if isinstance(ex, re.compile('').__class__):
                self.assert_regex_in_query(ex, _query_string=query_string)
            else:
                self.assert_in_query(ex, _query_string=query_string)
            # ensures the query executes and the grid renders without error
            self.get_pyq(_query_string=query_string)

        def page_func():
            query_string = urllib.parse.urlencode([('onpage', 2), ('perpage', 1), *qs_args])
            pg = self.get_session_grid(_query_string=query_string)
            if pg.page_count > 1:
                self.get_pyq(_query_string=query_string)

        if self.grid_cls.pager_on:
            page_func()

        return sub_func(expected)

    def test_filters(self):
        """Use filters attribute/property/method to run assertions."""
        cases = self.filters() if callable(self.filters) else self.filters
        for name, op, value, expected in cases:
            self.check_filter(name, op, value, expected)

    def check_sort(self, k, ex, asc):
        """Assertions to perform on a sort test.

        Args:
            k (str): Column key to sort.
            ex (str or regex): SQL string to find.
            asc (bool): Flag indicating ascending/descending order.
        """
        if not asc:
            k = '-' + k
        d = {'sort1': k}

        def sub_func():
            query_string = urllib.parse.urlencode(d)
            self.assert_in_query(
                'ORDER BY {}{}'.format(ex, '' if asc else ' DESC'),
                _query_string=query_string,
            )
            # ensures the query executes and the grid renders without error
            self.get_pyq(_query_string=query_string)

        return sub_func()

    def test_sort(self):
        """Use sort_tests attribute/property to run assertions."""
        for col, expect in self.sort_tests:
            self.check_sort(col, expect, True)
            self.check_sort(col, expect, False)

    def _compare_table_block(self, block_selector, tag, expect):
        print(block_selector)
        assert len(block_selector) == len(expect)

        for row_idx, row in enumerate(expect):
            cells = block_selector.eq(row_idx).find(tag)
            assert len(cells) == len(row)
            for col_idx, val in enumerate(row):
                read = cells.eq(col_idx).text()
                assert read == val, f'row {row_idx} col {col_idx} {read} != {val}'

    def expect_table_header(self, expect, grid=None, _query_string=None, **kwargs):
        """Run assertions to compare rendered headings with expected data.

        Args:
            expect (list): List representation of expected table data.

            grid (BaseGrid, optional): Grid to use instead of `self.get_session_grid`.
            Defaults to None.

            kwargs (dict, optional): Additional args passed to `self.get_session_grid`.
        """
        d = self.get_pyq(grid, _query_string=_query_string, **kwargs)
        self._compare_table_block(
            d.find('table.records thead tr'),
            'th',
            expect,
        )

    def expect_table_contents(self, expect, grid=None, _query_string=None, **kwargs):
        """Run assertions to compare rendered data rows with expected data.

        Args:
            expect (list): List representation of expected table data.

            grid (BaseGrid, optional): Grid to use instead of `self.get_session_grid`.
            Defaults to None.

            kwargs (dict, optional): Additional args passed to `self.get_session_grid`.
        """
        d = self.get_pyq(grid, _query_string=_query_string, **kwargs)
        self._compare_table_block(
            d.find('table.records tbody tr'),
            'td',
            expect,
        )

    def test_search_expr_passes(self, grid=None, _query_string=None):
        """Assert that a single-search query executes without error."""
        grid = grid or self.get_session_grid(_query_string=_query_string)
        if grid.enable_search:
            grid.records  # noqa: B018


class MSSQLGridBase(GridBase):
    """MSSQL dialect produces some string oddities compared to other dialects, such as
    having the N'foo' syntax for unicode strings instead of 'foo'. This can clutter
    tests a bit. Using MSSQLGridBase will patch that into the asserts, so that
    look_for will match whether it has the N-prefix or not.
    """

    def query_to_str_replace_type(self, compiled_query):
        """Same as query_to_str, but accounts for pyodbc type-specific rendering."""
        query_str = self.query_to_str(compiled_query)
        # pyodbc rendering includes an additional character for some strings,
        # like N'foo' instead of 'foo'. This is not relevant to what we're testing.
        return re.sub(r"(\(|WHEN|LIKE|ELSE|THEN|[,=\+])( ?)N'(.*?)'", r"\1\2'\3'", query_str)

    def assert_in_query(self, look_for, grid=None, context=None, _query_string=None, **kwargs):
        session_grid = grid or self.get_session_grid(_query_string=_query_string, **kwargs)
        query_str = self.query_to_str(session_grid.build_query())
        query_str_repl = self.query_to_str_replace_type(session_grid.build_query())
        assert look_for in query_str or look_for in query_str_repl, (
            f'"{look_for}" not found in: {query_str}'
        )

    def assert_not_in_query(self, look_for, grid=None, context=None, _query_string=None, **kwargs):
        session_grid = grid or self.get_session_grid(_query_string=_query_string, **kwargs)
        query_str = self.query_to_str(session_grid.build_query())
        query_str_repl = self.query_to_str_replace_type(session_grid.build_query())
        assert look_for not in query_str or look_for not in query_str_repl, (
            f'"{look_for}" found in: {query_str}'
        )

    def assert_regex_in_query(
        self,
        look_for,
        grid=None,
        context=None,
        _query_string=None,
        **kwargs,
    ):
        session_grid = grid or self.get_session_grid(_query_string=_query_string, **kwargs)
        query_str = self.query_to_str(session_grid.build_query())
        query_str_repl = self.query_to_str_replace_type(session_grid.build_query())

        if hasattr(look_for, 'search'):
            assert look_for.search(query_str) or look_for.search(query_str_repl), (
                f'"{look_for.pattern}" not found in: {query_str}'
            )
        else:
            assert re.search(look_for, query_str) or re.search(look_for, query_str_repl), (
                f'"{look_for}" not found in: {query_str}'
            )
