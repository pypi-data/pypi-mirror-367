from abc import ABC, abstractmethod
from collections import defaultdict
import csv
from dataclasses import asdict
import inspect
import io
import json
from operator import itemgetter
import re
import typing

from blazeutils.containers import HTMLAttributes, LazyDict
from blazeutils.functional import identity
from blazeutils.helpers import tolist
from blazeutils.jsonh import jsonmod
from blazeutils.spreadsheets import WriterX, openpyxl, xlsxwriter
from blazeutils.strings import randnumerics, reindent
import jinja2 as jinja
from markupsafe import Markup
import six
from werkzeug.datastructures import MultiDict
from werkzeug.routing import Map, Rule

from . import extensions, types
from .extensions import CustomJsonEncoder, ngettext, translation_manager
from .extensions import gettext as _
from .utils import current_url


if openpyxl:
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter
else:
    Font = Border = Side = Alignment = typing.Any

try:
    from morphi.helpers.jinja import configure_jinja_environment
except ImportError:
    configure_jinja_environment = lambda *args, **kwargs: None

try:
    from speaklater import is_lazy_string
except ImportError:
    is_lazy_string = lambda value: False


def fix_xls_value(value):
    """
    Perform any data type fixes that must be made
    prior to sending a value to be written by the spreadsheet library
    """
    if is_lazy_string(value):
        return six.text_type(value)

    return value


class RenderLimitExceeded(Exception):
    pass


class Renderer(ABC):
    """Abstract interface for a WebGrid renderer.

    If the renderer has an `init` callable, it will be called by the constructor.

    Renderers are callable, which will trigger the `render` method::

        renderer = HTML(my_grid)
        output = renderer()

    Args:
        grid (BaseGrid): Parent grid of this renderer instance.
    """

    _columns = None

    @property
    @abstractmethod
    def name(self):
        """Identifier used to find columns that will render on this target."""

    @property
    def columns(self):
        """Cache a set of columns from the grid that will render on this target."""
        if not self._columns:
            self._columns = list(self.grid.iter_columns(self.name))
        return self._columns

    def __init__(self, grid):
        self.grid = grid
        if hasattr(self, 'init') and callable(self.init):
            self.init()

    def __call__(self):
        return self.render()

    def can_render(self):
        """Guard method for preventing a renderer from overflowing the target format.

        For instance, spreadsheets have limitation in the number of possible rows. A
        renderer to that format should check that the record count does not exceed
        that limit.

        Returns:
            bool: True if the renderer can proceed.
        """
        return True

    @abstractmethod
    def render(self):
        """Main renderer method returning the output."""


class GroupMixin:
    def has_groups(self):
        """Returns True if any of the renderer's columns is part of a column group."""
        return any(col.group for col in self.columns)

    def get_group_heading_colspans(self):
        """Computes the number of columns spanned by various groups.

        Note, this may not be the number of columns in the group in the grid definition,
        because some of those columns may not render in this target.
        """
        heading_colspans = []
        buffer_colspan = 0
        group_colspan = 0
        current_group = None
        for col in self.columns:
            if col.group:
                if buffer_colspan:
                    heading_colspans.append((None, buffer_colspan))
                    buffer_colspan = 0

                if current_group and current_group != col.group:
                    heading_colspans.append((current_group, group_colspan))
                    group_colspan = 1
                    current_group = None
                else:
                    current_group = col.group
                    group_colspan += 1
            else:
                buffer_colspan += 1
                if current_group:
                    heading_colspans.append((current_group, group_colspan))
                    group_colspan = 0
                    current_group = None

        if current_group:
            heading_colspans.append((current_group, group_colspan))

        return heading_colspans


def _safe_id(idstring):
    """
    From webhelpers2.html.tags. This is included for backwards compatibility
    TODO: Set IDs explicitly and don't rely on this being applied to name attributes
    """
    # Transform all whitespace to underscore
    idstring = re.sub(r'\s', '_', f'{idstring}')
    # Remove everything that is not a hyphen or a member of \w
    idstring = re.sub(r'(?!-)\W', '', idstring).lower()
    return idstring


def render_html_attributes(attrs):
    """Escapes attrs for HTML markup."""
    if not attrs:
        return Markup('')

    def render_attr(key, value):
        if value is True:
            return Markup.escape(key)
        elif value is False or value is None:
            return Markup('')
        return Markup(f'{Markup.escape(key)}="{Markup.escape(value)}"')

    attrs = sorted(attrs.items(), key=itemgetter(0))
    rendered_attrs = filter(identity, (render_attr(k, v) for k, v in attrs))
    return Markup(' ' + ' '.join(rendered_attrs))


class JSON(Renderer):
    """Renderer for JSON output"""

    mime_type = 'application/json'

    @property
    def name(self):
        return 'json'

    def serialized_columns(self):
        """Usually we would use the renderer's column list. For JSON, though, we want to
        supply any labels possible for use in a front-end app. The front-end needs to
        know names for filters, for example."""
        return {col.key: str(col.label) for col in self.grid.columns}

    def serialized_column_types(self):
        return {col.key: col.json_type_helper for col in self.columns}

    def serialize_column_group(self, label, columns):
        return types.ColumnGroup(
            label=label,
            columns=columns,
        )

    def serialized_column_groups(self):
        group_to_keys = defaultdict(list)
        for column in filter(lambda col: col.group is not None, self.columns):
            group_to_keys[column.group].append(column.key)
        return [
            self.serialize_column_group(group.label, columns)
            for group, columns in group_to_keys.items()
        ]

    def serialize_filter(self, filter):
        return types.Filter(
            op=filter.op,
            value1=filter.value1_set_with,
            value2=filter.value2_set_with,
        )

    def serialized_filters(self):
        return {
            col.key: self.serialize_filter(col.filter)
            for key, col in self.grid.filtered_cols.items()
            if col.filter.is_active
        }

    def serialized_filter_specs(self):
        return {
            col.key: col.filter.serialize_filter_spec()
            for key, col in self.grid.filtered_cols.items()
        }

    def serialize_record(self, record):
        return {col.key: col.render('json', record) for col in self.columns}

    def serialized_records(self):
        return [self.serialize_record(record) for record in self.grid.records]

    def serialize_totals_record(self, record):
        cols = filter(lambda col: col.key in self.grid.subtotal_cols, self.columns)
        return {col.key: col.render('json', record) for col in cols}

    def serialized_totals(self):
        serialize = lambda record: self.serialize_totals_record(record) if record else None
        return {
            'page': serialize(self.grid.page_totals),
            'grand': serialize(self.grid.grand_totals),
        }

    def serialize_sort(self, sort):
        key, flag_desc = sort
        return types.Sort(key=key, flag_desc=flag_desc)

    def serialized_order_by(self):
        return [self.serialize_sort(sort) for sort in self.grid.order_by]

    def serialized_sortable_columns(self):
        return [col.key for col in filter(lambda col: col.can_sort, self.grid.columns)]

    def asdict(self):
        grid = types.Grid(
            errors=[],
            settings=types.GridSettings(
                search_expr=self.grid.search_value,
                filters=self.serialized_filters(),
                paging=types.Paging(
                    pager_on=self.grid.pager_on,
                    per_page=self.grid.per_page,
                    on_page=self.grid.on_page,
                ),
                sort=self.serialized_order_by(),
                export_to=None,
            ),
            spec=types.GridSpec(
                columns=self.serialized_columns(),
                column_groups=self.serialized_column_groups(),
                column_types=self.serialized_column_types(),
                export_targets=list(self.grid.allowed_export_targets.keys()),
                enable_search=self.grid.enable_search,
                enable_sort=self.grid.sorter_on,
                sortable_columns=self.serialized_sortable_columns(),
                filters=self.serialized_filter_specs(),
            ),
            state=types.GridState(
                page_count=self.grid.page_count,
                record_count=self.grid.record_count,
                warnings=self.grid.user_warnings,
            ),
            records=self.serialized_records(),
            totals=self.serialized_totals(),
        )
        return asdict(grid)

    def render(self):
        return json.dumps(self.asdict(), cls=CustomJsonEncoder)

    def as_response(self):
        """Return a response via the grid's manager."""
        buffer = io.BytesIO()
        buffer.write(self.render())
        buffer.seek(0)
        return self.grid.manager.file_as_response(buffer, None, self.mime_type)


class HTML(GroupMixin, Renderer):
    """Renderer for HTML output."""

    NBSP = Markup('&nbsp;')

    @property
    def name(self):
        return 'html'

    def init(self):
        self.manager = self.grid.manager
        if self.manager:
            self.jinja_env = self.manager.jinja_environment
        else:
            # if the grid is unmanaged for any reason (e.g. just not in a request/response
            # cycle and used only for render), fall back to a default jinja environment
            self.jinja_env = jinja.Environment(
                loader=jinja.PackageLoader('webgrid', 'templates'),
                finalize=lambda x: x if x is not None else '',
                autoescape=True,
            )
        self.jinja_env.filters['wg_safe'] = jinja.filters.do_mark_safe
        self.jinja_env.filters['wg_attributes'] = render_html_attributes
        self.jinja_env.filters['wg_gettext'] = _
        self._template_cache = {}

        configure_jinja_environment(self.jinja_env, translation_manager)

    def _render_jinja(self, source, **kwargs):
        if source not in self._template_cache:
            self._template_cache[source] = self.jinja_env.from_string(source)
        template = self._template_cache[source]
        return Markup(template.render(**kwargs))

    def render(self):
        if not self.can_render():
            raise RenderLimitExceeded('Unable to render HTML table')
        return self.load_content('grid.html')

    def grid_attrs(self):
        """HTML attributes to render on the main grid div element."""
        return self.grid.hah

    def header(self):
        """Return content for the grid header area. Used by the grid template."""
        if self.grid.hide_controls_box:
            return ''
        return self.load_content('grid_header.html')

    def header_form_attrs(self, **kwargs):
        """HTML attributes to render on the grid header form element."""
        return {'method': self.form_action_method(), 'action': self.form_action_url(), **kwargs}

    def form_action_method(self):
        """Detect whether the header form should have a GET or POST action.

        By default, we look at the grid manager's args_loaders for RequestFormLoader. If it
        is present, the form will be POST.
        """
        if extensions.RequestFormLoader in self.grid.manager.args_loaders:
            return 'post'
        return 'get'

    def form_action_url(self):
        """URL target for the grid header form."""
        return self.reset_url(session_reset=False)

    def header_filtering(self):
        """Return content for the grid filter area. Used by the header template."""
        return self.load_content('header_filtering.html')

    def filtering_table_attrs(self, **kwargs):
        """HTML attributes to render on the grid filter table element."""
        kwargs.setdefault('cellpadding', 1)
        kwargs.setdefault('cellspacing', 0)
        return kwargs

    def filtering_session_key(self):
        """Hidden input to preserve the session key on form submission."""
        return self._render_jinja(
            '<input type="hidden" name="session_key" value="{{value}}" />',
            value=self.grid.session_key,
        )

    def filtering_fields(self):
        """Table rows for the filter area."""
        rows = []
        for col in six.itervalues(self.grid.filtered_cols):
            rows.append(self.filtering_table_row(col))
        rows = Markup('\n'.join(rows))

        top_row = self.get_search_row() if self.grid.can_search() else self.get_add_filter_row()

        return Markup('\n'.join([top_row, rows]))

    def filtering_table_row(self, col):
        """Single filter row with op and inputs."""
        extra = getattr(col.filter, 'html_extra', {})
        return self._render_jinja(
            """
            <tr class="{{col.key}}_filter" {{- extra|wg_attributes }}>
                <th class="filter-label">{{renderer.filtering_col_label(col)}}</th>
                <td class="operator">{{renderer.filtering_col_op_select(col)}}</td>
                <td>
                    <div class="inputs1">
                        {{ renderer.filtering_col_inputs1(col) }}
                    </div>
                    <div class="inputs2">
                        {{ renderer.filtering_col_inputs2(col) }}
                    </div>
                </td>
            </tr>
            """,
            renderer=self,
            col=col,
            extra=extra,
        )

    def filtering_col_label(self, col):
        """Label getter for filter column."""
        return col.label

    def filtering_col_op_select(self, col):
        """Render select box for filter Operator options."""
        filter = col.filter
        current_selected = filter.op if filter.is_display_active else ''

        primary_op = filter.primary_op or filter.operators[0]
        is_primary = lambda op: 'primary' if op == primary_op else None

        field_name = f'op({col.key})'
        field_name = self.grid.prefix_qs_arg_key(field_name)

        return self.render_select(
            [(op.key, op.display, is_primary(op)) for op in filter.operators],
            current_selected,
            name=field_name,
        )

    def filtering_col_inputs1(self, col):
        """Render the first input, which can be freeform or select."""
        filter = col.filter
        field_name = f'v1({col.key})'
        field_name = self.grid.prefix_qs_arg_key(field_name)

        inputs = Markup()

        if 'input' in filter.input_types:
            ident = f'{col.key}_input1'
            inputs += self._render_jinja(
                '<input{{attrs|wg_attributes}} />',
                attrs={
                    'name': field_name,
                    'value': filter.value1_set_with,
                    'id': ident,
                    'type': 'text',
                },
            )
        if 'select' in filter.input_types:
            current_selected = tolist(filter.value1) or []
            inputs += self.render_select(
                filter.options_seq,
                current_selection=current_selected,
                placeholder=None,
                multiple=filter.receives_list,
                name=field_name,
            )
            if filter.receives_list:
                inputs += self.filtering_multiselect(
                    field_name,
                    current_selected,
                    self.filtering_filter_options_multi(filter, field_name),
                )
        return inputs

    def filtering_multiselect(self, field_name, current_selected, options):
        """Almost all selects are rendered with multiselect UI. Render that here.

        Structure is based on the jQuery Multiselect plugin. For efficiency of render with
        large numbers of options, we customized the plugin for WebGrid use and offloaded the
        main render/transform here.
        """
        return self._render_jinja(
            """
            <div class="ms-parent">
                <button type="button" class="ms-choice">
                    <span class="placeholder"></span>
                    <div></div>
                </button>
                <div class="ms-drop bottom">
                    <div class="ms-search">
                        <input type="text"
                            autocomplete="off"
                            autocorrect="off"
                            autocapitalize="off"
                            spellcheck="false"
                        />
                    </div>
                    <ul>
                        <li>
                            <label>
                                <input type="checkbox" name="selectAll{{field_name}}" />
                                [{{ 'Select all'|wg_gettext }}]
                            </label>
                        </li>
                        {{options}}
                        <li class="ms-no-results">
                            {{'No matches found'|wg_gettext}}
                        </li>
                    </ul>
                </div>
            </div>
            """,
            field_name=field_name,
            current_selected=current_selected,
            options=options,
        )

    def filtering_filter_options_multi(self, filter, field_name):
        """Render the multiselect options."""
        # Assume the multiselect filter has been set up with OptionsFilterBase.setup_validator
        selected = filter.value1 or []
        validator = filter.value_modifier
        if inspect.isclass(validator):
            validator = validator()
        return self._render_jinja(
            """
            {% for value, label in filter.options_seq %}
                <li>
                    <label>
                        <input
                            {% if transform(value) in selected %}checked{% endif %}
                            type="checkbox"
                            value="{{value}}"
                            name="selectItem{{field_name}}"
                        />
                        {{label}}
                    </label>
                </li>
            {% endfor %}
            """,
            filter=filter,
            field_name=field_name,
            selected=selected,
            transform=validator.process if validator else lambda x: x,
        )

    def filtering_col_inputs2(self, col):
        """Render the second filter input, currently only a freeform."""
        filter = col.filter
        field_name = f'v2({col.key})'
        field_name = self.grid.prefix_qs_arg_key(field_name)

        if 'input2' not in filter.input_types:
            return Markup('')

        # field will get modified by JS
        ident = f'{col.key}_input2'
        return self._render_jinja(
            '<input{{attrs|wg_attributes}} />',
            attrs={
                'name': field_name,
                'value': filter.value2_set_with,
                'id': ident,
                'type': 'text',
            },
        )

    def filtering_add_filter_select(self):
        """Render the select box for adding a new filter. Used by the filter template."""
        return self.render_select(
            [(col.key, col.label) for col in self.grid.filtered_cols.values()],
            name='datagrid-add-filter',
        )

    def filtering_json_data(self):
        """Export certain filter data as a JSON object for use by the JS asset."""
        for_js = {}
        for col_key, col in six.iteritems(self.grid.filtered_cols):
            html_input_types = getattr(col.filter, 'html_input_types', {})
            for_js[col_key] = opdict = {}
            for op in col.filter.operators:
                opdict[op.key] = {
                    'field_type': op.field_type,
                    'hint': op.hint,
                }
                html_input_type = html_input_types.get(op, 'text')
                opdict[op.key]['html_input_type'] = html_input_type

        if self.grid.can_search():
            for_js['search'] = {'contains': {'field_type': None}}
        return jsonmod.dumps(for_js)

    def confirm_export(self):
        """Export confirmation data as a JSON object for use by the JS asset."""
        count = self.grid.record_count
        if self.grid.unconfirmed_export_limit is None:
            confirmation_required = False
        else:
            confirmation_required = count > self.grid.unconfirmed_export_limit
        return jsonmod.dumps({'confirm_export': confirmation_required, 'record_count': count})

    def header_sorting(self):
        """Render the sort area. Used by the header template."""
        return self.load_content('header_sorting.html')

    def render_select(
        self,
        options,
        current_selection=None,
        placeholder=('', NBSP),
        name=None,
        id=None,
        **kwargs,
    ):
        """Generalized select box renderer.

        Args:
            options (iterable): Option tuples (value, label) or (value, label, data).
            If the data piece is present, it will be rendered as the value
            of a "data-render" attribute on the option tag.

            current_selection (iterable, optional): Option values to be marked as selected.
            Defaults to None.

            placeholder (tuple(str), optional): Option to use as a "blank" value.

            name (str, optional): Value for HTML name attribute. Defaults to None.

            id (str, optional): Value for HTML id attribute. Defaults to a sanitized value
            derived from `name`.

            kwargs: Passed as HTML attributes on the select tag.

        """
        current_selection = tolist(current_selection) if current_selection is not None else []
        if placeholder:
            options = [placeholder, *options]

        if name is not None:
            kwargs['name'] = name
        if id is None and kwargs.get('name'):
            id = _safe_id(kwargs.get('name'))
        kwargs['id'] = id

        return self._render_jinja(
            """
            <select{{attrs|wg_attributes}}>
                {% for value, label, data in options %}
                    <option value="{{value}}"
                        {%- if data %} data-render="{{data}}" {%- endif -%}
                        {%- if value in current_selection %} selected {%- endif -%}
                    >
                        {{- label -}}
                    </option>
                {% endfor %}
            </select>
            """,
            options=((tuple(opt) if len(opt) == 3 else (tuple(opt) + (None,))) for opt in options),  # noqa: RUF005
            current_selection=current_selection,
            placeholder=placeholder,
            attrs=kwargs,
        )

    def sorting_select_options(self):
        """Generate list of tuple pairs (key, label) and (-key, label DESC) for sort options.

        Returns:
            list: List of tuple pairs.
        """
        options = []
        for col in self.grid.columns:
            if col.can_sort:
                options.extend(
                    [
                        (col.key, col.label),
                        (f'-{col.key}', _('{label} DESC', label=col.label)),
                    ],
                )
        return options

    def sorting_select(self, number):
        """Render the dropdown select of sorting options.

        Args:
            number (int): Priority of ordering option.

        Returns:
            str: Jinja-rendered string.
        """
        key = f'sort{number}'
        sort_qsk = self.grid.prefix_qs_arg_key(key)

        if len(self.grid.order_by) < number:
            currently_selected = ''
        else:
            currently_selected, flag_desc = self.grid.order_by[number - 1]
            if flag_desc:
                currently_selected = '-' + currently_selected

        return self.render_select(
            self.sorting_select_options(),
            currently_selected,
            name=sort_qsk,
            id=sort_qsk,
        )

    def sorting_select1(self):
        """Render the first sort select."""
        return self.sorting_select(1)

    def sorting_select2(self):
        """Render the second sort select."""
        return self.sorting_select(2)

    def sorting_select3(self):
        """Render the third sort select."""
        return self.sorting_select(3)

    def header_paging(self):
        """Render the paging area of the grid header."""
        return self.load_content('header_paging.html')

    def paging_select(self):
        """Render the page selection input."""
        op_qsk = self.grid.prefix_qs_arg_key('onpage')
        return self._render_jinja(
            """
            <span>
                <input name="{{name}}" id="{{name}}" type="number" value="{{page}}"
                    min="1" max="{{page_count}}" /> {{text}}
            </span>
            """,
            name=op_qsk,
            page_count=self.grid.page_count,
            page=self.grid.on_page,
            text=_('of {page_count}', page_count=self.grid.page_count),
        )

    def paging_input(self):
        """Render the per-page input."""
        pp_qsk = self.grid.prefix_qs_arg_key('perpage')
        return self._render_jinja(
            '<input type="number" min="1" name="{{name}}" value="{{value}}" />',
            name=pp_qsk,
            value=self.grid.per_page,
        )

    def paging_url_first(self):
        """Generate a URL for the first page of the grid."""
        return self.current_url(onpage=1, perpage=self.grid.per_page)

    def _page_image(self, url, width, height, alt):
        return self._render_jinja(
            '<img src="{{url}}" width="{{width}}" height="{{height}}" alt="{{alt}}" />',
            url=url,
            width=width,
            height=height,
            alt=alt,
        )

    def paging_img_first(self):
        """Render the footer icon for the first page of the grid."""
        img_url = self.manager.static_url('b_firstpage.png')
        return self._page_image(img_url, width=16, height=13, alt='<<')

    def paging_img_first_dead(self):
        """Render the footer disabled icon for the first page of the grid."""
        img_url = self.manager.static_url('bd_firstpage.png')
        return self._page_image(img_url, width=16, height=13, alt='<<')

    def paging_url_prev(self):
        """Generate a URL for the previous page of the grid."""
        prev_page = self.grid.on_page - 1
        return self.current_url(onpage=prev_page, perpage=self.grid.per_page)

    def paging_img_prev(self):
        """Render the footer icon for the previous page of the grid."""
        img_url = self.manager.static_url('b_prevpage.png')
        return self._page_image(img_url, width=8, height=13, alt='<')

    def paging_img_prev_dead(self):
        """Render the footer disabled icon for the previous page of the grid."""
        img_url = self.manager.static_url('bd_prevpage.png')
        return self._page_image(img_url, width=8, height=13, alt='<')

    def paging_url_next(self):
        """Generate a URL for the next page of the grid."""
        next_page = self.grid.on_page + 1
        return self.current_url(onpage=next_page, perpage=self.grid.per_page)

    def paging_img_next(self):
        """Render the footer icon for the next page of the grid."""
        img_url = self.manager.static_url('b_nextpage.png')
        return self._page_image(img_url, width=8, height=13, alt='>')

    def paging_img_next_dead(self):
        """Render the footer disabled icon for the next page of the grid."""
        img_url = self.manager.static_url('bd_nextpage.png')
        return self._page_image(img_url, width=8, height=13, alt='>')

    def paging_url_last(self):
        """Generate a URL for the last page of the grid."""
        return self.current_url(onpage=self.grid.page_count, perpage=self.grid.per_page)

    def paging_img_last(self):
        """Render the footer icon for the last page of the grid."""
        img_url = self.manager.static_url('b_lastpage.png')
        return self._page_image(img_url, width=16, height=13, alt='>>')

    def paging_img_last_dead(self):
        """Render the footer disabled icon for the last page of the grid."""
        img_url = self.manager.static_url('bd_lastpage.png')
        return self._page_image(img_url, width=16, height=13, alt='>>')

    def table(self):
        """Render the table area of the grid from template."""
        return self.load_content('grid_table.html')

    def no_records(self):
        """Render a message paragraph indicating the current filters
        return no records."""
        return self._render_jinja(
            '<p class="no-records">{{msg}}</p>',
            msg=_('No records to display'),
        )

    def table_attrs(self, **kwargs):
        """Apply default HTML table attributes to the supplied kwargs.

        Returns:
            dict: keys/values to be rendered as attributes
        """
        kwargs.setdefault('class', 'records')
        return kwargs

    def table_column_headings(self):
        """Combine all rendered column headings and return as Markup."""
        headings = []
        for col in self.columns:
            headings.append(self.table_th(col))
        th_str = '\n'.join(headings)
        th_str = reindent(th_str, 12)
        return Markup(th_str)

    def table_group_headings(self):
        """Combine all rendered column group headings and return as Markup."""
        group_headings = [
            self.group_th(group, colspan) for group, colspan in self.get_group_heading_colspans()
        ]
        th_str = '\n'.join(group_headings)
        th_str = reindent(th_str, 12)
        return Markup(th_str)

    def buffer_th(self, colspan, **kwargs):
        """Render a placeholder TH tag for spacing between column groups."""
        kwargs.setdefault('class', 'buffer')
        kwargs['colspan'] = colspan
        return self._render_jinja('<th{{ attrs|wg_attributes }}></th>', attrs=kwargs)

    def group_th(self, group, colspan, **kwargs):
        """Render a column group heading with the needed span.

        Note, will render an empty placeholder if the column has no group."""
        if group is None:
            return self.buffer_th(colspan)

        kwargs.setdefault('class', group.class_)
        kwargs['colspan'] = colspan
        return self._render_jinja(
            '<th {{- attrs|wg_attributes }}>{{label}}</th>',
            label=group.label,
            attrs=kwargs,
        )

    def table_th(self, col):
        """Render a single column heading TH tag.

        Sortable columns are rendered as links with the needed URL args."""
        label = col.label
        if self.grid.sorter_on and col.can_sort:
            url_args = {}
            url_args['dgreset'] = None
            url_args['sort2'] = None
            url_args['sort3'] = None
            link_attrs = {}
            if self.grid.order_by and len(self.grid.order_by) == 1:
                current_sort, flag_desc = self.grid.order_by[0]
                if current_sort == col.key:
                    link_attrs['class'] = 'sort-' + ('desc' if flag_desc else 'asc')
                if current_sort != col.key or flag_desc:
                    url_args['sort1'] = col.key
                else:
                    url_args['sort1'] = f'-{col.key}'
            else:
                url_args['sort1'] = col.key
            label = self._render_jinja(
                '<a href="{{href}}" {{- attrs|wg_attributes }}>{{label}}</a>',
                href=self.current_url(**url_args),
                attrs=link_attrs,
                label=label,
            )
        return self._render_jinja(
            '<th{{attrs|wg_attributes}}>{{label}}</th>',
            attrs=col.head.hah,
            label=label,
        )

    def table_rows(self):
        """Combine rows rendered from grid records, return as Markup.

        Page/Grand totals are included here as rows if enabled in the grid."""
        rows = []
        # loop through rows
        for rownum, record in enumerate(self.grid.records):
            rows.append(self.table_tr(rownum, record))
        # process subtotals (if any)
        if rows and self.grid.page_totals:
            rows.append(self.table_pagetotals(rownum + 1, self.grid.page_totals))
        if rows and self.grid.grand_totals:
            rows.append(self.table_grandtotals(rownum + 2, self.grid.grand_totals))
        rows_str = '\n        '.join(rows)
        return Markup(rows_str)

    def table_tr_styler(self, rownum, record):
        """Compile the styling to be used on a given HTML grid row.

        Applies odd/even class based on the row number. Adds in any row stylers
        present in grid configuration.

        Args:
            rownum (int): row number in the rendered grid.
            record (Any): result record.

        Returns:
            HTMLAttributes: attributes collection to be applied on a TR.
        """
        # handle row styling
        row_hah = HTMLAttributes()

        # add odd/even classes to the rows
        if (rownum + 1) % 2 == 1:
            row_hah.class_ += 'odd'
        else:
            row_hah.class_ += 'even'

        for styler in self.grid._rowstylers:
            styler(self.grid, rownum, row_hah, record)

        return row_hah

    def table_tr_output(self, cells, row_hah):
        """Combine rendered cells and output a TR tag."""
        # do some formatting so that the source code is properly indented
        tds_str = '\n'.join(cells)
        tds_str = reindent(tds_str, 12)
        tds_str = f'\n{tds_str}\n        '

        return self._render_jinja(
            '<tr{{attrs|wg_attributes}}>{{tds}}</tr>',
            attrs=row_hah,
            tds=Markup(tds_str),
        )

    def table_tr(self, rownum, record):
        """Generate rendered cells and pass to table_tr_output for rendered result."""
        row_hah = self.table_tr_styler(rownum, record)

        # get the <td>s for this row
        cells = []
        for col in self.columns:
            cells.append(self.table_td(col, record))

        return self.table_tr_output(cells, row_hah)

    def table_totals(self, rownum, record, label, numrecords):
        """Render a totals row based on subtotal columns defined in the grid."""
        row_hah = self.table_tr_styler(rownum, record)
        row_hah.class_ += 'totals'

        # get the <td>s for this row
        cells = []
        colspan = 0
        firstcol = True
        for col in self.columns:
            if col.key not in list(self.grid.subtotal_cols.keys()):
                if firstcol:
                    colspan += 1
                else:
                    cells.append(Markup('<td>&nbsp;</td>'))
                continue
            if firstcol:
                bufferval = ngettext(
                    '{label} ({num} record):',
                    '{label} ({num} records):',
                    numrecords,
                    label=label,
                )
                buffer_hah = {'colspan': colspan, 'class': 'totals-label'}
                if colspan:
                    cells.append(
                        self._render_jinja(
                            '<td{{attrs|wg_attributes}}>{{val}}</td>',
                            attrs=buffer_hah,
                            val=bufferval,
                        ),
                    )
                firstcol = False
                colspan = 0
            cells.append(self.table_td(col, record))

        return self.table_tr_output(cells, row_hah)

    def table_pagetotals(self, rownum, record):
        """Render a Page totals row based on subtotal columns defined in the grid."""
        return self.table_totals(rownum, record, _('Page Totals'), rownum)

    def table_grandtotals(self, rownum, record):
        """Render a Grand totals row based on subtotal columns defined in the grid."""
        count = self.grid.record_count
        return self.table_totals(rownum, record, _('Grand Totals'), count)

    def table_td(self, col, record):
        """Render a table data cell.

        Value is obtained for render from the grid column's `render` method. To
        override how a column's data is rendered specifically for HTML, supply a
        `render_html` method on the column."""
        col_hah = HTMLAttributes(col.body.hah)

        # allow column stylers to set attributes
        for styler, cname in self.grid._colstylers:
            for_column = self.grid.column(cname)
            if col.key == for_column.key:
                styler(self.grid, col_hah, record)

        # extract the value from the record for this column and prep
        col_value = col.render('html', record, col_hah)

        # turn empty values into a non-breaking space so table cells don't
        # collapse
        if col_value is None:  # noqa: SIM114
            styled_value = self.NBSP
        elif isinstance(col_value, six.string_types) and col_value.strip() == '':
            styled_value = self.NBSP
        else:
            styled_value = col_value

        return self._render_jinja(
            '<td{{attrs|wg_attributes}}>{{value}}</td>',
            attrs=col_hah,
            value=styled_value,
        )

    def footer(self):
        """Render the grid footer area from template."""
        return self.load_content('grid_footer.html')

    def load_content(self, endpoint, **kwargs):
        """Load content via Jinja templates.

        Gives the grid manager a chance to render the template, in order to allow for
        application-level overrides on the grid templates. Otherwise, defaults to the
        internal Jinja environment set up for this renderer."""
        kwargs['renderer'] = self
        kwargs['grid'] = self.grid

        try:
            # give the adapter a chance to render
            if hasattr(self.grid.manager, 'render_template'):
                return self.grid.manager.render_template(endpoint, **kwargs)
        except jinja.exceptions.TemplateNotFound:
            # fail silently, will fail on the next step if there's really a problem
            pass

        # if the adapter doesn't want it, default to raw Jinja2
        template = self.jinja_env.get_template(endpoint)
        return template.render(**kwargs)

    def current_url(self, **kwargs):
        """Generate a URL from current request args and the given kwargs."""
        curl = current_url(self.grid.manager, strip_querystring=True, strip_host=True)

        map_adapter = Map(
            [
                Rule(curl, endpoint='magic'),
            ],
            sort_parameters=True,
        ).bind(curl)

        req_args = MultiDict(self.grid.manager.request_url_args())

        # reset key should not be retained from request for url generation. If we really
        # want it, we'll see it in kwargs.
        reset_key = (self.grid.qs_prefix or '') + 'dgreset'
        req_args.pop(reset_key, None)

        # kwargs will be modified with new keys if there is a prefix, so copy the original set
        # of keys first. Otherwise, the loop may pick up new keys and apply the prefix again
        key_list = list(kwargs.keys())
        for key in key_list:
            # arg keys may need to be prefixed
            if self.grid.qs_prefix:
                prefixed_key = self.grid.qs_prefix + key
                kwargs[prefixed_key] = kwargs[key]
                del kwargs[key]
                key = prefixed_key

            # multidicts extend, not replace, so we need to get rid of the
            # keys first
            if key in req_args:
                del req_args[key]

        # convert to md first so that if we have lists in the kwargs, they
        # are converted appropriately
        req_args.update(MultiDict(kwargs))
        return map_adapter.build('magic', req_args)

    def reset_url(self, session_reset=True):
        """Generate a URL that will trigger a reset of the grid's UI options."""
        url_args = {}
        url_args['perpage'] = None
        url_args['onpage'] = None
        url_args['search'] = None
        url_args['sort1'] = None
        url_args['sort2'] = None
        url_args['sort3'] = None
        url_args['export_to'] = None
        url_args['datagrid-add-filter'] = None

        for col in six.itervalues(self.grid.filtered_cols):
            url_args[f'op({col.key})'] = None
            url_args[f'v1({col.key})'] = None
            url_args[f'v2({col.key})'] = None

        url_args['session_key'] = self.grid.session_key
        url_args['session_override'] = None
        if session_reset:
            url_args['dgreset'] = 1

        return self.current_url(**url_args)

    def export_url(self, renderer):
        """Generate a URL that will trigger an export to one of the grid's renderers.

        Args:
            renderer (str): Export key (e.g. xlsx, csv) for rendering target."""
        return self.current_url(export_to=renderer)

    def _get_filter_select_info(self):
        return _('Add Filter:'), self.filtering_add_filter_select()

    def get_search_row(self):
        """Render the single-search input, along with filter select."""
        filter_label, filter_select = self._get_filter_select_info()
        return self._render_jinja(
            """
            <tr class="search">
                <th>{{label}}</th>
                <td colspan="2">
                    <input name="search" type="text" value="{{search_value}}" id="search_input" />
                    <div class="add-filter">
                        <label>{{filter_label}}</label>
                        {{filter_select}}
                    </div>
                </td>
            </tr>
            """,
            label=_('Search'),
            search_value=self.grid.search_value,
            filter_label=filter_label,
            filter_select=filter_select,
        )

    def get_add_filter_row(self):
        """Render just the Add Filter area on a row."""
        filter_label, filter_select = self._get_filter_select_info()
        return self._render_jinja(
            """
            <tr class="add-filter">
                <th>
                    {{ filter_label }}
                </th>
                <td colspan="3">
                    {{ filter_select }}
                </td>
            </tr>
            """,
            filter_label=filter_label,
            filter_select=filter_select,
        )


class XLSXWriterWorkbookManager:
    def __init__(self, *args, **kwargs):
        self.base_style_attrs = {
            'bold': True,
            'top': 6,  # Double thick border
        }
        self._workbook = self.create_workbook()
        self._xlsx_format_cache = {}
        self.heading_style = self._workbook.add_format({'bold': True})
        self.totals_style = self._workbook.add_format(self.base_style_attrs)

        self.styles_cache = LazyDict()
        self.default_style = {}
        super().__init__(*args, **kwargs)

    def create_workbook(self):
        buf = io.BytesIO()
        return xlsxwriter.Workbook(buf, options={'in_memory': True})

    def get_xlsx_format(self, style_dict):
        """
        This method is meant to solve a major performance issue with how xlsxwriter manages formats.
        Xlsxwriter maintains a cache of formats, however generating the cache key is surprisingly
        expensive since it must join together every property of the format.

        The upshot of this is that if we have several columns with identical style properties but
        separate xlsxwriter Format objects, the cache key will have to be generated multiple times
        per cell. It is much faster to use the same Format object for all columns sharing the same
        style properties.

        See xlsxwriter.format::Format._get_xf_index for how the caching works.
        """
        key = tuple(sorted(style_dict.items(), key=itemgetter(0)))
        if key not in self._xlsx_format_cache:
            self._xlsx_format_cache[key] = self._workbook.add_format(style_dict)
        return self._xlsx_format_cache[key]

    def merged_totals_cell(self, xlh, value, style, colspan):
        xlh.ws.merge_range(
            xlh.rownum,
            xlh.colnum,
            xlh.rownum,
            xlh.colnum + colspan - 1,
            value,
            style,
        )

    def merged_heading_cell(self, xlh, value, style, col_index, colspan):
        xlh.ws.merge_range(0, col_index, 0, col_index + (colspan - 1), value, style)

    def style_for_column(self, col):
        if col.key not in self.styles_cache:
            style_dict = getattr(col, 'xlsx_style', self.default_style).copy()
            if col.xls_num_format:
                style_dict['num_format'] = col.xls_num_format
            self.styles_cache[col.key] = self.get_xlsx_format(style_dict)
        return self.styles_cache[col.key]

    def set_column_widths(self, writer, idx, width):
        writer.ws.set_column(idx, idx, width)

    def add_worksheet(self, sheet_name):
        return self._workbook.add_worksheet(sheet_name)

    @property
    def fileclosed(self):
        return self._workbook.fileclosed

    def close(self):
        self._workbook.close()

    @property
    def filename(self):
        return self._workbook.filename

    def write_group_header(self, xlh, col_index, data):
        xlh.ws.write(0, col_index, data, self.heading_style)

    def write_totals_row(self, xlh, col, value):
        style = self.base_style_attrs.copy()
        default_style = {}
        style.update(getattr(col, 'xlsx_style', default_style))
        style = self._workbook.add_format(style)
        xlh.awrite(fix_xls_value(value), style)


class OpenpyxlWorkbookManager:
    def __init__(self, *args, **kwargs):
        self._workbook = openpyxl.Workbook()
        self._file = None
        self._named_styles = set()
        self.totals_style_spec = {
            'font': Font(bold=True),
            'border': Border(top=Side(border_style='double')),
        }

        self.heading_style = self.get_heading_style()
        self.totals_style = self.get_totals_style()

    def add_named_style(self, style: dict[str, Font | Border | Alignment], name: str) -> str:
        """
        Registers a named style based on the column key, and caches to avoid name conflicts later.
        https://openpyxl.readthedocs.io/en/stable/styles.html#creating-a-named-style

        Returns: string representing the name of the named style
        """

        if name in self._named_styles:
            return name

        named_style = openpyxl.styles.NamedStyle(name=name)
        for key, val in style.items():
            setattr(named_style, key, val)
        self._workbook.add_named_style(named_style)
        self._named_styles.add(name)
        return named_style.name

    def get_heading_style(self):
        return self.add_named_style({'font': Font(bold=True)}, 'heading')

    def get_totals_style(self):
        return self.add_named_style(self.totals_style_spec, 'totals')

    @property
    def filename(self):
        """
        A misnomer but is consistent with xlsxwriter. This actually returns the file handle,
        in this case a BytesIO
        """
        if not self._file:
            buf = io.BytesIO()
            self._workbook.save(buf)
            self._file = buf
        return self._file

    def close(self):
        self._workbook.close()

    @property
    def fileclosed(self):
        """
        Property does not exist in openpyxl like it does in xlsxwriter,
        and trying to close an already closed book does not seem to cause issues.
        """
        return False

    def merged_totals_cell(self, xlh, value, style, colspan):
        """
        openpyxl-specific creation of a merged cell for the totals row
        """
        merged_cells = xlh.ws.cell(row=xlh.rownum, column=xlh.colnum)
        merged_cells.value = value
        merged_cells.style = style
        xlh.ws.merge_cells(
            start_row=xlh.rownum,
            start_column=xlh.colnum,
            end_row=xlh.rownum,
            end_column=xlh.colnum + colspan - 1,
        )

    def merged_heading_cell(self, xlh, value, style, col_index, colspan):
        """
        openpyxl-specific creation of a merged cell for group headers
        """
        merged_cells = xlh.ws.cell(row=1, column=col_index + 1)

        # The editable cell within a merged cell (top left cell) is a Cell,
        # while the other cells in that merge will yield a MergedCell, which
        # has an immutable value
        if isinstance(merged_cells, openpyxl.cell.Cell):
            merged_cells.value = value

        merged_cells.style = style
        xlh.ws.merge_cells(
            start_row=1,
            start_column=col_index + 1,
            end_row=1,
            end_column=col_index + colspan,
        )

    def style_for_column(self, col):
        """Return a column's style from cached information."""
        style_dict = getattr(col, 'xlsx_style', None)
        if style_dict:
            if 'num_format' in style_dict:
                # translate the key for openpyxl
                style_dict['number_format'] = style_dict['num_format']

            # Only use certain keys, as they become style attributes
            style_dict = {
                key: style_dict[key]
                for key in ['font', 'alignment', 'border', 'number_format']
                if key in style_dict
            }
            return self.add_named_style(style_dict, col.key)

    def set_column_widths(self, writer, idx, width):
        col_ltr = get_column_letter(idx + 1)
        writer.ws.column_dimensions[col_ltr].width = width

    def add_worksheet(self, sheet_name):
        # Workbook is created with blank worksheet, just rename and return
        sheet = self._workbook.worksheets[0]
        sheet.title = sheet_name
        return sheet

    def write_group_header(self, xlh, col_index, data):
        group_hdr_cell = xlh.ws.cell(row=1, column=col_index + 1)
        group_hdr_cell.style = self.heading_style

    def write_totals_row(self, xlh, col, value):
        style = self.totals_style_spec.copy()
        xlsx_style = getattr(col, 'xlsx_style', None)
        if xlsx_style:
            if 'num_format' in xlsx_style:
                xlsx_style['number_format'] = xlsx_style['num_format']
            style.update(xlsx_style)

        named_style = self.add_named_style(style, f'{col.key}_totals')
        xlh.awrite(fix_xls_value(value), named_style)


class XLSX(GroupMixin, Renderer):
    """Renderer for Excel XLSX output."""

    mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def __init__(self, grid):
        self._manager_cls = None
        super().__init__(grid)

    def __call__(self, manager_cls=None):
        self._manager_cls = manager_cls
        return super().__call__()

    @property
    def name(self):
        return 'xlsx'

    def init(self):
        self.col_widths = {}

    def create_workbook(self):
        if self._manager_cls:
            return self._manager_cls()
        if openpyxl:
            return OpenpyxlWorkbookManager()
        if xlsxwriter:
            return XLSXWriterWorkbookManager()
        raise ImportError('You must have either openpyxl or xlsxwriter installed.')

    def update_column_width(self, col, data):
        """Compute and store a column width from length of current data."""
        width = max((col.xls_width_calc(data), self.col_widths.get(col.key, 0)))
        self.col_widths[col.key] = width

    def adjust_column_widths(self, writer, wb):
        """Apply stored column widths to the XLSX worksheet."""
        for idx, col in enumerate(self.columns):
            if col.key in self.col_widths:
                wb.set_column_widths(writer, idx, self.col_widths[col.key])

    def build_sheet(self, wb=None, sheet_name=None):
        """Create and populate a worksheet for the current grid.

        Args:
            wb (Workbook, optional): xlsxwriter Workbook. Defaults to None (create one).
            sheet_name (str, optional): Sheet name. Defaults to None (use grid identity).

        Raises:
            ImportError: No suitable XLSX library installed.
            RenderLimitExceeded: Too many records to render to the target.

        Returns:
            Workbook: Created/supplied workbook with the rendered worksheet added.
        """

        if not self.can_render():
            raise RenderLimitExceeded('Unable to render XLSX sheet')

        if wb is None:
            wb = self.create_workbook()

        sheet = wb.add_worksheet(self.sanitize_sheet_name(sheet_name or self.grid.ident))
        writer = WriterX(sheet)

        self.sheet_header(writer, wb)
        self.sheet_body(writer, wb)
        self.sheet_footer(writer, wb)
        self.adjust_column_widths(writer, wb)
        return wb

    def render(self):
        return self.build_sheet()

    def can_render(self):
        total_rows = self.grid.record_count + 1
        if self.grid.subtotals != 'none':
            total_rows += 1
        return total_rows <= 1048576 and sum(1 for _ in self.columns) <= 16384

    def sanitize_sheet_name(self, sheet_name):
        """Work around Excel limitations on names of worksheets."""
        return sheet_name if len(sheet_name) <= 30 else (sheet_name[:27] + '...')

    def sheet_header(self, xlh, wb):
        """Placeholder method for app-specific sheet header rendering.

        Args:
            xlh (WriterX): Helper for writing worksheet cells.
            wb (Workbook): xlsxwriter Workbook object for direct usage.
        """

    def sheet_body(self, xlh, wb):
        """Render the headings/records area of the worksheet.

        Args:
            xlh (WriterX): Helper for writing worksheet cells.
            wb (Workbook): xlsxwriter Workbook object for direct usage.
        """
        self.body_headings(xlh, wb)
        self.body_records(xlh, wb)

    def sheet_footer(self, xlh, wb):
        """Placeholder method for app-specific sheet footer rendering.

        Args:
            xlh (WriterX): Helper for writing worksheet cells.
            wb (Workbook): xlsxwriter Workbook object for direct usage.
        """

    def body_headings(self, xlh, wb):
        """Render group and column label rows.

        Args:
            xlh (WriterX): Helper for writing worksheet cells.
            wb (Workbook): xlsxwriter Workbook object for direct usage.
        """

        # Render group labels above column headings.
        if self.has_groups():
            col_index = 0
            for group, colspan in self.get_group_heading_colspans():
                data = fix_xls_value(group.label) if group else None
                if colspan == 1:
                    xlh.awrite(data, wb.heading_style)
                    wb.write_group_header(xlh, col_index, data)

                else:
                    wb.merged_heading_cell(xlh, data, wb.heading_style, col_index, colspan)

                col_index += colspan

            xlh.nextrow()

        for col in self.columns:
            xlh.awrite(fix_xls_value(col.label), wb.heading_style)
            self.update_column_width(col, col.label)
        xlh.nextrow()

    def body_records(self, xlh, wb):
        """Render records and totals rows.

        Args:
            xlh (WriterX): Helper for writing worksheet cells.
            wb (Workbook): xlsxwriter Workbook object for direct usage.
        """
        # turn off paging
        self.grid.set_paging(None, None)

        rownum = 0
        for rownum, record in enumerate(self.grid.records):
            self.record_row(xlh, rownum, record, wb)

        # totals
        if rownum and self.grid.subtotals != 'none' and self.grid.subtotal_cols:
            self.totals_row(xlh, rownum + 1, self.grid.grand_totals, wb)

    def record_row(self, xlh, rownum, record, wb):
        """Render a single row from data record.

        Value is obtained for render from the grid column's `render` method. To
        override how a column's data is rendered specifically for XLSX, supply a
        `render_xlsx` method on the column.

        Args:
            xlh (WriterX): Helper for writing worksheet cells.
            rownum (int): Not used by default, but helpful for style overrides.
            record (Any): Object containing row data.
            wb (Workbook): xlsxwriter Workbook object for direct usage.
        """
        for col in self.columns:
            value = col.render('xlsx', record)
            style = wb.style_for_column(col)
            xlh.awrite(fix_xls_value(value), style)
            self.update_column_width(col, value)
        xlh.nextrow()

    def totals_row(self, xlh, rownum, record, wb):
        """Render a totals row based on subtotal columns defined in the grid."""
        colspan = 0
        firstcol = True

        for col in self.columns:
            if col.key not in list(self.grid.subtotal_cols.keys()):
                if firstcol:
                    colspan += 1
                else:
                    xlh.awrite('', wb.totals_style)
                continue
            if firstcol:
                numrecords = self.grid.record_count
                bufferval = 'Totals ({} record{}):'.format(
                    numrecords,
                    's' if numrecords != 1 else '',
                )
                if colspan > 1:
                    wb.merged_totals_cell(xlh, bufferval, wb.totals_style, colspan)
                    xlh.colnum = xlh.colnum + colspan
                else:
                    xlh.awrite(bufferval, wb.totals_style)

                firstcol = False
                colspan = 0

            value = col.render('xlsx', record)
            wb.write_totals_row(xlh, col, value)
            self.update_column_width(col, value)

        xlh.nextrow()

    def file_name(self):
        """Return an output filename based on grid identifier.

        A random numeric suffix is added. This is due to Excel's limitation to having
        only one workbook open with a given name. Excel will not allow a second file
        with the same name to open, even if the files are in different paths.
        """
        return f'{self.grid.ident}_{randnumerics(6)}.xlsx'

    def as_response(self, wb=None, sheet_name=None):
        """Return an attachment file via the grid's manager."""
        wb = self.build_sheet(wb, sheet_name)
        if not wb.fileclosed:
            wb.close()
        wb.filename.seek(0)
        return self.grid.manager.file_as_response(wb.filename, self.file_name(), self.mime_type)


class CSV(Renderer):
    """Renderer for CSV output."""

    mime_type = 'text/csv'

    @property
    def name(self):
        return 'csv'

    def render(self):
        self.output = six.StringIO()
        self.writer = csv.writer(self.output, delimiter=',', quotechar='"')
        self.body_headings()
        self.body_records()

    def file_name(self):
        """Return an output filename based on grid identifier.

        A random numeric suffix is added. This is due to Excel's limitation to having
        only one workbook open with a given name. Excel will not allow a second file
        with the same name to open, even if the files are in different paths.
        """
        return f'{self.grid.ident}_{randnumerics(6)}.csv'

    def build_csv(self):
        """Render grid output as CSV and return the contents in an IO stream."""
        self.render()
        byte_data = six.BytesIO()
        byte_data.write(self.output.getvalue().encode('utf-8'))
        return byte_data

    def body_headings(self):
        """Render the column headers.

        Note, column groups do not have real meaning in the CSV context, so
        they are left out here."""
        headings = []
        for col in self.columns:
            headings.append(col.label)
        self.writer.writerow(headings)

    def body_records(self):
        """Render all rows from grid records.

        Value is obtained for render from each grid column's `render` method. To
        override how a column's data is rendered specifically for CSV, supply a
        `render_csv` method on the column."""
        # turn off paging
        self.grid.set_paging(None, None)

        for record in self.grid.records:
            row = []
            for col in self.columns:
                row.append(col.render('csv', record))
            self.writer.writerow(row)

    def as_response(self):
        """Return an attachment file via the grid's manager."""
        buffer = self.build_csv()
        buffer.seek(0)
        return self.grid.manager.file_as_response(buffer, self.file_name(), self.mime_type)
