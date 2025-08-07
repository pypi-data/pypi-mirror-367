import csv
import datetime as dt
from enum import Enum
import io
import json
from typing import ClassVar

import arrow
from markupsafe import Markup
import openpyxl
from pyquery import PyQuery
import pytest
from werkzeug.datastructures import MultiDict
import xlsxwriter

from webgrid import (
    BoolColumn,
    Column,
    ColumnGroup,
    DateTimeColumn,
    LinkColumnBase,
    NumericColumn,
    YesNoColumn,
    col_filter,
    col_styler,
    extensions,
    row_styler,
)
from webgrid.filters import DateFilter, OptionsEnumFilter, TextFilter
from webgrid.renderers import (
    CSV,
    HTML,
    JSON,
    XLSX,
    OpenpyxlWorkbookManager,
    RenderLimitExceeded,
    XLSXWriterWorkbookManager,
    render_html_attributes,
)
from webgrid_ta.grids import (
    ArrowCSVGrid,
    ArrowGrid,
    Grid,
    StopwatchGrid,
)
from webgrid_ta.grids import (
    PeopleGrid as PG,
)
from webgrid_ta.model.entities import (
    AccountType,
    ArrowRecord,
    Email,
    Person,
    Status,
    db,
)

from .helpers import _inrequest, eq_html, render_in_grid


def _query_exclude_person(query):
    # this is pretty limited, but only used in the below couple of grids to
    # exclude the third Person record
    persons = Person.query.order_by(Person.id).limit(3).all()
    exclude_id = persons[2].id if len(persons) >= 3 else -1
    return query.filter(Person.id != exclude_id)


class PeopleGrid(PG):
    def query_prep(self, query, has_sort, has_filters):
        query = PG.query_prep(self, query, True, True)

        # default sort
        if not has_sort:
            query = query.order_by(Person.id.desc())

        # default filter
        if not has_filters:
            query = _query_exclude_person(query)

        return query


class PeopleCSVGrid(PG):
    allowed_export_targets: ClassVar = {'csv': CSV}

    def query_prep(self, query, has_sort, has_filters):
        query = PG.query_prep(self, query, True, True)

        # default sort
        if not has_sort:
            query = query.order_by(Person.id.desc())

        # default filter
        if not has_filters:
            query = _query_exclude_person(query)

        return query


def setup_module():
    Status.delete_cascaded()
    sp = Status(label='pending')
    sip = Status(label='in process')
    sc = Status(label='complete', flag_closed=1)
    db.session.add_all([sp, sip, sc])

    for x in range(1, 5):
        p = Person()
        p.firstname = f'fn{x:03d}'
        p.lastname = f'ln{x:03d}'
        p.sortorder = x
        p.numericcol = '2.13'
        p.state = f'st{x:03d}'
        if x != 2:
            p.createdts = dt.datetime(2012, 0o2, 22, 10, x, 16)
            p.due_date = dt.date(2012, 0o2, x)
        db.session.add(p)
        p.emails.append(Email(email=f'email{x:03d}@example.com'))
        p.emails.append(Email(email=f'email{x:03d}@gmail.com'))
        if x % 4 == 1:
            p.status = sip
            p.account_type = AccountType.admin
        elif x % 4 == 2:
            p.status = sp
            p.account_type = AccountType.employee
        elif x % 4 == 0:
            p.status = None

    db.session.commit()


class SimpleGrid(Grid):
    on_page = 1
    per_page = 1

    Column('ID', 'id')
    Column('Name', 'name', filter=TextFilter(Person.firstname))
    Column('Status', 'status')
    Column('Emails', 'emails', can_sort=False)


class ColorColumn(Column):
    def format_data(self, data):
        if data == 'blue':
            return 'blue :)'
        return data


class EditColumn(LinkColumnBase):
    link_attrs: ClassVar = {'target': '_blank'}

    def create_url(self, record):
        return '/vehicle-edit/{}'.format(record['id'])


class DealerColumn(LinkColumnBase):
    def create_url(self, record):
        return '/dealer-edit/{}'.format(record['dealer_id'])

    def extract_data(self, record):
        return record['dealer'] + record['dealer_id']


class CarGrid(Grid):
    EditColumn('ID', 'id')
    EditColumn('Edit', 'edit', link_label='edit')
    DealerColumn('Dealer', 'dealer')
    Column('Make', 'make')
    Column('Model', 'model', class_='model')
    ColorColumn('Color', 'color')
    BoolColumn('Active', 'active')
    BoolColumn('Active Reverse', 'active', reverse=True)
    YesNoColumn('Active Yes/No', 'active')

    @row_styler
    def style_row(self, rownum, attrs, record):
        attrs.id = 'tr_{}'.format(record['id'])

    @col_styler('model')
    def highlight_1500(self, attrs, record):
        if record['model'] == '1500':
            attrs.class_ += 'red'

    @col_filter('color')
    def pink_is_ugly(self, value):
        if value == 'pink':
            return 'pink :('
        return value


def find_tag(html, tag, id_=None, class_=None, **attrs):
    selector = tag
    if id_:
        selector += f'#{id_}'
    if class_:
        selector += f'.{class_}'
    for k, v in attrs.items():
        if v is None:
            selector += f'[{k}]'
        else:
            selector += f'[{k}="{v}"]'

    return PyQuery(html)(selector)


def assert_tag(html, tag, text=None, **kwargs):
    results = find_tag(html, tag, **kwargs)
    assert results

    if text is not None:
        assert any(i.text(squash_space=False) == text for i in results.items()), (
            f'{text} not found in {results}'
        )
    return results


class TestHtmlRenderer:
    key_data = (
        {'id': 1, 'name': 'one', 'status': 'new', 'emails': ''},
        {'id': 2, 'name': 'two', 'status': 'new', 'emails': ''},
        {'id': 3, 'name': 'three', 'status': 'pending', 'emails': ''},
        {'id': 4, 'name': 'three', 'status': 'pending', 'emails': ''},
        {'id': 5, 'name': 'three', 'status': 'complete', 'emails': ''},
    )

    @_inrequest('/')
    def test_car_html(self):
        key_data = (
            {
                'id': 1,
                'make': 'ford',
                'model': 'F150&',
                'color': 'pink',
                'dealer': 'bob',
                'dealer_id': '7',
                'active': True,
                'active_1': True,
                'active_2': True,
            },
            {
                'id': 2,
                'make': 'chevy',
                'model': '1500',
                'color': 'blue',
                'dealer': 'fred',
                'dealer_id': '9',
                'active': False,
                'active_1': False,
                'active_2': False,
            },
        )

        mg = CarGrid()
        mg.set_records(key_data)
        eq_html(mg.html.table(), 'basic_table.html')

    @pytest.mark.skipif(db.engine.dialect.name != 'sqlite', reason='IDs will not line up')
    @_inrequest('/')
    def test_people_html(self):
        pg = render_in_grid(PeopleGrid, 'html')()
        eq_html(pg.html.table(), 'people_table.html')

    @pytest.mark.skipif(db.engine.dialect.name != 'sqlite', reason='IDs will not line up')
    @_inrequest('/')
    def test_stopwatch_html(self):
        # Test Stopwatch grid with column groups.
        grid = StopwatchGrid()
        eq_html(grid.html.table(), 'stopwatch_table.html')

    @_inrequest('/')
    def test_default_jinja_env(self):
        class TGrid(Grid):
            manager = None
            hide_controls_box = True
            session_on = False
            allowed_export_targets = None

            Column('ID', 'id', can_sort=False)
            Column('Value', 'value', can_sort=False)

        tg = TGrid()
        tg.set_records(
            [
                {'id': 1, 'value': 'foo'},
            ],
        )
        tg.html()

    def test_render_html_attributes(self):
        result = render_html_attributes({})
        assert isinstance(result, Markup)
        assert result == ''

        result = render_html_attributes(
            {
                'text': 'abc',
                'empty': '',
                'bool1': True,
                'bool2': False,
                'none': None,
                'esc&': '<>"',
            },
        )
        assert isinstance(result, Markup)
        assert result == ' bool1 empty="" esc&amp;="&lt;&gt;&#34;" text="abc"'

    @_inrequest('/')
    def test_no_filters(self):
        class TGrid(Grid):
            Column('Test', Person.id)

        tg = TGrid()
        assert 'Add Filter' not in tg.html()

    def get_grid(self, **kwargs):
        g = SimpleGrid(**kwargs)
        g.set_records(self.key_data)
        g.apply_qs_args()
        return g

    @_inrequest('/thepage?perpage=5&onpage=1')
    def test_current_url(self):
        g = self.get_grid()
        assert g.html.current_url() == '/thepage?onpage=1&perpage=5'
        assert g.html.current_url(perpage=10) == '/thepage?onpage=1&perpage=10'

    @_inrequest('/thepage')
    def test_current_url_qs_prefix(self):
        g = self.get_grid(qs_prefix='dg_')
        assert g.html.current_url(perpage=10) == '/thepage?dg_perpage=10'

    @_inrequest('/thepage?perpage=5&onpage=1&dgreset=1')
    def test_current_url_reset_removed(self):
        g = self.get_grid()
        assert g.html.current_url() == '/thepage?onpage=1&perpage=5'
        assert g.html.current_url(perpage=10) == '/thepage?onpage=1&perpage=10'

    @_inrequest('/thepage?foo_dgreset=1')
    def test_current_url_reset_removed_prefix(self):
        g = self.get_grid(qs_prefix='foo_')
        assert g.html.current_url(perpage=5) == '/thepage?foo_perpage=5'

    @_inrequest('/thepage?perpage=5&onpage=1')
    def test_current_url_reset_added(self):
        g = self.get_grid()
        assert g.html.current_url(dgreset=1) == '/thepage?dgreset=1&onpage=1&perpage=5'

    @_inrequest('/thepage?perpage=5&onpage=1')
    def test_export_url(self):
        g = self.get_grid()
        assert g.html.export_url('xlsx') == '/thepage?export_to=xlsx&onpage=1&perpage=5'
        assert g.html.export_url('csv') == '/thepage?export_to=csv&onpage=1&perpage=5'

    @_inrequest('/thepage?onpage=3')
    def test_paging_url_first(self):
        g = self.get_grid()
        assert g.html.paging_url_first() == '/thepage?onpage=1&perpage=1'

    @_inrequest('/thepage?onpage=3')
    def test_paging_url_next(self):
        g = self.get_grid()
        assert g.html.paging_url_next() == '/thepage?onpage=4&perpage=1'

    @_inrequest('/thepage?onpage=3')
    def test_paging_url_prev(self):
        g = self.get_grid()
        assert g.html.paging_url_prev() == '/thepage?onpage=2&perpage=1'

    @_inrequest('/thepage?onpage=3')
    def test_paging_url_last(self):
        g = self.get_grid()
        assert g.html.paging_url_last() == '/thepage?onpage=5&perpage=1'

    @_inrequest(
        '/thepage?foo=bar&onpage=5&perpage=10&sort1=1&sort2=2&sort3=3&op(name)=eq&v1(name)'
        '=bob&v2(name)=fred&search=bar',
    )
    def test_reset_url(self):
        g = self.get_grid()
        assert g.html.reset_url() == f'/thepage?dgreset=1&foo=bar&session_key={g.session_key}'

    @_inrequest('/thepage?foo=bar&onpage=5')
    def test_form_action_url(self):
        g = self.get_grid()
        assert g.html.form_action_url() == f'/thepage?foo=bar&session_key={g.session_key}'

    @_inrequest('/thepage?foo=bar&onpage=5&session_override=1')
    def test_form_action_url_session_opts(self):
        class TestManager(SimpleGrid.manager.__class__):
            def get_args(self, previous_args):
                return MultiDict({'foo': 'bar', 'onpage': 5, 'session_override': 1})

        class TGrid(Grid):
            manager = TestManager()

            Column('ID', 'id')

        g = TGrid()
        g.set_records(self.key_data)
        g.apply_qs_args()
        assert g.html.form_action_url() == f'/thepage?foo=bar&session_key={g.session_key}'

    @_inrequest('/thepage?foo=bar&onpage=5')
    def test_form_action_method_get(self):
        g = self.get_grid()
        assert g.html.form_action_method() == 'get'

        pyq = PyQuery(g.html())
        assert pyq('form.header').attr('method') == 'get'
        assert not pyq('form.header > input[name="csrf_token"]')

    @_inrequest('/thepage?foo=bar&onpage=5')
    def test_form_action_method_post(self):
        class TestManager(SimpleGrid.manager.__class__):
            args_loaders = (
                extensions.RequestArgsLoader,
                extensions.RequestFormLoader,
            )

        class TGrid(Grid):
            manager = TestManager()

            Column('ID', 'id')

        g = TGrid()
        g.set_records(self.key_data)
        assert g.html.form_action_method() == 'post'

        pyq = PyQuery(g.html())
        assert pyq('form.header').attr('method') == 'post'
        assert pyq('form.header > input[name="csrf_token"]')

    @pytest.mark.parametrize(
        'page_param,input_value',
        [
            (1, '1'),
            (2, '2'),
            (3, '3'),
            (4, '4'),
            (5, '5'),
            (0, '1'),
            (6, '5'),
            ('abc', '1'),
        ],
    )
    def test_paging_select(self, page_param, input_value):
        @_inrequest(f'/thepage?onpage={page_param}')
        def check_paging():
            g = self.get_grid()
            select_html = g.html.paging_select()
            assert PyQuery(select_html).text().strip() == 'of 5'
            assert_tag(
                select_html,
                'input',
                id_='onpage',
                name='onpage',
                type='number',
                value=input_value,
                min='1',
                max='5',
            )

        check_paging()

    @_inrequest('/thepage?onpage=2')
    def test_paging_html(self):
        g = self.get_grid()
        input_html = g.html.paging_input()
        assert_tag(input_html, 'input', name='perpage', type='number', value='1')

        img_html = g.html.paging_img_first()
        assert_tag(
            img_html,
            'img',
            alt='<<',
            width='16',
            height='13',
            src='/static/webgrid/b_firstpage.png',
        )

        img_html = g.html.paging_img_first_dead()
        assert_tag(
            img_html,
            'img',
            alt='<<',
            width='16',
            height='13',
            src='/static/webgrid/bd_firstpage.png',
        )

        img_html = g.html.paging_img_prev()
        assert_tag(
            img_html,
            'img',
            alt='<',
            width='8',
            height='13',
            src='/static/webgrid/b_prevpage.png',
        )

        img_html = g.html.paging_img_prev_dead()
        assert_tag(
            img_html,
            'img',
            alt='<',
            width='8',
            height='13',
            src='/static/webgrid/bd_prevpage.png',
        )

        img_html = g.html.paging_img_next()
        assert_tag(
            img_html,
            'img',
            alt='>',
            width='8',
            height='13',
            src='/static/webgrid/b_nextpage.png',
        )

        img_html = g.html.paging_img_next_dead()
        assert_tag(
            img_html,
            'img',
            alt='>',
            width='8',
            height='13',
            src='/static/webgrid/bd_nextpage.png',
        )

        img_html = g.html.paging_img_last()
        assert_tag(
            img_html,
            'img',
            alt='>>',
            width='16',
            height='13',
            src='/static/webgrid/b_lastpage.png',
        )

        img_html = g.html.paging_img_last_dead()
        assert_tag(
            img_html,
            'img',
            alt='>>',
            width='16',
            height='13',
            src='/static/webgrid/bd_lastpage.png',
        )

        # since we are on page 2, all links should be live
        footer_html = g.html.footer()
        assert g.html.paging_img_first() in footer_html
        assert g.html.paging_img_next() in footer_html
        assert g.html.paging_img_prev() in footer_html
        assert g.html.paging_img_last() in footer_html

        g.set_paging(1, 1)
        g.set_records(self.key_data)
        footer_html = g.html.footer()
        assert g.html.paging_img_first() not in footer_html, footer_html
        assert g.html.paging_img_first_dead() in footer_html
        assert g.html.paging_img_prev_dead() in footer_html

        g.set_paging(2, 3)
        g.set_records(self.key_data)
        footer_html = g.html.footer()
        assert g.html.paging_img_last() not in footer_html, footer_html
        assert g.html.paging_img_next_dead() in footer_html
        assert g.html.paging_img_last_dead() in footer_html

    @_inrequest('/thepage?sort1=name&sort2=-id')
    def test_sorting_html(self):
        g = self.get_grid()

        select_html = g.html.sorting_select1()
        assert_tag(select_html, 'select', id_='sort1', name='sort1')
        assert_tag(select_html, 'option', value='', text='\N{NO-BREAK SPACE}')
        assert_tag(select_html, 'option', text='Name', selected=None, value='name')
        assert_tag(select_html, 'option', text='Name DESC', value='-name')
        assert_tag(select_html, 'option', text='ID', value='id')
        assert not find_tag(select_html, 'option', value='emails')

        select_html = g.html.sorting_select2()
        assert find_tag(select_html, 'option', text='Name', value='name').attr('selected') is None
        assert_tag(select_html, 'option', text='ID DESC', selected=None, value='-id')

        select_html = g.html.sorting_select3()
        assert_tag(select_html, 'option', selected=None, value='', text='\N{NO-BREAK SPACE}')

        heading_row = g.html.table_column_headings()
        assert 'sort-asc' not in heading_row
        assert 'sort-desc' not in heading_row

    @_inrequest('/thepage?sort1=name')
    def test_sorting_headers_asc(self):
        g = self.get_grid()
        heading_row = g.html.table_column_headings()
        assert_tag(heading_row, 'a', text='Name', class_='sort-asc', href='/thepage?sort1=-name')

    @_inrequest('/thepage?sort1=-name')
    def test_sorting_headers_desc(self):
        g = self.get_grid()
        heading_row = g.html.table_column_headings()
        assert_tag(heading_row, 'a', text='Name', class_='sort-desc', href='/thepage?sort1=name')

    @_inrequest(
        '/thepage?op(firstname)=eq&v1(firstname)=foo&op(createdts)=between&v1(createdts)='
        '2%2F15%2F12&&v2(createdts)=2012-02-16',
    )
    def test_filtering_input_html(self):
        g = PeopleGrid()

        filter_html = g.html.filtering_col_inputs1(g.key_column_map['firstname'])
        assert '<input id="firstname_input1" name="v1(firstname)" type="text" />' in filter_html, (
            filter_html
        )

        filter_html = g.html.filtering_col_inputs1(g.key_column_map['createdts'])
        assert '<input id="createdts_input1" name="v1(createdts)" type="text" />' in filter_html, (
            filter_html
        )

        filter_html = g.html.filtering_col_inputs2(g.key_column_map['createdts'])
        assert '<input id="createdts_input2" name="v2(createdts)" type="text" />' in filter_html, (
            filter_html
        )

        g.apply_qs_args()

        filter_html = g.html.filtering_col_inputs1(g.key_column_map['firstname'])
        assert (
            '<input id="firstname_input1" name="v1(firstname)" type="text" value="foo" />'
            in filter_html
        ), filter_html

        filter_html = g.html.filtering_col_inputs1(g.key_column_map['createdts'])
        assert (
            '<input id="createdts_input1" name="v1(createdts)" type="text" value='
            '"2012-02-15T00:00" />' in filter_html
        ), filter_html

        filter_html = g.html.filtering_col_inputs2(g.key_column_map['createdts'])
        assert (
            '<input id="createdts_input2" name="v2(createdts)" type="text" value='
            '"2012-02-16T23:59" />' in filter_html
        ), filter_html

    @_inrequest('/thepage?op(firstname)=foobar&v1(firstname)=baz')
    def test_filtering_invalid_operator(self):
        g = PeopleGrid()

        filter_html = g.html.filtering_col_inputs1(g.key_column_map['firstname'])
        assert '<input id="firstname_input1" name="v1(firstname)" type="text" />' in filter_html, (
            filter_html
        )

    @_inrequest('/thepage')
    def test_extra_filter_attrs(self):
        g = PeopleGrid()
        g.key_column_map['firstname'].filter.html_extra = {'data-special-attr': 'foo'}
        filter_html = g.html.filtering_table_row(g.key_column_map['firstname'])
        assert '<tr class="firstname_filter" data-special-attr="foo">' in filter_html, filter_html

    @_inrequest('/thepage')
    def test_filter_primary_op_specified(self):
        g = PeopleGrid()
        g.key_column_map['firstname'].filter.primary_op = '!eq'
        filter_html = g.html.filtering_table_row(g.key_column_map['firstname'])
        assert PyQuery(filter_html).find('option[value="!eq"]').attr('data-render') == 'primary'
        assert not PyQuery(filter_html).find('option[value="eq"]').attr('data-render')

    @_inrequest('/thepage')
    def test_filter_primary_op_not_specified(self):
        g = PeopleGrid()
        g.key_column_map['firstname'].filter.primary_op = None
        filter_html = g.html.filtering_table_row(g.key_column_map['firstname'])
        assert PyQuery(filter_html).find('option[value="eq"]').attr('data-render') == 'primary'
        assert not PyQuery(filter_html).find('option[value="!eq"]').attr('data-render')

    @_inrequest('/thepage')
    def test_multiselect_enum_options(self):
        class PeopleType(Enum):
            bob = 'Bob'
            charlie = 'Charlie'

        g = PeopleGrid()
        name_filter = OptionsEnumFilter(Person.firstname, enum_type=PeopleType).new_instance()
        name_filter.set('is', ['bob'])
        output = g.html.filtering_filter_options_multi(name_filter, 'foo')
        assert PyQuery(output).find('input[value="bob"]').attr('checked')
        assert not PyQuery(output).find('input[value="charlie"]').attr('checked')

    @_inrequest('/thepage')
    def test_confirm_export(self):
        g = PeopleGrid()
        assert json.loads(g.html.confirm_export()) == {'confirm_export': False, 'record_count': 3}

        g.unconfirmed_export_limit = 2
        assert json.loads(g.html.confirm_export()) == {'confirm_export': True, 'record_count': 3}

        g.unconfirmed_export_limit = None
        assert json.loads(g.html.confirm_export()) == {'confirm_export': False, 'record_count': 3}

    @_inrequest('/thepage')
    def test_grid_rendering(self):
        g = PeopleGrid()
        # really just making sure no exceptions come through at this point
        assert g.html()

    @_inrequest('/thepage')
    def test_no_records(self):
        g = PeopleGrid()
        g.set_records([])
        assert '<p class="no-records">No records to display</p>' in g.html()

    @_inrequest('/thepage')
    def test_no_pager(self):
        class PgNP(PeopleGrid):
            pager_on = False

        g = PgNP()
        assert '<td class="page">' not in g.html()
        assert '<td class="perpage">' not in g.html()
        assert '<th class="page">' not in g.html()
        assert '<th class="perpage">' not in g.html()

    def test_can_render(self):
        assert PeopleGrid().html.can_render() is True

    def test_render_error(self):
        class Renderer(HTML):
            def can_render(self):
                return False

        class TestGrid(PeopleGrid):
            def set_renderers(self):
                super().set_renderers()
                self.html = Renderer(self)

        with pytest.raises(RenderLimitExceeded):
            TestGrid().html()

    @_inrequest('/thepage?search=foo')
    def test_render_search_filter(self):
        g = PeopleGrid()
        g.enable_search = True

        filter_html = g.html.filtering_fields()
        tag = assert_tag(filter_html, 'input', id='search_input', name='search', type='text')
        assert tag.val() == ''

        g.apply_qs_args()
        filter_html = g.html.filtering_fields()
        assert_tag(filter_html, 'input', id='search_input', name='search', type='text', value='foo')

    @_inrequest('/thepage?search=foo&dgreset=1')
    def test_render_search_filter_reset(self):
        g = PeopleGrid()
        g.enable_search = True

        filter_html = g.html.filtering_fields()
        tag = assert_tag(filter_html, 'input', id='search_input', name='search', type='text')
        assert tag.val() == ''

        g.apply_qs_args()
        filter_html = g.html.filtering_fields()
        assert_tag(filter_html, 'input', id='search_input', name='search', type='text', value='')

    def test_search_disabled(self):
        class PeopleGrid2(PeopleGrid):
            enable_search = False

        g = PeopleGrid2()

        filter_html = g.html.filtering_fields()
        assert '<input id="search_input"' not in filter_html

    def test_no_searchable_columns(self):
        class TGrid(Grid):
            enable_search = True
            Column('Test', Person.id)

        tg = TGrid()
        filter_html = tg.html.filtering_fields()
        assert '<input id="search_input"' not in filter_html

    @_inrequest('/thepage?op(firstname)=contains&v1(firstname)=Fred')
    def test_filtering_operator_selected(self):
        g = PeopleGrid()
        g.apply_qs_args()

        pyq = PyQuery(g.html())

        # Ensures the op we have in the URL is valid.  If it wasn't valid, webgrid would not render
        # a value for input1.
        assert pyq('#firstname_input1').val() == 'Fred'

        selected_opts = pyq('tr.firstname_filter td.operator option[selected]')
        assert selected_opts.val() == 'contains'


class PGPageTotals(PeopleGrid):
    subtotals = 'page'


class TestPageTotals:
    @_inrequest('/')
    def test_people_html(self):
        g = PGPageTotals()
        html = g.html()
        elem = find_tag(html, 'td', class_='totals-label', colspan='7')
        assert len(elem) == 1
        assert elem.text() == 'Page Totals (3 records):'


class PGGrandTotals(PeopleGrid):
    subtotals = 'grand'


class TestGrandTotals:
    @_inrequest('/')
    def test_people_html(self):
        g = PGGrandTotals()
        assert '<td class="totals-label" colspan="7">Grand Totals (3 records):</td>' in g.html()
        assert '<td class="totals-label" colspan="7">Page Totals (3 records):</td>' not in g.html()


class TestFooterRendersCorrectly:
    @_inrequest('/')
    def test_people_html_footer(self):
        g = PeopleGrid()
        assert '<a class="export-link" href="/?export_to=xlsx">XLSX</a>' in g.html()

    @_inrequest('/')
    def test_people_html_footer_only_csv(self):
        g = PeopleCSVGrid()
        assert '<a class="export-link" href="/?export_to=csv">CSV</a>' in g.html()


class PGAllTotals(PeopleGrid):
    subtotals = 'all'


class TestAllTotals:
    @_inrequest('/')
    def test_people_html(self):
        g = PGAllTotals()
        html = g.html()
        assert_tag(html, 'td', text='Grand Totals (3 records):', class_='totals-label', colspan='7')
        assert_tag(html, 'td', text='Page Totals (3 records):', class_='totals-label', colspan='7')


class PGTotalsStringExpr(PeopleGrid):
    subtotals = 'all'
    Column('FloatCol', 'float_col', has_subtotal=True)

    def query_prep(self, query, has_sort, has_filters):
        query = super().query_prep(query, has_sort, has_filters)
        return query.add_columns(Person.floatcol.label('float_col'))


class TestStringExprTotals:
    @_inrequest('/')
    def test_people_html(self):
        g = PGTotalsStringExpr()
        html = g.html()

        assert_tag(html, 'td', text='Grand Totals (3 records):', class_='totals-label', colspan='7')
        assert_tag(html, 'td', text='Page Totals (3 records):', class_='totals-label', colspan='7')


class TestJSONRenderer:
    def get_json(self, grid):
        return json.loads(JSON(grid).render())

    def test_json_format_records(self):
        status_options = [
            {'key': status.id, 'value': status.label} for status in reversed(Status.list())
        ]
        expected = {
            'errors': [],
            'settings': {
                'filters': {},
                'paging': {'pager_on': True, 'on_page': 1, 'per_page': 50},
                'search_expr': None,
                'sort': [],
                'export_to': None,
            },
            'spec': {
                'columns': {
                    'account_type': 'Account Type',
                    'createdts': 'Created',
                    'due_date': 'Due Date',
                    'emails': 'Emails',
                    'firstname': 'First Name',
                    'full_name': 'Full Name',
                    'inactive': 'Active',
                    'numericcol': 'Number',
                    'state': 'State',
                    'status': 'Status',
                },
                'column_groups': [{'columns': ['firstname', 'inactive'], 'label': 'foo'}],
                'column_types': {
                    'account_type': None,
                    'createdts': 'datetime',
                    'due_date': 'date',
                    'emails': None,
                    'firstname': None,
                    'full_name': None,
                    'inactive': 'boolean',
                    'numericcol': 'number_general',
                    'status': None,
                },
                'enable_search': True,
                'enable_sort': True,
                'export_targets': ['xlsx'],
                'sortable_columns': [
                    'firstname',
                    'inactive',
                    'status',
                    'createdts',
                    'due_date',
                    'state',
                    'numericcol',
                    'account_type',
                ],
                'filters': {
                    'account_type': {
                        'operators': [
                            {'field_type': 'select', 'hint': None, 'key': 'is', 'label': 'is'},
                            {'field_type': 'select', 'hint': None, 'key': '!is', 'label': 'is not'},
                            {'field_type': None, 'hint': None, 'key': 'empty', 'label': 'empty'},
                            {
                                'field_type': None,
                                'hint': None,
                                'key': '!empty',
                                'label': 'not empty',
                            },
                        ],
                        'options': [
                            {'key': 'admin', 'value': 'Admin'},
                            {'key': 'manager', 'value': 'Manager'},
                            {'key': 'employee', 'value': 'Employee'},
                        ],
                        'primary_op': {
                            'field_type': 'select',
                            'hint': None,
                            'key': 'is',
                            'label': 'is',
                        },
                    },
                    'createdts': {
                        'operators': [
                            {
                                'field_type': 'input.datetime-local',
                                'hint': None,
                                'key': 'eq',
                                'label': 'is',
                            },
                            {
                                'field_type': 'input.datetime-local',
                                'hint': None,
                                'key': '!eq',
                                'label': 'is not',
                            },
                            {
                                'field_type': None,
                                'hint': None,
                                'key': 'past',
                                'label': 'in the past',
                            },
                            {
                                'field_type': None,
                                'hint': None,
                                'key': 'future',
                                'label': 'in the future',
                            },
                            {
                                'field_type': 'input.datetime-local',
                                'hint': None,
                                'key': 'lte',
                                'label': 'less than or equal',
                            },
                            {
                                'field_type': 'input.datetime-local',
                                'hint': None,
                                'key': 'gte',
                                'label': 'greater than or equal',
                            },
                            {
                                'field_type': '2inputs.datetime-local',
                                'hint': None,
                                'key': 'between',
                                'label': 'between',
                            },
                            {
                                'field_type': '2inputs.datetime-local',
                                'hint': None,
                                'key': '!between',
                                'label': 'not between',
                            },
                            {
                                'field_type': 'input',
                                'hint': 'days',
                                'key': 'da',
                                'label': 'days ago',
                            },
                            {
                                'field_type': 'input',
                                'hint': 'days',
                                'key': 'ltda',
                                'label': 'less than days ago',
                            },
                            {
                                'field_type': 'input',
                                'hint': 'days',
                                'key': 'mtda',
                                'label': 'more than days ago',
                            },
                            {'field_type': None, 'hint': None, 'key': 'today', 'label': 'today'},
                            {
                                'field_type': None,
                                'hint': None,
                                'key': 'thisweek',
                                'label': 'this week',
                            },
                            {
                                'field_type': None,
                                'hint': None,
                                'key': 'lastweek',
                                'label': 'last week',
                            },
                            {
                                'field_type': 'input',
                                'hint': 'days',
                                'key': 'ind',
                                'label': 'in days',
                            },
                            {
                                'field_type': 'input',
                                'hint': 'days',
                                'key': 'iltd',
                                'label': 'in less than days',
                            },
                            {
                                'field_type': 'input',
                                'hint': 'days',
                                'key': 'imtd',
                                'label': 'in more than days',
                            },
                            {'field_type': None, 'hint': None, 'key': 'empty', 'label': 'empty'},
                            {
                                'field_type': None,
                                'hint': None,
                                'key': '!empty',
                                'label': 'not empty',
                            },
                            {
                                'field_type': None,
                                'hint': None,
                                'key': 'thismonth',
                                'label': 'this month',
                            },
                            {
                                'field_type': None,
                                'hint': None,
                                'key': 'lastmonth',
                                'label': 'last month',
                            },
                            {
                                'field_type': 'select+input',
                                'hint': None,
                                'key': 'selmonth',
                                'label': 'select month',
                            },
                            {
                                'field_type': None,
                                'hint': None,
                                'key': 'thisyear',
                                'label': 'this year',
                            },
                        ],
                        'primary_op': None,
                        'options': [
                            {'key': 1, 'value': '01-Jan'},
                            {'key': 2, 'value': '02-Feb'},
                            {'key': 3, 'value': '03-Mar'},
                            {'key': 4, 'value': '04-Apr'},
                            {'key': 5, 'value': '05-May'},
                            {'key': 6, 'value': '06-Jun'},
                            {'key': 7, 'value': '07-Jul'},
                            {'key': 8, 'value': '08-Aug'},
                            {'key': 9, 'value': '09-Sep'},
                            {'key': 10, 'value': '10-Oct'},
                            {'key': 11, 'value': '11-Nov'},
                            {'key': 12, 'value': '12-Dec'},
                        ],
                    },
                    'firstname': {
                        'operators': [
                            {'field_type': 'input', 'hint': None, 'key': 'eq', 'label': 'is'},
                            {'field_type': 'input', 'hint': None, 'key': '!eq', 'label': 'is not'},
                            {
                                'field_type': 'input',
                                'hint': None,
                                'key': 'contains',
                                'label': 'contains',
                            },
                            {
                                'field_type': 'input',
                                'hint': None,
                                'key': '!contains',
                                'label': "doesn't contain",
                            },
                            {'field_type': None, 'hint': None, 'key': 'empty', 'label': 'empty'},
                            {
                                'field_type': None,
                                'hint': None,
                                'key': '!empty',
                                'label': 'not empty',
                            },
                        ],
                        'primary_op': {
                            'field_type': 'input',
                            'hint': None,
                            'key': 'contains',
                            'label': 'contains',
                        },
                    },
                    'status': {
                        'operators': [
                            {'field_type': None, 'hint': None, 'key': 'o', 'label': 'open'},
                            {'field_type': 'select', 'hint': None, 'key': 'is', 'label': 'is'},
                            {'field_type': 'select', 'hint': None, 'key': '!is', 'label': 'is not'},
                            {'field_type': None, 'hint': None, 'key': 'c', 'label': 'closed'},
                            {'field_type': None, 'hint': None, 'key': 'empty', 'label': 'empty'},
                            {
                                'field_type': None,
                                'hint': None,
                                'key': '!empty',
                                'label': 'not empty',
                            },
                        ],
                        'options': status_options,
                        'primary_op': {
                            'field_type': 'select',
                            'hint': None,
                            'key': 'is',
                            'label': 'is',
                        },
                    },
                },
            },
            'state': {
                'page_count': 1,
                'record_count': 3,
                'warnings': [],
            },
            'records': [
                {
                    'account_type': None,
                    'createdts': '2012-02-22T10:04:16',
                    'due_date': '2012-02-04',
                    'emails': 'email004@example.com, email004@gmail.com',
                    'firstname': 'fn004',
                    'full_name': 'fn004 ln004',
                    'inactive': 'Yes',
                    'numericcol': 2.13,
                    'status': None,
                },
                {
                    'account_type': 'Employee',
                    'createdts': None,
                    'due_date': None,
                    'emails': 'email002@example.com, email002@gmail.com',
                    'firstname': 'fn002',
                    'full_name': 'fn002 ln002',
                    'inactive': 'Yes',
                    'numericcol': 2.13,
                    'status': 'pending',
                },
                {
                    'account_type': 'Admin',
                    'createdts': '2012-02-22T10:01:16',
                    'due_date': '2012-02-01',
                    'emails': 'email001@example.com, email001@gmail.com',
                    'firstname': 'fn001',
                    'full_name': 'fn001 ln001',
                    'inactive': 'Yes',
                    'numericcol': 2.13,
                    'status': 'in process',
                },
            ],
            'totals': {
                'page': None,
                'grand': None,
            },
        }
        grid = PeopleGrid()
        group = ColumnGroup('foo')
        grid.column('firstname').group = grid.column('inactive').group = group
        assert self.get_json(grid) == expected

    def test_subtotals(self):
        grid = PGAllTotals()
        grid.set_paging(1, 1)
        grid_json = self.get_json(grid)
        assert grid_json['totals'] == {
            'page': {'numericcol': 2.13},
            'grand': {'numericcol': 6.39},
        }

    def test_warnings(self):
        grid = PeopleGrid()
        grid.user_warnings = ['foo', 'bar', 'baz']
        assert self.get_json(grid)['state']['warnings'] == ['foo', 'bar', 'baz']

    def test_json_format_settings(self):
        grid = PeopleGrid()
        firstname = grid.column('firstname')
        firstname.filter.set('eq', 'bar', 'baz')
        # Ensure that raw filter value is serialized
        firstname.filter.value1 = 'bong'
        firstname.filter.value2 = 'bing'
        grid.set_paging(20, 2)
        grid.search_value = 'foo'
        grid.set_sort('firstname', '-status')
        assert self.get_json(grid)['settings'] == {
            'filters': {
                'firstname': {'op': 'eq', 'value1': 'bar', 'value2': 'baz'},
            },
            'paging': {'pager_on': True, 'on_page': 2, 'per_page': 20},
            'search_expr': 'foo',
            'sort': [
                {'key': 'firstname', 'flag_desc': False},
                {'key': 'status', 'flag_desc': True},
            ],
            'export_to': None,
        }

    def test_json_format_lists(self):
        grid = PeopleGrid()
        account_type = grid.column('account_type')
        account_type.filter.set('is', ['admin', 'manager'])
        assert self.get_json(grid)['settings'] == {
            'filters': {
                'account_type': {'op': 'is', 'value1': ['admin', 'manager'], 'value2': None},
            },
            'paging': {'pager_on': True, 'on_page': 1, 'per_page': 50},
            'search_expr': None,
            'sort': [],
            'export_to': None,
        }

    def test_json_format_arrow(self):
        ArrowRecord.query.delete()
        ArrowRecord.testing_create(created_utc=arrow.Arrow(2016, 8, 10, 1, 2, 3))
        reloaded = self.get_json(ArrowGrid())
        assert reloaded['records'] == [{'created_utc': '2016-08-10T01:02:03+00:00'}]


class TestXLSXRenderer:
    def test_using_xlsxwriter_library(self):
        g = render_in_grid(PeopleGrid, 'xlsx')(per_page=1)
        wb = g.xlsx(manager_cls=XLSXWriterWorkbookManager)
        assert isinstance(wb._workbook, xlsxwriter.workbook.Workbook)

    def test_using_openpyxl_library_default(self):
        g = render_in_grid(PeopleGrid, 'xlsx')(per_page=1)
        wb = g.xlsx()
        assert isinstance(wb._workbook, openpyxl.Workbook)

    def test_some_basics(self):
        g = render_in_grid(PeopleGrid, 'xlsx')(per_page=1)
        wb = g.xlsx()
        wb.filename.seek(0)
        book = openpyxl.load_workbook(wb.filename)
        sh = book['render_in_grid']

        # headers
        assert sh.max_column == 10
        assert sh.cell(1, 1).value == 'First Name'
        assert sh.cell(1, 8).value == 'State'

        # last data row
        assert sh.max_row == 4
        assert sh.cell(4, 1).value == 'fn001'
        assert sh.cell(4, 8).value == 'st001'

    def test_group_headings(self):
        grid = StopwatchGrid()
        wb = grid.xlsx()
        wb.filename.seek(0)
        book = openpyxl.load_workbook(wb.filename)
        sheet = book[book.sheetnames[0]]
        #   [ A | B | C | D | E | F | G | H | I ]
        # 1 [       | Lap 1 |   | Lap 2 | Lap 3 ]
        row_values = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        assert row_values == [None, None, 'Lap 1', None, None, 'Lap 2', None, 'Lap 3', None]
        assert sheet.cell(2, 2).value == 'Label'
        assert sheet.cell(3, 2).value == 'Watch 1'
        assert {str(range_) for range_ in sheet.merged_cells.ranges} == {
            'A1:B1',
            'C1:D1',
            # E is a single cell
            'F1:G1',
            'H1:I1',
        }
        assert sheet.max_column == 9

    def test_group_headings_xlsxwriter(self):
        grid = StopwatchGrid()
        wb = grid.xlsx(manager_cls=XLSXWriterWorkbookManager)
        wb.filename.seek(0)
        wb.close()
        book = openpyxl.load_workbook(wb.filename)
        sheet = book[book.sheetnames[0]]
        #   [ A | B | C | D | E | F | G | H | I ]
        # 1 [       | Lap 1 |   | Lap 2 | Lap 3 ]
        row_values = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        assert row_values == [None, None, 'Lap 1', None, None, 'Lap 2', None, 'Lap 3', None]
        assert sheet.cell(2, 2).value == 'Label'
        assert sheet.cell(3, 2).value == 'Watch 1'
        assert {str(range_) for range_ in sheet.merged_cells.ranges} == {
            'A1:B1',
            'C1:D1',
            # E is a single cell
            'F1:G1',
            'H1:I1',
        }
        assert sheet.max_column == 9

    def test_subtotals_with_no_records(self):
        g = PGGrandTotals()
        g.column('firstname').filter.op = 'eq'
        g.column('firstname').filter.value1 = 'foobar'
        wb = g.xlsx()
        wb.filename.seek(0)

    def test_long_grid_name_xlsxwriter(self):
        class PeopleGridWithAReallyReallyLongName(PeopleGrid):
            pass

        g = PeopleGridWithAReallyReallyLongName()
        wb = g.xlsx(manager_cls=XLSXWriterWorkbookManager)
        wb.close()
        wb.filename.seek(0)

        book = openpyxl.load_workbook(wb.filename)
        assert book['people_grid_with_a_really_r...']

    def test_long_grid_name_openpyxl(self):
        class PeopleGridWithAReallyReallyLongName(PeopleGrid):
            pass

        g = PeopleGridWithAReallyReallyLongName()
        wb = g.xlsx(manager_cls=OpenpyxlWorkbookManager)
        wb.filename.seek(0)

        book = openpyxl.load_workbook(wb.filename)
        assert book['people_grid_with_a_really_r...']

    def test_totals_xlsxwriter(self):
        g = PeopleGrid()
        g.subtotals = 'grand'

        wb = g.xlsx(manager_cls=XLSXWriterWorkbookManager)
        wb.filename.seek(0)
        wb.close()

        book = openpyxl.load_workbook(wb.filename)
        sheet = book[book.sheetnames[0]]
        assert sheet.max_row == 5
        assert sheet.cell(5, 1).value == 'Totals (3 records):'
        assert sheet.cell(5, 9).value == 6.39
        assert [str(range_) for range_ in sheet.merged_cells.ranges] == ['A5:H5']

    def test_totals_openpyxl(self):
        g = PeopleGrid()
        g.subtotals = 'grand'

        wb = g.xlsx()
        wb.filename.seek(0)

        book = openpyxl.load_workbook(wb.filename)
        sheet = book[book.sheetnames[0]]
        assert sheet.max_row == 5
        assert sheet.cell(5, 1).value == 'Totals (3 records):'
        assert sheet.cell(5, 9).value == 6.39
        assert [str(range_) for range_ in sheet.merged_cells.ranges] == ['A5:H5']

    def test_totals_no_merge(self):
        class TestGrid(Grid):
            subtotals = 'all'
            Column('First Name', Person.firstname)
            NumericColumn('Number', Person.numericcol, has_subtotal=True)

        g = TestGrid()
        wb = g.xlsx()
        wb.filename.seek(0)

        book = openpyxl.load_workbook(wb.filename)
        sheet = book[book.sheetnames[0]]

        assert sheet.max_row == 6
        assert sheet.cell(6, 1).value == 'Totals (4 records):'
        assert set(sheet.merged_cells.ranges) == set()

    def test_can_render(self):
        class FakeCountsGrid(PeopleGrid):
            def __init__(self, record_count, col_count, has_subtotals):
                self._num_records = record_count
                self._col_count = col_count
                self.subtotals = 'all' if has_subtotals else 'none'
                super().__init__()

            @property
            def record_count(self):
                return self._num_records

            def iter_columns(self, render_type):
                for _ in range(self._col_count):
                    yield None

        assert FakeCountsGrid(1048575, 16384, False).xlsx.can_render() is True
        assert FakeCountsGrid(1048576, 16384, False).xlsx.can_render() is False
        assert FakeCountsGrid(1048575, 16384, True).xlsx.can_render() is False
        assert FakeCountsGrid(1048574, 16384, True).xlsx.can_render() is True
        assert FakeCountsGrid(1048575, 16385, False).xlsx.can_render() is False

    def test_render_error(self):
        class Renderer(XLSX):
            def can_render(self):
                return False

        class TestGrid(PeopleGrid):
            def set_renderers(self):
                super().set_renderers()
                self.xlsx = Renderer(self)

        with pytest.raises(RenderLimitExceeded):
            TestGrid().xlsx()

    def test_xlsx_format_caching(self):
        grid = PeopleGrid()
        wb = XLSXWriterWorkbookManager()
        format1 = wb.style_for_column(grid.column('status'))
        format2 = wb.style_for_column(grid.column('state'))
        format3 = wb.style_for_column(grid.column('numericcol'))

        assert format1 is not None
        assert format2 is not None
        assert format3 is not None

        assert format1 is format2
        assert format3 is not format1

        grid = PeopleGrid()
        grid.column('status').xlsx_style = {'bold': True, 'border': 1}
        grid.column('state').xlsx_style = {'bold': True, 'border': 2}
        wb = XLSXWriterWorkbookManager()

        format1 = wb.style_for_column(grid.column('status'))
        format2 = wb.style_for_column(grid.column('state'))
        assert format1 is not format2

        grid = PeopleGrid()
        grid.column('status').xlsx_style = {'bold': True, 'border': 1}
        grid.column('state').xlsx_style = {'bold': True, 'border': 1}
        wb = XLSXWriterWorkbookManager()

        format1 = wb.style_for_column(grid.column('status'))
        format2 = wb.style_for_column(grid.column('state'))
        assert format1 is format2


class TestCSVRenderer:
    def test_some_basics(self):
        g = render_in_grid(PeopleCSVGrid, 'csv')(per_page=1)
        csv_data = g.csv.build_csv()
        csv_data.seek(0)
        byte_str = io.StringIO(csv_data.read().decode('utf-8'))
        reader = csv.reader(byte_str, delimiter=',', quotechar='"')
        data = []
        for row in reader:
            data.append(row)
        assert len(data[0]) == 9
        assert data[0][0] == 'First Name'
        assert data[0][2] == 'Active'
        assert data[1][0] == 'fn004'

    def test_it_renders_date_time_with_tz(self):
        ArrowRecord.query.delete()
        ArrowRecord.testing_create(created_utc=arrow.Arrow(2016, 8, 10, 1, 2, 3))
        g = ArrowCSVGrid()
        g.allowed_export_targets = {'csv': CSV}
        csv_data = g.csv.build_csv()
        csv_data.seek(0)
        byte_str = io.StringIO(csv_data.read().decode('utf-8'))
        reader = csv.reader(byte_str, delimiter=',', quotechar='"')
        data = []
        for row in reader:
            data.append(row)
        assert data[0][0] == 'Created'
        assert data[1][0] == '2016-08-10 01:02:03+0000'

    def test_it_renders_date_time_with_custom_format(self):
        class CSVGrid(Grid):
            session_on = True
            allowed_export_targets: ClassVar = {'csv': CSV}
            DateTimeColumn('Created', ArrowRecord.created_utc, csv_format='%m/%d/%Y %I:%M %p')

        ArrowRecord.query.delete()
        ArrowRecord.testing_create(created_utc=arrow.Arrow(2016, 8, 10, 1, 2, 3))
        g = CSVGrid()
        csv_data = g.csv.build_csv()
        csv_data.seek(0)
        byte_str = io.StringIO(csv_data.read().decode('utf-8'))
        reader = csv.reader(byte_str, delimiter=',', quotechar='"')
        data = []
        for row in reader:
            data.append(row)
        assert data[0][0] == 'Created'
        assert data[1][0] == '08/10/2016 01:02 AM'


class TestHideSection:
    @_inrequest('/')
    def test_controlls_hidden(self):
        class NoControlBoxGrid(PG):
            hide_controls_box = True

        g = NoControlBoxGrid()
        assert '<tr class="status"' not in g.html()
        assert '<div class="footer">' not in g.html()


class TestArrowDate:
    def setup_method(self):
        ArrowRecord.query.delete()

    @_inrequest('/')
    def test_arrow_render_html(self):
        ArrowRecord.testing_create(created_utc=arrow.Arrow(2016, 8, 10, 1, 2, 3))
        g = ArrowGrid()
        assert '<td>08/10/2016 01:02 AM</td>' in g.html(), g.html()

        g.column('created_utc').html_format = 'YYYY-MM-DD HH:mm:ss ZZ'
        assert '<td>2016-08-10 01:02:03 +00:00</td>' in g.html(), g.html()

    @_inrequest('/')
    def test_arrow_timezone(self):
        # regardless of timezone given, ArrowType stored as UTC and will display that way
        ArrowRecord.testing_create(created_utc=arrow.Arrow(2016, 8, 10, 1, 2, 3).to('US/Pacific'))
        g = ArrowGrid()
        assert '<td>08/10/2016 01:02 AM</td>' in g.html(), g.html()

        g.column('created_utc').html_format = 'YYYY-MM-DD HH:mm:ss ZZ'
        assert '<td>2016-08-10 01:02:03 +00:00</td>' in g.html(), g.html()

    def test_filter_handles_arrow(self):
        ArrowRecord.testing_create(created_utc=arrow.Arrow(2016, 8, 10, 1, 2, 3))
        g = ArrowGrid()
        g.column('created_utc').filter = DateFilter(g.column('created_utc').expr)
        g.apply_qs_args(grid_args={'op(created_utc)': 'eq', 'v1(created_utc)': '2018-01-01'})
        assert g.column('created_utc').filter.is_active
        g.record_count  # noqa: B018
