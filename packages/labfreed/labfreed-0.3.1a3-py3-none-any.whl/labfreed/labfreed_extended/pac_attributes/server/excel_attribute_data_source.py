from datetime import datetime

import os
from urllib.parse import urlparse
from cachetools import TTLCache, cached

from labfreed.pac_attributes.api_data_models.response import AttributeGroup
from labfreed.pac_attributes.server.server import AttributeGroupDataSource
from labfreed.labfreed_extended.pac_attributes.py_attributes import pyAttribute, pyAttributes


try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("Please install labfreed with the [extended] extra: pip install labfreed[extended]")


cache = TTLCache(maxsize=128, ttl=0)


class ExcelAttributeDataSource(AttributeGroupDataSource):
    '''
    Demonstrates how to analyze the PAC-ID and it's extensions to provide some data
    '''
    def __init__(self, file_path:str, cache_duration_seconds:int=0, base_url:str="", *args, **kwargs):
        self._file_path = file_path
        self._base_url = base_url

        
        if is_sharepoint_url(file_path):
            self._file_location = "sharepoint"
        else: 
            self._file_location = "local"
            
        super().__init__(*args, **kwargs)
            
         
    def is_static(self) -> bool:
        return False
    
    @property
    def provides_attributes(self):
        if self._file_location == "local":
            rows, last_changed = read_excel_openpyxl(self._file_path)
            headers = [self._base_url + r for r in rows[0][1:] ]
        elif self._file_location == "sharepoint":
            raise NotImplementedError('Sharepoint Access not implemented')
        return headers
    
    
    def attributes(self, pac_url: str) -> AttributeGroup:
        if not self._include_extensions:
            pac_url = pac_url.split('*')[0]
        
        if self._file_location == "local":
            rows, last_changed = read_excel_openpyxl(self._file_path)
        elif self._file_location == "sharepoint":
            raise NotImplementedError('Sharepoint Access not implemented')
        
        d = get_row_by_first_cell(rows, pac_url, self._base_url)
        if not d:
            return None
        
        attributes = [pyAttribute(key=k, value=v) for k,v in d.items()]
        
            
        return AttributeGroup(key=self._attribute_group_key, 
                              attributes=pyAttributes(attributes).to_payload_attributes(),
                              state_of=last_changed)
    
    
    
@cached(cache)
def read_excel_openpyxl(path: str, worksheet: str = None) -> list[tuple]:
    """
    Read and cache Excel worksheet as list of rows (including headers),
    then close the workbook.
    """
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[worksheet] if worksheet else wb.active

    rows = list(ws.iter_rows(values_only=True))
    
    last_changed: datetime | None = wb.properties.modified
    wb.close()  # immediately release the file

    return rows, last_changed  # list of tuples (header + data rows)


def get_row_by_first_cell(sheet_rows: list[tuple], match_value: str, base_url:str) -> dict | None:
    """
    Takes list of rows and returns the first row where the first cell == match_value,
    as a dict using headers from the first row.
    """
    if not sheet_rows:
        return None

    headers = sheet_rows[0]
    for row in sheet_rows[1:]:
        if not row:
            continue
        first = str(row[0]).strip() if row[0] is not None else ""
        if first == match_value:
            return {
                base_url + str(headers[i]).strip(): row[i]
                for i in range(1, len(headers))
                if headers[i] is not None
            }

    return None

        
        
def is_sharepoint_url(s: str) -> bool:
    try:
        parsed = urlparse(s)
        if parsed.scheme not in {"http", "https"}:
            return False
        if "sharepoint.com" in parsed.netloc or "1drv.ms" in parsed.netloc:
            return True
        return False
    except Exception:
        return False

def is_local_path(s: str) -> bool:
    if is_sharepoint_url(s):
        return False
    # Treat anything that's not a URL and points to an existing file as a local path
    return os.path.exists(s)