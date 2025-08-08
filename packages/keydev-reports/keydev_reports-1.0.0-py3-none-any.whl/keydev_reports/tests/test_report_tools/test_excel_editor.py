import pytest
from openpyxl import Workbook
from keydev_reports.report_tools.excel_editor import ExcelEditor
from openpyxl.cell.cell import MergedCell


@pytest.fixture
def mock_load_workbook(mocker):
    # Create a mock for the load_workbook function
    mock_workbook = Workbook()
    mocker.patch(
        'keydev_reports.report_tools.excel_editor.load_workbook',
        return_value=mock_workbook)


@pytest.fixture
def excel_editor(mock_load_workbook):
    # Instantiate ExcelEditor with a mocked load_workbook function
    return ExcelEditor(file_name='test.xlsx')


@pytest.fixture
def mock_worksheet(excel_editor):
    # Return the active sheet from the excel_editor's workbook
    return excel_editor.book.active


class TestExcelEditor:

    def test_find_target_cell(self, excel_editor, mock_worksheet):
        mock_worksheet['A3'] = 'Cell 3'

        # Test when the cell is found
        target_cell = excel_editor.find_target_cell(mock_worksheet, 'Cell 3')
        assert target_cell.value == 'Cell 3'

        # Test when the cell is not found
        with pytest.raises(ValueError):
            excel_editor.find_target_cell(mock_worksheet, 'Non-existing Cell')

    @pytest.mark.parametrize(
        'data, start_row, start_col, expected_data',
        [
            (
                [[1, 2, 3], [4, 5, 6], [7, 8, 9]], 2, 3,  # input data
                [[1, 2, 3], [4, 5, 6], [7, 8, 9]]  # expected output
            ),
            (
                [[1, 2, '=SUM(A2:C2)'], [4, '=A3*2', 6], [7, 8, 9]], 2, 3,  # input data
                [[1, 2, '=SUM(A2:C2)'], [4, '=A3*2', 6], [7, 8, 9]]  # expected output
            )
        ]
    )
    def test_populate_sheet_data(
            self,
            excel_editor,
            mock_worksheet,
            data,
            start_row,
            start_col,
            expected_data
    ):
        excel_editor.populate_sheet_data(mock_worksheet, data, start_row, start_col)

        for row_index, row_data in enumerate(expected_data, start=start_row):
            for col_index, value in enumerate(row_data, start=start_col):
                assert mock_worksheet.cell(row=row_index, column=col_index).value == value

    @pytest.mark.parametrize(
        'search_text, replace_text, initial_value, expected_value',
        [
            ('Cell 1', 'Cell 2', 'Cell 1', 'Cell 2'),  # Case 1: Single Replacement
            ('Cell 1', 'Cell 2', 'Cell 1 Cell 1', 'Cell 2 Cell 2'),  # Case 3: Multiple Replacements
            ('!@#$%^&*()', 'Special', 'Cell !@#$%^&*()', 'Cell Special'),  # Case 4: Special Characters
            ('12345', 'Number', '12345', 'Number'),  # Case 5: Numeric Search Text
        ]
    )
    def test_replace_cell_value(
        self,
        excel_editor,
        mock_worksheet,
        search_text,
        replace_text,
        initial_value,
        expected_value
    ):
        mock_worksheet['A3'] = initial_value
        excel_editor.replace_cell_value(search_text=search_text, replace_text=replace_text)
        assert mock_worksheet['A3'].value == expected_value

    def test_replace_data(self, excel_editor, mock_worksheet):
        mock_worksheet['A1'] = '${1} World 1'
        mock_worksheet['B1'] = 'Hello ${2} 2'
        mock_worksheet['C1'] = 'Hello World ${3} 3'
        excel_editor.replace_data(data={'1': 'Hello', '2': 'World', '3': '!!!'})
        assert mock_worksheet['A1'].value == 'Hello World 1'
        assert mock_worksheet['B1'].value == 'Hello World 2'
        assert mock_worksheet['C1'].value == 'Hello World !!! 3'

    def test_shift_cells(self, excel_editor, mock_worksheet):
        sheet = mock_worksheet
        sheet['A1'].value = 'Hello'
        sheet['B1'].value = 'World'
        sheet['C1'].value = '!!!'
        sheet.merge_cells('A1:C1')
        excel_editor.shift_cells(sheet, 1, 2)

        assert sheet['A1'].value is None
        assert sheet['B1'].value is None
        assert sheet['C1'].value is None
        assert sheet['A3'].value == 'Hello'
        assert sheet['B3'].value is None
        assert sheet['C3'].value is None
        assert isinstance(sheet['B3'], MergedCell)
        assert isinstance(sheet['C3'], MergedCell)

    def test_fill_data_from_startpoint(self, excel_editor, mock_worksheet):
        # Test data
        data = {
            'SYMBOL1': [
                ['Value11', 'Value12'],
                ['Value21', 'Value22']
            ],
            'SYMBOL2': [
                ['Value31', 'Value32'],
                ['Value41', 'Value42']
            ]
        }

        # Placeholders in the sheet
        mock_worksheet['A1'] = '${SYMBOL1}'
        mock_worksheet['D5'] = '${SYMBOL2}'

        # Calling the method
        excel_editor.fill_data_from_startpoint(data)

        # Check if the data was filled correctly
        assert mock_worksheet['A1'].value == 'Value11'
        assert mock_worksheet['B1'].value == 'Value12'
        assert mock_worksheet['D5'].value == 'Value31'
        assert mock_worksheet['E5'].value == 'Value32'

    # @pytest.mark.parametrize(
    #     'data, start_row, start_col, expected_values',
    #     [
    #         (
    #                 [['V1', 'V2', 'V3', 'V4'], ['V11', 'V22', 'V33', 'V44']],
    #                 1, 1,
    #                 [['V1', 'V2', 'V3', 'V4'], ['V11', 'V22', 'V33', 'V44']],
    #         )
    #     ]
    # )
    # def test_populate_data_in_merge_range(self, excel_editor, data, start_row, start_col, expected_values):
    #     # Merge cells in the sheet
    #     sheet = excel_editor.book.active
    #     sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    #     excel_editor.populate_data_in_merge_range(sheet, data, start_row, start_col)
    #     # print(excel_editor.get_merged_cell_ranges_dict(sheet))
    #     for row_idx, row_data in enumerate(expected_values, start=1):
    #         for col_idx, expected_value in enumerate(row_data, start=1):
    #             # For merged cells, check the value of the starting cell
    #             # print(row_idx, col_idx, expected_value)
    #             # sheet.cell(row=row_idx, column=start_col).value
    #             # print(f'{sheet.cell(row=row_idx, column=start_col)} {starting_cell_value}')
    #             # assert starting_cell_value == expected_value
    #     assert 1 == 2

    def test_get_merged_cell_ranges_dict(self, excel_editor, mocker):
        # Create a mock sheet with merged cells
        mock_sheet = mocker.Mock()

        # Mock MergedCellRange objects
        mock_merged_cell_range_1 = mocker.Mock()
        mock_merged_cell_range_1.cells = [(1, 1), (1, 2)]

        mock_merged_cell_range_2 = mocker.Mock()
        mock_merged_cell_range_2.cells = [(2, 1), (2, 2), (2, 3)]

        mock_sheet.merged_cells.ranges = [mock_merged_cell_range_1, mock_merged_cell_range_2]

        # Mock cell objects for different coordinates
        mock_cell_1_1 = mocker.Mock()
        mock_cell_1_2 = mocker.Mock()
        mock_cell_2_1 = mocker.Mock()
        mock_cell_2_2 = mocker.Mock()
        mock_cell_2_3 = mocker.Mock()

        def cell_side_effect(*args):
            calls = {
                (1, 1): mock_cell_1_1,
                (1, 2): mock_cell_1_2,
                (2, 1): mock_cell_2_1,
                (2, 2): mock_cell_2_2,
                (2, 3): mock_cell_2_3,
            }
            return calls[args]

        mock_sheet.cell.side_effect = cell_side_effect

        merged_cell_ranges_dict = excel_editor.get_merged_cell_ranges_dict(mock_sheet)

        assert len(merged_cell_ranges_dict) == 5

        # Test specific mappings
        assert merged_cell_ranges_dict[mock_cell_1_1] == mock_merged_cell_range_1
        assert merged_cell_ranges_dict[mock_cell_1_2] == mock_merged_cell_range_1
        assert merged_cell_ranges_dict[mock_cell_2_1] == mock_merged_cell_range_2
        assert merged_cell_ranges_dict[mock_cell_2_2] == mock_merged_cell_range_2
        assert merged_cell_ranges_dict[mock_cell_2_3] == mock_merged_cell_range_2

        # Test empty input
        mock_sheet.merged_cells.ranges = []
        merged_cell_ranges_dict = excel_editor.get_merged_cell_ranges_dict(mock_sheet)
        assert len(merged_cell_ranges_dict) == 0

        # Test edge cases
        mock_sheet.merged_cells.ranges = [mock_merged_cell_range_1]
        merged_cell_ranges_dict = excel_editor.get_merged_cell_ranges_dict(mock_sheet)
        assert len(merged_cell_ranges_dict) == 2

    # @pytest.mark.parametrize("data, expected_merged_cells", [
    #     ([(1, 1), (2, 1), (3, 1)], [('A1', 'A3')]),  # All cells in the range are the same
    #     # ([(1, 1), (2, 2), (3, 3)], []),  # All cells in the range are different
    #     # ([(1, 1), (1, 1), (1, 1), (2, 2), (3, 3)], [('A1', 'A3')]),  # Some cells are the same
    #     # ([(1, 1), (2, 1), (2, 1), (3, 3), (3, 3)], [('B1', 'B3')]),  # Some cells in a different column are the same
    #     # ([], []),  # Empty range
    # ])
    # def test_merge_rows(self, excel_editor, mocker, data, expected_merged_cells):
    #     # Create a mock worksheet with the provided data
    #     wb = Workbook()
    #     ws = wb.active
    #     for row, col in data:
    #         ws.cell(row=row, column=col).value = 1
    #
    #     # Mock get_merged_cell_ranges_dict to return an empty dict
    #     mocker.patch.object(excel_editor, 'get_merged_cell_ranges_dict', return_value={})
    #     print(f'{len(data)} sdsdfsdf')
    #     # Call merge_rows
    #     excel_editor.merge_rows(ws, start_row=1, end_row=len(data))
    #
    #     # Get the actual merged cell ranges
    #     actual_merged_cells = [(str(range_ref)) for range_ref in ws.merged_cells.ranges]
    #     print(actual_merged_cells)
    #     # Check the result
    #     assert actual_merged_cells == expected_merged_cells
