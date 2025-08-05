"""
Integration tests for verifying Microsoft Office Excel functionality through SharePoint.

Tests included:
1. `test_list_files_from_sharepoint_folder`:
   - Verifies that a known test Excel file exists in the SharePoint folder.

2. `test_append_row_to_excel`:
   - Appends a row to a specific sheet in the Excel file.
   - Verifies the row was successfully added.

3. `test_format_and_sort_excel_file`:
   - Sorts the Excel sheet by date/time (descending).
   - Applies formatting (bold headers, auto column widths).
   - Verifies that sorting has changed the order.

Important constants:
- FOLDER_NAME = "MSOffice tests"
- FILE_NAME = "Test_Append_Rows.xlsx"
- SHEET_NAME = "Upload logs"
- HEADERS = ["Upload date", "Upload time"]

Requires environment variables:
- MSOFFICE_USERNAME
- MSOFFICE_PASSWORD
"""

import os
import datetime

from io import BytesIO

import pytest
import openpyxl

from mbu_dev_shared_components.msoffice365.sharepoint_api.files import Sharepoint

FOLDER_NAME = "MSOffice tests"
CURRENT_DATE = datetime.datetime.now().strftime("%d-%m-%Y")
CURRENT_TIME = datetime.datetime.now().strftime("%H:%M")


@pytest.fixture(scope="module")
def sharepoint_api():
    """
    Authenticates with SharePoint and returns a Sharepoint API instance.
    Skips test if required environment variables are missing.
    """

    def _get_cfg(key: str) -> str:
        val = os.getenv(key)

        if not val:
            pytest.skip(f"env var '{key}' not set → skipping integration test")

        return val

    username = _get_cfg("MSOFFICE_USERNAME")
    password = _get_cfg("MSOFFICE_PASSWORD")

    site_url = "https://aarhuskommune.sharepoint.com"
    site_name = "MBURPA"
    document_library = "Delte dokumenter"

    sp = Sharepoint(
        username=username,
        password=password,
        site_url=site_url,
        site_name=site_name,
        document_library=document_library
    )

    assert sp.ctx is not None, "SharePoint authentication failed"

    return sp


@pytest.mark.dependency(name="list_files")
def test_list_files_from_sharepoint_folder(sharepoint_api: Sharepoint):
    """
    Test 1: Check that the expected Excel file exists in the SharePoint folder.
    """

    files = sharepoint_api.fetch_files_list(FOLDER_NAME)

    file_names = [f["Name"] for f in files]

    assert "Test_Append_Rows.xlsx" in file_names


@pytest.mark.dependency(name="append_row", depends=["list_files"])
def test_append_row_to_excel(sharepoint_api: Sharepoint):
    """
    Test 2: Append a row to the Excel file and verify it was added correctly.
    """

    file_name = "Test_Append_Rows.xlsx"

    sheet_name = "Upload logs"

    headers = ["Upload date", "Upload time"]

    data = {
        "Upload date": CURRENT_DATE,
        "Upload time": CURRENT_TIME
    }

    # Append the row to the SharePoint-hosted Excel file
    sharepoint_api.append_row_to_sharepoint_excel(
        required_headers=headers,
        folder_name=FOLDER_NAME,
        excel_file_name=file_name,
        sheet_name=sheet_name,
        new_row=data,
    )

    # Fetch the updated file and check that the last row matches the new data
    binary_file = sharepoint_api.fetch_file_using_open_binary(file_name, FOLDER_NAME)

    wb = openpyxl.load_workbook(BytesIO(binary_file))

    ws = wb[sheet_name]

    newest_row = list(ws.iter_rows(values_only=True))[-1]

    assert newest_row[0] == CURRENT_DATE
    assert newest_row[1] == CURRENT_TIME


@pytest.mark.dependency(depends=["append_row"])
def test_format_and_sort_excel_file(sharepoint_api: Sharepoint):
    """
    Test 3: Format and sort the Excel file by date and time.
    Ensures that sorting changes the order and formatting is applied.
    """

    file_name = "Test_Append_Rows.xlsx"

    sheet_name = "Upload logs"

    # Download the file before sorting to capture original row order
    test_file = sharepoint_api.fetch_file_using_open_binary(file_name, FOLDER_NAME)

    wb = openpyxl.load_workbook(BytesIO(test_file))

    ws = wb[sheet_name]

    all_rows = list(ws.iter_rows(values_only=True))

    top_before = all_rows[1]      # First data row

    bottom_before = all_rows[-1]  # Last data row

    # Sort by first two columns (date, time) in descending order
    sorting_keys = [
        {"key": "A", "ascending": False, "type": "str"},
        {"key": "B", "ascending": False, "type": "str"},
    ]

    # Apply sorting and formatting
    sharepoint_api.format_and_sort_excel_file(
        folder_name=FOLDER_NAME,
        excel_file_name=file_name,
        sheet_name=sheet_name,
        sorting_keys=sorting_keys,
        bold_rows=[1],
        column_widths="auto",
    )

    # Re-fetch the file after sorting and verify order has changed
    test_file = sharepoint_api.fetch_file_using_open_binary(file_name, FOLDER_NAME)

    wb = openpyxl.load_workbook(BytesIO(test_file))

    ws = wb[sheet_name]

    all_rows = list(ws.iter_rows(values_only=True))

    top_after = all_rows[1]

    bottom_after = all_rows[-1]

    # Check that sorting actually changed the row order
    assert top_before != top_after
    assert bottom_before != bottom_after

    # Check that the top row after sorting is the one just added
    assert top_after[0] == CURRENT_DATE
