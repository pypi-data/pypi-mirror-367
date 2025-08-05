print('Módulo Local: PLANO\nEste módulo contiene la ejecución del cálculo local de las hojas de Liquidez y Portafolio del PLANO para así subirlas a GCP.')

#-----------------------------------------------------------------
# Librerias

import pandas as pd
import numpy as np
import QuantLib as ql
import os
from datetime import datetime, timedelta, date
from openpyxl import load_workbook
#from tkinter import messagebox, Tk, Button
from scipy import optimize
import sys
import warnings

import blpapi

#adsa=input('hola y/n:')

# Para BigQuery
# pip install google-cloud-bigquery-storage
from google.cloud import bigquery

from .A_TRANSVERSAL import dicc_static,RGB_F,Riesgo_F,project,dataset_id,ideas_querys,festivos_manuales_PAN
from .A_TRANSVERSAL import fechas_relevantes,generate_paths,charge_serv,multiple_query,checker_fechas_cargue,upload_table,Formateo_df2GCP,all_path_dyncomp
# QuantLib, os, openpyxl, tkinter, scipy, sys, warnings, blpapi 

#-----------------------------------------------------------------
# Funciones a usar

def EXTRACCION_BG_HOY(cusips, fields, fecha_corte_flat, aki):

    """
    Esta función extrae los valores de BG para múltiples securities 
    a partir del CUSIP. Se admiten también múltiples fields, aunque solamente
    se utilizará uno. También la fecha de corte debe ser únicamente una sola
    fecha.

    Args:
    1. CUSIPS: Lista con los distintos cusips.
    2. fields: Lista con los fields con nomenclatura BG.
    3. fecha_corte_flat: Una sola fecha en formato 'YYYYMMDD'.
    4. aki: Se obtiene de APRE en BG, es la clave.


    Output:
    df para los securities que fueron encontrados.

    """
    
    sessionOptions = blpapi.SessionOptions()
    # AIK is currently optional
    sessionOptions.setApplicationIdentityKey(aki) # aki
    
    # Create a Session
    session = blpapi.Session(sessionOptions)
    
    # Start a Session
    if not session.start():
        print("Failed to start session.")
        return

    try:
        # Open service to get historical data from
        if not session.openService("//blp/refdata"):
            print("Failed to open //blp/refdata")
            return

        # Obtain previously opened service
        refDataService = session.getService("//blp/refdata")

        # Create and fill the request for the reference data
        service_BG = "Reference"

        full_service_req = f"{service_BG}DataRequest"
        request = refDataService.createRequest(full_service_req)

        # Define securities
        for cusip in cusips:

            value_s = f"/cusip/{cusip}"

            securities = request.getElement("securities")
            securities.appendValue(value_s)

        # Define fields
        for field in fields:
            request.getElement("fields").appendValue(field)

        # Optional - Apply overrides
        #overrides = request.getElement("overrides")

        #override1 = overrides.appendElement()
        #override1.setElement("fieldId","FUND_PER")
        #override1.setElement("value","Q")

        #override2 = overrides.appendElement()
        #override2.setElement("fieldId","EQY_FUND_CRNCY")
        #override2.setElement("value","EUR")

        #request.set("startDate", fecha_corte_flat)
        #request.set("endDate", fecha_corte_flat)
        
        # Send the request
        print("Sending Request:", request)
        session.sendRequest(request)
        
        data = []
        # Process received events
        while(True):
            ev = session.nextEvent()
            if ev.eventType() in [
                blpapi.Event.PARTIAL_RESPONSE,
                blpapi.Event.RESPONSE
            ]:
                for msg in ev:
                    if msg.hasElement("responseError"):
                        print(f"REQUEST FAILED: {msg.getElement('responseError')}")
                        continue
                
                    security_data_array = msg.getElement("securityData")
                    for i in range(security_data_array.numValues()):
                        security_data = security_data_array.getValueAsElement(i)
                        security_name = security_data.getElementAsString("security")

                        if security_data.hasElement("fieldData"):
                            field_data_array = security_data.getElement("fieldData")

                            date = fecha_corte_flat
                            row = {"Security": security_name, "Date": date}

                            for j in range(field_data_array.numElements()):
                                field_data = field_data_array.getElement(j)
                                
                                row[fields[j]] = field_data.getValue()
                                

                            data.append(row)
                        else: 
                            print(f"No field data available for {security_name}.")

            if ev.eventType() == blpapi.Event.RESPONSE:
                # Response completley received, so we could exit
                break

    finally:
        # Stop the session
        session.stop()

        
    df = pd.DataFrame(data)
    return df


def EXTRACCION_BG_HIST(cusips, fields, fecha_corte_flat, aki):

    """
    Esta función extrae los valores de BG para múltiples securities 
    a partir del CUSIP. Se admiten también múltiples fields, aunque solamente
    se utilizará uno. También la fecha de corte debe ser únicamente una sola
    fecha.

    Args:
    1. CUSIPS: Lista con los distintos cusips.
    2. fields: Lista con los fields con nomenclatura BG.
    3. fecha_corte_flat: Una sola fecha en formato 'YYYYMMDD'.
    4. aki: Se obtiene de APRE en BG, es la clave.


    Output:
    df para los securities que fueron encontrados.

    """
    
    sessionOptions = blpapi.SessionOptions()
    # AIK is currently optional
    sessionOptions.setApplicationIdentityKey(aki) # aki
    
    # Create a Session
    session = blpapi.Session(sessionOptions)
    
    # Start a Session
    if not session.start():
        print("Failed to start session.")
        return

    try:
        # Open service to get historical data from
        if not session.openService("//blp/refdata"):
            print("Failed to open //blp/refdata")
            return

        # Obtain previously opened service
        refDataService = session.getService("//blp/refdata")

        # Create and fill the request for the reference data
        service_BG = "Historical"


        full_service_req = f"{service_BG}DataRequest"
        request = refDataService.createRequest(full_service_req)

        # Define securities
        for cusip in cusips:

            value_s = f"/cusip/{cusip}"

            securities = request.getElement("securities")
            securities.appendValue(value_s)

        # Define fields
        for field in fields:
            request.getElement("fields").appendValue(field)

        # Optional - Apply overrides
        #overrides = request.getElement("overrides")

        #override1 = overrides.appendElement()
        #override1.setElement("fieldId","FUND_PER")
        #override1.setElement("value","Q")

        #override2 = overrides.appendElement()
        #override2.setElement("fieldId","EQY_FUND_CRNCY")
        #override2.setElement("value","EUR")

        request.set("startDate", fecha_corte_flat)
        request.set("endDate", fecha_corte_flat)
        
        # Send the request
        print("Sending Request:", request)
        session.sendRequest(request)
        
        data = []
        # Process received events
        while(True):
            ev = session.nextEvent()
            if ev.eventType() in [
                blpapi.Event.PARTIAL_RESPONSE,
                blpapi.Event.RESPONSE
            ]:
                for msg in ev:
                    if msg.hasElement("responseError"):
                        print(f"REQUEST FAILED: {msg.getElement('responseError')}")
                        continue
                
                    security_data = msg.getElement("securityData")
                    security_name = security_data.getElementAsString("security")
                    
                    if security_data.hasElement("fieldData"):
                        field_data_array = security_data.getElement("fieldData")
                        
                        for i in range(field_data_array.numValues()):
                            field_data = field_data_array.getValueAsElement(i)
                            date = field_data.getElementAsString("date")

                            row = {"Security": security_name, "Date": date}

                            for field in fields:
                                if field_data.hasElement(field):
                                    row[field] = field_data.getElementAsFloat(field)
                                else:
                                    row[field] = None

                            data.append(row)
                    else: 
                        print(f"No field data available for {security_name}.")

            if ev.eventType() == blpapi.Event.RESPONSE:
                # Response completley received, so we could exit
                break

    finally:
        # Stop the session
        session.stop()

        
    df = pd.DataFrame(data)
    return df

# def confirm_search(TOTAL_CONSULTAS_HOY):
            
#             message = f"{TOTAL_CONSULTAS_HOY} datapoints will be searched in Bloomberg.\n Do you want to proceed?"
#             response = messagebox.askyesno("Confirmar Búsqueda", message)

#             if response: 
#                 messagebox.showinfo("Search Accepted","Search will proceed")
#             else:
#                 messagebox.showwarning("Search denied", "Search has been canceled.")
#                 sys.exit()

# Función para formateo de fechas.
def fecha_timestamp_2_qldate(date_timestamp):
    return(ql.Date(date_timestamp.day, date_timestamp.month, date_timestamp.year))

# Función para formatear todas las fechas y ajustar frequency.
def prepare_bond_data(row, fecha_corte,base_map):
    issue_date = row['Fecha de emisión']
    maturity_date = row['Fecha de vencimiento']
    frequency_raw = row['Frecuencia']
    base_raw = row['Base']
    analysis_date = pd.to_datetime(fecha_corte)

    # Convertir las fechas al formato que espera Quantlib
    maturity_date = fecha_timestamp_2_qldate(maturity_date)
    issue_date    = fecha_timestamp_2_qldate(issue_date)
    analysis_date = fecha_timestamp_2_qldate(analysis_date)
    base_ql       = base_map.get(base_raw)

    if frequency_raw == 1:
         frequency_ql  = ql.Annual
    elif frequency_raw == 2:
         frequency_ql = ql.Semiannual
    elif frequency_raw == 4:
         frequency_ql = ql.Quarterly
    elif frequency_raw == 12:
         frequency_ql = ql.Monthly

    return maturity_date, issue_date, analysis_date, frequency_ql, base_ql

def maca_dur(issue_date, maturity_date, analysis_date, coupon_rate, frequency_ql, notional, bond_yield, day_count):

    #issue_date = ql.Date(issue_date.day, issue_date.month, issue_date.year)
    #maturity_date = ql.Date(maturity_date.day, maturity_date.month, maturity_date.year)
    #analysis_date = ql.Date(analysis_date.day, analysis_date.month, analysis_date.year)

    calendar = ql.UnitedStates(ql.UnitedStates.NYSE)

    schedule = ql.Schedule(issue_date, maturity_date, ql.Period(frequency_ql), calendar,
                        ql.Unadjusted, ql.Unadjusted, ql.DateGeneration.Backward, False)

    bond = ql.FixedRateBond(1, notional, schedule, [coupon_rate], day_count)

    # Keep cashflows that are after analysis_date
    next_cashflows = [cf for cf in bond.cashflows() if cf.date() > analysis_date]
    
    cashflows_amounts = np.array([cf.amount() for cf in next_cashflows])
    cashflow_dates = [cf.date() for cf in next_cashflows]
    # Years from analysis to cf.
    years_to_cf = np.array([day_count.yearFraction(analysis_date, date) for date in cashflow_dates])

    # Define discount curve as flat (IRR).
    compounding = ql.Compounded
    discount_curve = ql.YieldTermStructureHandle(
        ql.FlatForward(analysis_date, bond_yield, day_count, compounding, frequency_ql)
    )

    discount_factors = np.array([discount_curve.discount(date) for date in cashflow_dates])
    discounted_cashflows = cashflows_amounts * discount_factors

    # Calculation of MACA duration.
    numerator = np.sum(years_to_cf * discounted_cashflows)
    VP = np.sum(discounted_cashflows)

    # Calculate clean price
    accrued_interest_base100 = bond.accruedAmount(analysis_date)
    accrued_interest = accrued_interest_base100 * (notional/100)
    VP = VP - accrued_interest # Holdings viene con el precio limpio.

    maca_dur = numerator / VP
    return maca_dur


def df_dur_maca(df, fecha_corte,base_map):

    results = []

    for _, row in df.iterrows():

        maturity_date, issue_date, analysis_date, frequency_ql, day_count = prepare_bond_data(row, fecha_corte,base_map)
        coupon_rate = row['Facial']
        notional = row['Nominal']
        bond_yield = row['Yield']
        

        dur_MACA = maca_dur(issue_date, maturity_date, analysis_date, coupon_rate,
                             frequency_ql, notional, bond_yield, day_count)
        
        results.append(dur_MACA)
    
    return results

def VP_yield(bond_yield, issue_date, maturity_date, analysis_date, coupon_rate, frequency_ql,
             notional, day_count):

    ql.Settings.instance().evaluationDate = analysis_date
    calendar = ql.UnitedStates(ql.UnitedStates.NYSE)

    schedule = ql.Schedule(issue_date, maturity_date, ql.Period(frequency_ql), calendar,
                        ql.Unadjusted, ql.Unadjusted, ql.DateGeneration.Backward, False)

    bond = ql.FixedRateBond(1, notional, schedule, [coupon_rate], day_count)

    # Keep cashflows that are after analysis_date
    next_cashflows = [cf for cf in bond.cashflows() if cf.date() > analysis_date]
    
    cashflows_amounts = np.array([cf.amount() for cf in next_cashflows])
    cashflow_dates = [cf.date() for cf in next_cashflows]

    # Define discount curve as flat (IRR).
    compounding = ql.Compounded
    discount_curve = ql.YieldTermStructureHandle(
        ql.FlatForward(analysis_date, bond_yield, day_count, compounding, frequency_ql)
    )

    discount_factors = np.array([discount_curve.discount(date) for date in cashflow_dates])
    discounted_cashflows = cashflows_amounts * discount_factors
    VP = np.sum(discounted_cashflows)

    # Calculate clean price
    accrued_interest_base100 = bond.accruedAmount(analysis_date)
    accrued_interest = accrued_interest_base100 * (notional/100)
    VP = VP - accrued_interest # Holdings viene con el precio limpio.

    VP = (VP/ notional) * 100

    return VP

def objec_f_irr(bond_yield, bond_clean_price, issue_date, maturity_date, analysis_date, coupon_rate, frequency_ql,
                notional, day_count):

    dirty_VP = VP_yield(bond_yield, issue_date, maturity_date, analysis_date, coupon_rate, frequency_ql,
                          notional, day_count)

    return dirty_VP - bond_clean_price

def IRR(bond_clean_price, issue_date, maturity_date, analysis_date, coupon_rate, frequency_ql,
                notional, day_count):

    result_irr = optimize.root_scalar(objec_f_irr, bracket = [-1, 1], method = 'brentq',
                                      args = (bond_clean_price, issue_date, maturity_date, 
                                              analysis_date, coupon_rate, frequency_ql,
                                              notional, day_count),
                                      xtol= 1e-8)
    if result_irr.converged:
        return result_irr.root
    else:
        raise ValueError("IRR calculation did not converge")


def df_IRR(df, fecha_corte,base_map):
        
        results = []

        for _, row in df.iterrows():


            maturity_date, issue_date, analysis_date, frequency_ql, day_count = prepare_bond_data(row, fecha_corte,base_map)
            coupon_rate = row['Facial']
            notional = row['Nominal']
            bond_clean_price = row['PRECIO']
            

            yield_C = IRR(bond_clean_price, issue_date, maturity_date, analysis_date, coupon_rate, frequency_ql,
                          notional, day_count)
            
            results.append(yield_C)
        return results

def verify_dataframe(df):
    issues = []  # List to store error messages

    # 2. Check 'Frecuencia' contains only integers in {1, 2, 4, 12}
    if 'Frecuencia' in df.columns:
        invalid_freq = df[~df['Frecuencia'].isin([1, 2, 4, 12])]
        if not invalid_freq.empty:
            issues.append(f"'Frecuencia' has {len(invalid_freq)} invalid values (must be 1, 2, 4, or 12).")

    # 3. Check 'Facial' is a float and < 1
    if 'Facial' in df.columns:
        invalid_facial = df[(df['Facial'].apply(lambda x: isinstance(x, (int, float))) == False) | (df['Facial'] >= 1)]
        if not invalid_facial.empty:
            issues.append(f"'Facial' has {len(invalid_facial)} invalid values (not float or >= 1).")

    # 4. Check 'PRECIO' is a float and < 1000
    if 'PRECIO' in df.columns:
        invalid_precio = df[(df['PRECIO'].apply(lambda x: isinstance(x, (int, float))) == False) | (df['PRECIO'] >= 1000)]
        if not invalid_precio.empty:
            issues.append(f"'PRECIO' has {len(invalid_precio)} invalid values (not float or >= 1000).")

    # 5. Check 'Base' contains only '360', '365', 'ACT/ACT', '30/360', 'ACT/365', 'ACT/360'
    if 'Base' in df.columns:
        valid_base_values = {'360', '365', 'ACT/ACT', '30/360', 'ACT/365', 'ACT/360',"ISMA-30/360"}
        invalid_base = df[~df['Base'].isin(valid_base_values)]
        if not invalid_base.empty:
            issues.append(f"'Base' has {len(invalid_base)} invalid values (must be one of {', '.join(valid_base_values)}).")

    # Raise an error if any issues found
    if issues:
        error_message = "\n".join(issues)
        raise ValueError(f"Data verification failed with the following issues:\n{error_message}")
    else:
        print("No issue found on Renta Fija Holdings, everything good to calculate modified duration")

def verify_manual(df):
    if 'PX_DIRTY_BID' in df.columns:
        invalid_values = df[(df['PX_DIRTY_BID'].apply(lambda x: isinstance(x, (int, float))) == False) | (df['PX_DIRTY_BID'] < 0) | (df['PX_DIRTY_BID'] > 1000)]
        if not invalid_values.empty:
            raise ValueError(f"'PX_DIRTY_BID' has {len(invalid_values)} invalid values (must be numerical and between 0 and 1000).")
        else:
            print("No issue found in 'PX_DIRTY_BID', ready to use for modified duration calculations.")
    else:
        raise ValueError("Column 'PX_DIRTY_BID' not found in the DataFrame.")


def global_treatment_Plano_Pan(fecha:str=None,where_to_run:str = 'local'):
    # Cargue objetos estáticos
    month_map, categorias, Type_out, pp_titulosmanuales, NAC_REG = (dicc_static[k] for k in ('dicc_month_map', 'tipo_activos_holdings', 'tipo_producto_plano',
                                                                                         'pp_titulosmanuales', 'dicc_nac_2_cont'))
    # Cargue de Fechas
    #fecha = None
    fechas = fechas_relevantes(pais = "PAN", fecha_analisis= fecha,festivos_manuales_PAN= festivos_manuales_PAN)
    fecha_corte,fecha_corte_ayer,fecha_corte_d = [fechas[k] for k in ['f_corte','f_corte_ayer','f_corte_d']]
    # Diccionario con todos los path
    all_paths = generate_paths(fecha_corte= fecha_corte, fecha_corte_ayer=fecha_corte_ayer, RGB_F=RGB_F, Riesgo_F=Riesgo_F, month_map=month_map)

    # Desagregación de Paths
    data_pershing_path, manual_data_path, AKI_path, Calificaciones_path, ISIN_path, TRM_path, Saldos_path, folder_salida, nombre_salida, path_BG = (all_paths[k] for k in ('pershing','manual_data_plano','AKI_BG','calificaciones','ISIN','TRM','saldos_liq','output_folder_plano','nombre_plano','output_BG'))
    
    fecha_corte_flat = all_path_dyncomp(fecha_corte, fecha_corte_ayer, RGB_F, Riesgo_F, month_map)['fecha_corte_flat']
    
    # Verificar que la carpeta de salida existe
    if not os.path.exists(folder_salida):
        os.makedirs(folder_salida)
        print('Se ha creado la carpeta: {}'.format(folder_salida))

    path_salida = rf"{folder_salida}/{nombre_salida}"                                                                                                                                                                                            
                                                                                                                                                                                                        
    #-----------------------------------------------------------------
    # Importe de Inputs Manuales

    # Input 1
    # Lectura de Definición de Portafolios.
    # Este input lleva registro de todos los ADPT y PP. Se debe modificar cada que hay un nuevo cliente para ADTP.
    Porta_df = pd.read_excel(manual_data_path, sheet_name="Input", usecols=list(range(5)), skiprows= 1, dtype={'NUMERO':str})
    Porta_df = Porta_df.dropna(how = "all")

    # Input 2
    # Títulos que no se encuentran en Bloomberg (Incorpora Formateo de Fechas).
    # Este input se debe realizar para cada uno de los activos vigentes que no se encuentre en Bloomberg. (e.g. Bladex).

    Titulos_df = pd.read_excel(manual_data_path, sheet_name="Input", usecols=list(range(6,25)), skiprows= 1)
    Titulos_df = Titulos_df.iloc[:Titulos_df.isnull().any(axis = 1).idxmax()]

    col_dates = ['Fecha de emisión','Fecha de vencimiento','Fecha de compra']
    for col in col_dates:
        Titulos_df[col] = pd.to_datetime(Titulos_df[col], errors= 'coerce')
    Titulos_df['Frecuencia'] = Titulos_df['Frecuencia'].astype(int)
    Titulos_df['Base'] = Titulos_df['Base'].astype(int).astype(str)
        
    # Input 3 
    # Lectura de Saldos Cuentas
    # Se debe actualizar cuando Contabilidad envíe una nueva cuenta de ahorros.
    Saldosc_df = pd.read_excel(manual_data_path, sheet_name="Input", usecols=list(range(26,28)), skiprows= 1, 
                            dtype={'ID_Cuentas':str})
    Saldosc_df = Saldosc_df.iloc[:Saldosc_df.isnull().any(axis = 1).idxmax()]

    # Input 4 
    # Calificaciones
    Calif_RF = pd.read_excel(manual_data_path, sheet_name="Input", usecols=list(range(29,31)), skiprows= 1, dtype={'NUMERO':str})

    #-----------------------------------------------------------------
    # Warmimgs de Inputs Manuales

    # Verificar el primer input de portafolio
    for col in Porta_df.columns:
        missing_rows = Porta_df[((Porta_df)[col].isna()) & (Porta_df['PORTAFOLIO'] != "POSICION PROPIA") & (Porta_df["VIGENTE"] != 0)].index.tolist()

        if missing_rows:
            raise ValueError(f"Porta_df: Column '{col}' has missing values in rows: {missing_rows}")

    # Verificar Titutlos y Saldos
    dfs = [Titulos_df, Saldosc_df]
    df_names = ["Títulos Manuales", "Saldos"]

    # Check for missing values
    for i, df in enumerate(dfs):
        for col in df.columns:
                
            missing_rows = df[df[col].isna()].index.tolist()
        
        if missing_rows:
            raise ValueError(f"{df_names[i]}: Column '{col}' has missing values in rows: {missing_rows}")
        


    #-----------------------------------------------------------------
    # Importe y Limpieza de Data Pershig (HOLDINGS) 

    holdings = pd.read_excel(data_pershing_path, skiprows=6, engine= "openpyxl")
    # Drop excesses on tail.
    excesos_tail_row = holdings[holdings['Account'] == "TOTAL"].index[0]
    holdings = holdings.loc[:excesos_tail_row - 1]

    # Holdings tiene unas filas que deben ser eliminadas para garantizar que la info sea un cuadrado limpio.
    # Estas filas clasifican los valores y son:
    categorias = ['CASH AND EQUIVALENTS','EQUITY','ETF','FIXED INCOME', 'MUTUAL FUND']
    filas_separadoras = holdings[holdings['Account'].isin(categorias)].index
    # Append last row de modo que podamos separar el bloque final.
    num_row_Persh = holdings.shape[0]
    filas_separadoras = filas_separadoras.append(pd.Index([num_row_Persh])).to_list()

    # Se debe entonces eliminar estas filas que separan e incluirlas como característica en la 1era columna.
    holdings['Type'] = None
    holdings.insert(0, 'Type', holdings.pop('Type'))
    # Esta será la columna a incluir en Holdings.
    Type_out = ['Cash', 'Equity', 'ETF', 'FixedIncome', 'MutualFund'] 

    # Se eliminan las filas separadoras y se incluye una primera columna de tipo de activo que reemplaza lo eliminado.
    for i in range(len(filas_separadoras) - 1):
        start_idx = filas_separadoras[i]
        final_idx = filas_separadoras[i+1]
        holdings.loc[start_idx:final_idx, 'Type'] = Type_out[i]

    holdings = holdings.drop(filas_separadoras[:-1], axis = 0)
    holdings = holdings.reset_index(drop=True)

    # Se eliminan aquellos títulos que no se encuentran en Bloomberg. La información de estos títulos será introducida
    # manualmente con los inputs previos. Estos títulos en este momento se entiende como los pertenecientes al Portafolio
    # PXG898019 y cuyo Security Name no sea Cash.

    holdings = holdings[~((holdings['Account'].isin(pp_titulosmanuales)) & (holdings['Security Name'] != 'CASH'))]
    #holdings.shape

    #-----------------------------------------------------------------
    # Inputs de Base de Datos (Son 6)

    # Input 1: Calificaciones de riesgo
    calificaciones = pd.read_excel(Calificaciones_path)
    calificaciones = calificaciones.loc[:, : "Clasificación"]

    # Input 2: ISINES: los fundamentales de un activo.
    isines = pd.read_excel(ISIN_path, sheet_name="ISINES", dtype={'DAY_COUNT':str})
    isines['CODIGOPERIODO'] = isines['CODIGOPERIODO'].astype('Int64')
    isines['CUSIP'] = (
        isines['CUSIP']
        .astype(str)
        .str.strip()
        .replace({'': np.nan, 'nan': np.nan, 'NaN': np.nan})
    )

    isines = isines[isines['CUSIP'].notna()]

    # Input 3 
    # Traer la TRM
    TRM_data = pd.read_excel(TRM_path, engine='xlrd')
    TRM = TRM_data[TRM_data.iloc[:,3].isin(['TRM','USD'])].iloc[0,4]
    print("La TRM cargada es:",TRM)

    # Input 4
    # Traer los saldos para liquidez.
    saldos_L = pd.read_excel(Saldos_path, skiprows= 5, dtype={'CUENTA':str})

    #-----------------------------------------------------------------
    # Importe de AKI para BG

    with open(AKI_path, "r") as file:
        aki = file.readline().strip()

    #-----------------------------------------------------------------
    # Carpintería: Creación de Output (Hoja Portafolio)

    # Traer los portafolios vigentes
    cuentas_vigentes = Porta_df.loc[Porta_df['VIGENTE'] == 1, 'CUENTA'].tolist()
    # Vamos a identificar cuales filas dentro de holdings cumplen las condiciones para la hoja de portafolio.
    # Para ello se deben verificar 3 condiciones: La cuenta asociada coincide con una cuenta vigente, no es Cash y hay Trade.
    filas_portafolio = holdings[
        (holdings['Account'].isin(cuentas_vigentes)) & 
        (holdings['Type'] != 'Cash') & 
        (holdings['Trade Date Quantity'] != 0)
    ]
    filas_portafolio = filas_portafolio.reset_index(drop = False)
    # Formateo en caso de ser necesario
    filas_portafolio.loc[:,'CUSIP'] = filas_portafolio['CUSIP'].astype(str)

    # Número de activos a incluir dentro de Portafolio
    # Esto será la suma de los que vienen de Holdings y los que se introducen manualmente.
    titulos_corte_Port = len(filas_portafolio) + len(Titulos_df)

    len(filas_portafolio)


    Portafolio = pd.DataFrame(columns= ['FECHAPORTAFOLIO', 'CLA', 'FTE', 'ESPECIE',	'SERIE', 'EMISOR',
                                        'TITULO', 'TC', 'CUSIP', 'ISIN', 'CALIF', 'COD', 'FECHAEMISION',
                                        'FECHAVENCIMIENTO',	'CANTIDADVALORNOMINAL',	'FACIAL', 'CODIGOPERIODO',
                                        'F_COMPRA', 'VALOR_COMPRA_ORIGEN', 'MONEDA_ORIGEN', 'BASEMONETARIA',
                                        'VALOR_COMPRA', 'T_COMPRA',	'VALOR_TIR', 'TIR', 'VALORMERCADO',	'T_MER',
                                        'DIAS',	'MET',	'PRECIO',	'TASA',	'VR_TASA',	'MARGEN',	'T_DCTO',
                                        'ES', 'FG', 'FV', 'POR', 'VR_MDO_AYER', 'INGRESO_VALORACION',
                                        'CAUSACION',	'DOS_MENOS_UNO', 'EFECTO_PATRIMONIAL', 'DURACION_MODIFICADA',
                                        'CONVEXIDAD', 'NOMINAL_PESOS',	'CLASIFICACION1', 'CLASIFICACION2',
                                        'GRUPO_TASA', 'TIPO_TASA', 'EMISOR2', 'SECTOR', 'NACIONALIDAD',	
                                        'TIPO_EMISOR', 'GRUPO_FINANCIERO',	'PORTAFOLIO', 'EMPRESA', 'ID_TITULO', 'CUPON_REPRECIO',
                                        'DV1', 'ASSET_CLASS', 'CODIGOTIPOOPERACIONPORTAFOLIO', 'NIT_PORTAFOLIO',
                                        'TIPO_ACTIVO',	'TIPO_DE_PRODUCTO',	'DETALLE_PRODUCTO',	'CLASIF_RF', 'REGION_PERFIL',
                                        'REGION','PERFIL','DUR_MACA', 'IF_TRADE'
                                        ], index = range(titulos_corte_Port))
    



    #-----------------------------------------------------------------
    # Variables Estáticas

    # Fecha de hoy
    Portafolio['FECHAPORTAFOLIO'] = fecha_corte_d
    # CLA 
    Portafolio['CLA'] = "1CP"
    # Base Monetaria
    Portafolio['BASEMONETARIA'] = "DOLAR"
    # MET
    Portafolio['MET'] = "1" 
    # Empresa
    Portafolio['EMPRESA'] = "DAVIVIENDA CORREDORES PANAMA"
    # Código Tipo Operación Portafolio es un tipo de Inversión I.
    Portafolio['CODIGOTIPOOPERACIONPORTAFOLIO'] = "I"

    #-----------------------------------------------------------------
    # Variables Dinámicas que se traen directamente de Holdings

    # ESPECIE
    ESPECIE = filas_portafolio['CUSIP']
    # Titulo
    TITULO = filas_portafolio['Security Name']
    # CUSIP
    CUSIP = filas_portafolio['CUSIP']
    # Volumen
    CANTIDAD_VAL_NOM = filas_portafolio['Trade Date Quantity']
    # Valor Mercado
    VAL_MERCADO = filas_portafolio['Trade Date Market Value']
    # Nominal Pesos
    NOMINAL_COP = filas_portafolio['Trade Date Quantity'] * TRM


    #-----------------------------------------------------------------
    # Variables VLOOKUP

    # Emisor
    EMISOR = filas_portafolio['CUSIP'].map(isines.set_index('CUSIP')['EMISOR'])

    # ISIN
    ISIN = filas_portafolio['CUSIP'].map(isines.set_index('CUSIP')['ISIN'])

    # Fecha de emisión
    FEC_EMIS = filas_portafolio['CUSIP'].map(isines.set_index('CUSIP')['FECHAEMISION'])
    FEC_EMIS = pd.to_datetime(FEC_EMIS, errors= 'coerce', dayfirst=True)

    # Fecha de Vencimiento
    FEC_VENC = filas_portafolio['CUSIP'].map(isines.set_index('CUSIP')['FECHAVENCIMIENTO'])
    FEC_VENC = pd.to_datetime(FEC_VENC, errors= 'coerce', dayfirst=True)

    # Facial - Tasa cupón # Hay un error de los datos de Coeasy.
    # Las unidades se expresan en varios órdenes de magnitud, con lo cual lo mejor que se
    # puede hacer es dividir por 100 si es mayor a 1 y por 10000 si es mayor a 100.
    FACIAL = filas_portafolio['CUSIP'].map(isines.set_index('CUSIP')['FACIAL'])

    # Frecuencia de los cupones - CODIGOPERIODO
    # # Primero Vlookup luego mapeo de strings
    FREQ_CUPON = filas_portafolio['CUSIP'].map(isines.set_index('CUSIP')['CODIGOPERIODO'])

    # Moneda Origen
    MON_ORIGEN = filas_portafolio['CUSIP'].map(isines.set_index('CUSIP')['MONEDA_ORIGEN'])

    # Días hasta maturity
    DIAS = (FEC_VENC - pd.to_datetime(fecha_corte)).dt.days

    # Precio = Valor Mercado / Nocional
    # Debemos hacer la distinción de aquellas entradas que son Fixed Income.
    # El precio de Fixed Income por convención se dispone por 100.
    FIC_r = filas_portafolio.index[filas_portafolio['Type'] == 'FixedIncome'].tolist()
    PRECIO = filas_portafolio['Trade Date Market Value'] / filas_portafolio['Trade Date Quantity']
    PRECIO.loc[FIC_r] *= 100

    # POR Identifica el número del portafolio y si es Fixed Income. 
    # Mezclamos el número del portafolio con el identificador.
    NUMERO = filas_portafolio['Account'].map(Porta_df.set_index('CUENTA')['NUMERO'])
    IDENT = np.array(["-I2"] * len(NUMERO))
    IDENT[FIC_r] = "-I1"
    POR = list(map(lambda x, y: f"{x}{y}", NUMERO, IDENT))

    # BASE
    BASE = filas_portafolio['CUSIP'].map(isines.set_index('CUSIP')['DAY_COUNT'])

    # Clasificación2 VAR
    CLAS2 = np.array(["ACCIONES"] * len(filas_portafolio))
    CLAS2[FIC_r] = "YANKEES_USD"

    # Sector
    SECTOR = EMISOR.map(isines.drop_duplicates(subset= 'EMISOR').set_index('EMISOR')['SECTOR'])

    # Nacionalidad
    NACIONALIDAD = EMISOR.map(isines.drop_duplicates(subset= 'EMISOR').set_index('EMISOR')['NACIONALIDAD'])

    # Portafolio
    PORTAFOLIO = filas_portafolio['Account'].map(Porta_df.set_index('CUENTA')['PORTAFOLIO'])

    # Asset Class
    ASSET_CLASS = filas_portafolio['CUSIP'].map(isines.set_index('CUSIP')['ASSET_CLASS'])

    # NIT Portafolio
    NIT_P = filas_portafolio['Account']

    # Tipo de Activo
    TIPO_ACTIVO = pd.Series(filas_portafolio['CUSIP'].map(isines.set_index('CUSIP')['TIPO_ACTIVO']))

    # Tipo de Producto
    TIPO_PRODUCTO = filas_portafolio['Type']

    # Detalle Producto
    # Se define detalle producto a partir de TIPO_DE_PRODUCTO. Si el valor es inexistente,
    # se rellena con el código ISINES.
    DETALLE_PRODUCTO = filas_portafolio['CUSIP'].map(isines.set_index('CUSIP')['DETALLE PRODUCTO'])

    # Region
    # Ya se cuenta con Nacionalidad y se mapea hacia Region con:
    NAC_REG = {"ESTADOS UNIDOS": "EEUU", "LUXEMBURGO": "EUROPA", "GRAND CAYMAN": "EUROPA",
            "ALEMANIA": "EUROPA", "NEDERLAND": "EUROPA", "MEXICO": "LATINOAMERICA", 
            "PERU": "LATINOAMERICA", "COLOMBIA": "LATINOAMERICA", "SPAIN": "EUROPA",
            "FRANCIA": "EUROPA", "PANAMA": "LATINOAMERICA", "IRELAND": "EUROPA", 
            "IRLANDA": "EUROPA"}

    REGION = NACIONALIDAD.map(NAC_REG)

    # Perfil de Riesgo
    PERFIL = filas_portafolio['Account'].map(Porta_df.set_index('CUENTA')['PERFIL'])

    # CALIF de Riesgo
    CALIF_CERT = ISIN.map(calificaciones.drop_duplicates(subset='ISIN').set_index('ISIN')['Calificación'])
    FIL_SINC_rows = TIPO_PRODUCTO[(TIPO_PRODUCTO == "MutualFund") | (TIPO_PRODUCTO == "ETF")].index
    FIL_acc_rows = np.where((np.array(TIPO_PRODUCTO) == "Equity") | (np.array(DETALLE_PRODUCTO) == "Equity"))
    CALIF_CERT[FIL_SINC_rows] = "SIN CALIFICACION"
    CALIF_CERT[FIL_acc_rows] = "ACCIONES"

    # CALIF Clasificación binaria de renta fija según el código de las calificadoras.
    CODIGO_CALIF = CALIF_CERT.map(Calif_RF.drop_duplicates(subset = 'Calif').set_index('Calif')['Codigo'])
    CLASIF_RF = pd.Series(np.where((CODIGO_CALIF >= 1) & (CODIGO_CALIF <= 14), "Especulativo",
                        np.where(CODIGO_CALIF > 14, "Grado de Inversión", CODIGO_CALIF)))
    CLASIF_RF.replace(to_replace='nan', value= np.nan, inplace=True)

    # Si el título se holdea hasta el vencimiento
    IF_TRADE = np.array([1] * len(filas_portafolio))

    #-----------------------------------------------------------------
    # Variables BLOOMBERG


    #-----------------------------------------------------------------
    # Extraccion desde Bloomberg

    # Vamos a identificar los títulos que requieren del cálculo de duración modificada
    # Estos son los de renta fija y eliminamos duplicados con CUSIP.
    index_not_dup_CUSIP = CUSIP.drop_duplicates(keep='first').index

    index_rf_positions = index_not_dup_CUSIP[TIPO_PRODUCTO[index_not_dup_CUSIP] == 'FixedIncome']

    # Hay un detalle y es que los BLADEX son personalizados por lo que esta información no se encuentra
    # registrada en BG, también deben omitirse.
    index_BLADEX = TITULO[TITULO.str.contains('BANCO LATINOAMERICANO DE COMERCIO EXTERIOR', case=False, na=False)].index

    # Final index to send to BG.
    index_rf_BG = index_rf_positions.difference(index_BLADEX) 

    # Este dataset contiene la información fundamental que se requiere para el cálculo de DM
    # y que se buscará en BG para los títulos viejos.

    column_names = ['Fecha de emisión', 'Fecha de vencimiento', 'Frecuencia', 'Base', 'Facial']
    variables_v = [FEC_EMIS, FEC_VENC, FREQ_CUPON, BASE, FACIAL]
    df_DM_Holdings = pd.DataFrame(dict(zip(column_names, variables_v)))
    df_DM_Holdings = df_DM_Holdings.iloc[index_rf_BG]

    missing_value_rf_index = df_DM_Holdings[df_DM_Holdings.isna().any(axis=1)].index
    Consulta_Hold = missing_value_rf_index.intersection(index_rf_BG)

    # Consultas para los títulos manuales
    Consulta_P_Manual = Titulos_df[Titulos_df['IF_BMG'] == 1].index
    CUSIP_BMG_M = Titulos_df.loc[Titulos_df.index[Consulta_P_Manual], 'ID'].tolist()
    CUSIP_BMG_CLEAN_M = set(CUSIP_BMG_M)


    TOTAL_CONSULTAS_HOY = len(Consulta_Hold)*5 + len(CUSIP_BMG_CLEAN_M)
    print(f"Consultas (datapoints) requeridas para {fecha_corte}: {TOTAL_CONSULTAS_HOY}")

    #-----------------------------------------------------------------
    # Encapsulación de BG

    # Este pedazo del código está encapsulado. La protección es que si el archivo de Excel con la data de BG
    # ya fue creado, entonces aparece la alerta y no se realiza la consulta. 
    # En caso contrario si se realiza la consulta de BG.

    if os.path.isfile(path_BG):
        print("La información de BG ya se encuentra en {path_BG}, la información ya ha sido extraída para la fecha de corte:{fecha_corte}")
    else:
        # Explicación:    
        # La extracción desde Bloomberg se realiza distino para los títulos de Pershing y para los manuales.
        # En el caso de Pershing necesitamos la DM para calcular el DV01 (La principal razón es que el precio
        # de pershing viene bueno pero la duración no se puede calcular puesto que la Tasa Cupón y la frecuencia,
        # no viene muy limpia. # En cambio para los manuales como se introduce manualmente todos los valores relevantes
        # entonces nos basta con extraer el precio.

        #################################
        #################################
        # BMG Holdings
        # BMG Holdings
        # BMG Holdings
        # BMG Holdings
        #################################
        #################################

        # This is the last security check, user must accept search and know how many datapoints will be extracted.
        print(f"{TOTAL_CONSULTAS_HOY} datapoints will be searched in Bloomberg")
        response = input("Do you want to continue (yes/no): ").lower()

        if response == 'no':
            sys.exit('Se cancela la corrida.')

        if not Consulta_Hold.empty:

            ###################################
            # Consulta #1: Holdings
            ###################################

            # Se extraen los valores para la consulta y se verifica que no hayan securities repetidos.
            CUSIP_BMG_CLEAN = CUSIP[Consulta_Hold].tolist()

            #Valores de consulta en BG
            cusips = CUSIP_BMG_CLEAN
            fields = ['ISSUE_DT','MATURITY','DAY_CNT_DES','CPN','CPN_FREQ']
            aki = aki

            BG_Holdings = EXTRACCION_BG_HOY(cusips = cusips, fields = fields, fecha_corte_flat = fecha_corte_flat,
                                                aki = aki)

            n_sec_matches_hold = BG_Holdings.shape[0]
            # Guardar la información en un csv
            BG_Holdings.to_excel(path_BG, sheet_name = "BG_Holdings", index = False)
            print(f"La información de holdings ha sido correctamente guardada en {path_BG}, con {n_sec_matches_hold} security matches.")

        else:
            pass

        #################################
        #################################
        # BMG Manual
        # BMG Manual
        # BMG Manual
        # BMG Manual
        #################################
        #################################
        

        # Se extraen los valores para la consulta y se verifica que no hayan securities repetidos.
        CUSIP_BMG_M = Titulos_df.loc[Titulos_df.index[Consulta_P_Manual], 'ID'].tolist()
        CUSIP_BMG_CLEAN_M = set(CUSIP_BMG_M)

        # Valores de consulta en BG
        cusips = CUSIP_BMG_CLEAN_M
        fields = ['PX_DIRTY_BID']
        aki = aki

        if len(cusips) > len(Consulta_P_Manual): 
            print("El número de securities a consultar es superior al total de títulos Manuales con BG.")

        BG_Manual = EXTRACCION_BG_HIST(cusips = cusips, fields = fields, fecha_corte_flat = fecha_corte_flat,
                                       aki = aki)


        n_sec_matches_M = BG_Manual.shape[0]

        # Guardar la información en un Excel
        if os.path.exists(path_BG): 
            # Load the existing Excel file

            with pd.ExcelWriter(path_BG, engine='openpyxl', mode='a') as writer:
                BG_Manual.to_excel(writer, sheet_name="BG_Manual", index = False)
        else:
            BG_Manual.to_excel(path_BG, sheet_name="BG_Manual", index = False)   


        if n_sec_matches_M == 0:
            raise Warning("Warning: No se ha recuperado información de Bloomberg para títulos manuales, vacío. !\n",
                          "Se debe actualizar manualmente el precio (utilizando el del día anterior seguramente) de los siguientes !\n"
                          f"securities {CUSIP_BMG_CLEAN_M}")
        else:
            print(f"La información de los títulos manuales ha sido correctamente guardada en {path_BG}, con {n_sec_matches_M} security matches.")


    #-----------------------------------------------------------------
    # Lectura de Info de BG para Holdings

    if not Consulta_Hold.empty:

        BMG_data_Hold = pd.read_excel(path_BG, sheet_name= "BG_Holdings")
        BMG_data_Hold['Security'] = BMG_data_Hold['Security'].str.replace('/cusip/', '', regex=False)
        BMG_data_Hold = BMG_data_Hold.rename(columns = {'Security': 'CUSIP'})
        BMG_data_Hold['CPN'] = BMG_data_Hold['CPN']/100 # Cupon en términos no porcentuales.
        BMG_data_Hold['CPN_FREQ'] = pd.to_numeric(BMG_data_Hold['CPN_FREQ'], errors='coerce').astype('Int64') # Mantener la frecuencia como un entero.
        # Mantener el formato string
        #str_map_m = {0: 'CeroCupon', 1: 'AV', 2:'SV', 4:'TV'}
        #BMG_data_Hold['CPN_FREQ'] = BMG_data_Hold['CPN_FREQ'].map(str_map_m)

    #-----------------------------------------------------------------
    # Funciones para Duracion de Macaulay y Yield

    # El primero es para garantizar un ACT/360 el día llegue una base con formato erróneo,
    base_map = {"360": ql.Thirty360(ql.Thirty360.ISMA),
                "365": ql.Actual365Fixed(ql.Actual365Fixed.NoLeap),
                "ACT/ACT": ql.ActualActual(ql.ActualActual.ISMA),
                "30/360": ql.Thirty360(ql.Thirty360.ISMA), 
                "ACT/365": ql.Actual365Fixed(ql.Actual365Fixed.NoLeap),
                "ACT/360": ql.Actual360(),
                "ISMA-30/360": ql.Thirty360(ql.Thirty360.ISMA)}

    #-----------------------------------------------------------------
    # Imcorporacion de Fundamentales de BG para DM y DV01

    # Se incorporan las variables fundamentales traidas de BG.
    # Se debe encapsular por si el número de títulos a buscar es cero.

    if not Consulta_Hold.empty:

        FEC_EMIS = CUSIP.map(BMG_data_Hold.set_index('CUSIP')['ISSUE_DT']).fillna(FEC_EMIS)
        FEC_VENC = CUSIP.map(BMG_data_Hold.set_index('CUSIP')['MATURITY']).fillna(FEC_VENC)
        BASE = CUSIP.map(BMG_data_Hold.set_index('CUSIP')['DAY_CNT_DES']).fillna(BASE)
        FACIAL = CUSIP.map(BMG_data_Hold.set_index('CUSIP')['CPN']).fillna(FACIAL)   
        FREQ_CUPON = CUSIP.map(BMG_data_Hold.set_index('CUSIP')['CPN_FREQ']).fillna(FREQ_CUPON)

    #-----------------------------------------------------------------
    # Calculo de Yield Para Holdings

    # Se vuelve a crear el df de los fundamentales para tener en cuenta la información de BG.

    variables_v = [FEC_EMIS, FEC_VENC, FREQ_CUPON, BASE, FACIAL]
    df_DM_Holdings = pd.DataFrame(dict(zip(column_names, variables_v)))
    df_DM_Holdings = df_DM_Holdings.iloc[index_rf_BG]
    df_DM_Holdings['PRECIO'] = PRECIO[index_rf_BG]
    df_DM_Holdings['Nominal'] = NOMINAL_COP[index_rf_BG]/TRM

    # Verificación de que DM_Holdings se encuentre en el formato adecuado para el cálculo de la Duración Modificada.
    verify_dataframe(df_DM_Holdings)

    rows_with_issues = df_DM_Holdings.isna().any(axis = 1).sum()
    if rows_with_issues > 0:
        raise ValueError("Hay una o más filas con información faltante")
    else:
        IRR_Hold = df_IRR(df_DM_Holdings, fecha_corte,base_map) # Acá se realiza el cálculo de la yield.

    # Se incluye nuevamente en df para el cálculo final de la DM.
    df_DM_Holdings['Yield'] = IRR_Hold

    #-----------------------------------------------------------------
    # Calaculo de DM par Holdings

    # El precio sucio debe estar en términos porcentuales * 100.
    # Se crea un df porque es la manera en la que se diseño la entrada de MD

    DMaca_Hold = df_dur_maca(df_DM_Holdings, fecha_corte,base_map)
    DM_Hold =  DMaca_Hold/(1 + df_DM_Holdings['Yield'])
    df_DM_Holdings['DurMACA'] = DMaca_Hold
    df_DM_Holdings['DurMod'] = DM_Hold
    df_DM_Holdings['CUSIP'] = CUSIP[index_rf_BG]
    df_DM_Holdings['CLASIFICACION'] = CLASIF_RF[index_rf_BG]

    ### El paso final es crear el vector completo donde pueden haber CUSIPS repetidos.
    TIR = CUSIP.map(df_DM_Holdings.set_index('CUSIP')['Yield'])
    DUR_MACA = CUSIP.map(df_DM_Holdings.set_index('CUSIP')['DurMACA'])
    DM_H = CUSIP.map(df_DM_Holdings.set_index('CUSIP')['DurMod'])
    # Se vuelven date las columnas de df_DM_Holdings
    df_DM_Holdings[['Fecha de emisión','Fecha de vencimiento']] = df_DM_Holdings[['Fecha de emisión','Fecha de vencimiento']].apply(lambda x: x.dt.normalize())
    ## DV01
    DV01_H = VAL_MERCADO * DM_H * 0.0001 



    #-----------------------------------------------------------------
    # Consolidar la Informacion de ISINES

    cols_ISIN = ['FECHAEMISION','FECHAVENCIMIENTO','FACIAL', 'CODIGOPERIODO',
                'DAY_COUNT', 'CLASIFICACION']

    cols_from_BG = ['Fecha de emisión', 'Fecha de vencimiento', 'Facial', 'Frecuencia','Base','CLASIFICACION']


    conteo_titulos_añadidos = 0

    for index, row in df_DM_Holdings.iterrows():
        cusip = row['CUSIP']

        # Find match in ISIN
        match_index = isines[isines['CUSIP'] == cusip].index

        if not match_index.empty:
            isines.loc[match_index, cols_ISIN] = row[cols_from_BG].values
        
        else:    
            nuevo_titulo = row.rename(dict(zip(cols_from_BG, cols_ISIN)))
            nuevo_titulo = nuevo_titulo[cols_ISIN + ['CUSIP','CLASIFICACION']]
            isines = pd.concat([isines, nuevo_titulo.to_frame().T], ignore_index=True)
            conteo_titulos_añadidos += 1
    print("El número de títulos añadidos a ISIN es:", conteo_titulos_añadidos)
    # Formateo Final de Fechas antes de cargue a Excel
    cols = ['FECHAEMISION','FECHAVENCIMIENTO']
    for col in cols:
        isines[col] = pd.to_datetime(isines[col], errors = 'coerce').dt.date
    #-----------------------------------------------------------------
    # Guardar ISIN

    titulos_unicos_df = isines[isines['CUSIP'].isin(CUSIP)]
    titulos_unicos_convacios = titulos_unicos_df[titulos_unicos_df.isnull().any(axis = 1)]

    if titulos_unicos_convacios.shape[0]:
        warnings.warn(rf"Hay {titulos_unicos_convacios.shape[0]} títulos con al menos una columna vacía en ISIN")
        print(titulos_unicos_convacios['CUSIP'])

    isines.to_excel(ISIN_path, sheet_name="ISINES", index = False)

    #-----------------------------------------------------------------
    # Variables Manuales

    # Lectura de BG para los títulos manuales
    BMG_data_manual = pd.read_excel(path_BG, sheet_name= "BG_Manual")

    # Verificación de que el precio obtenido de Bloomberg sea el correcto.
    verify_manual(BMG_data_manual)

    # Se elimina el string que no aporta del CUSIP para dejar el código limpio.
    BMG_data_manual['Security'] = BMG_data_manual['Security'].str.replace('/cusip/', '', regex=False)
    BMG_data_manual = BMG_data_manual.rename(columns = {'Security': 'CUSIP'})

    # Se genera la Tabla que contiene toda la información para Portafolio
    # Estas son las variables que se generan a partir de la información ingresada manualmente y
    # algunos valores a través de Bloomberg.

    # Días al vencimiento
    DAYS_MAT = (Titulos_df['Fecha de vencimiento'] - pd.to_datetime(fecha_corte)).dt.days
    # Facial Tasa cupón
    FACIAL_MAN = Titulos_df['Facial']

    # Más que una valoración es una contabilización financiera, de cuantos intereses se han "causado"
    # En realidad no se ha causado nada, porque el CDT no es para vender, pero el cálculo se realiza por norma
    # o es estándar para el sector financiero.

    # Base numérica:
    base_num = Titulos_df['Base'].astype(float)

    VAL_hoy = Titulos_df['Nominal'] * (1 + FACIAL_MAN)**(((pd.to_datetime(fecha_corte) - Titulos_df['Fecha de emisión']).dt.days)/base_num)
    VAL_ayer = Titulos_df['Nominal'] * (1 + FACIAL_MAN)**(((pd.to_datetime(fecha_corte_ayer) - Titulos_df['Fecha de emisión']).dt.days)/base_num)
    # Aqui en val hoy debo editar para incluir los valores de Bloomberg
    BMG_manuales_portafolio = Titulos_df[Titulos_df['IF_BMG'] == 1].index 
    PRECIO_BMG_M = Titulos_df.loc[BMG_manuales_portafolio, 'ID'].map(BMG_data_manual.set_index('CUSIP')['PX_DIRTY_BID'])
    VAL_hoy[BMG_manuales_portafolio] = (PRECIO_BMG_M * Titulos_df.loc[BMG_manuales_portafolio, 'Nominal'])/100
    Titulos_df['PRECIO'] = (VAL_hoy/ Titulos_df['Nominal'])*100

    # Nominal de USD a COP.
    NOMINAL_COP_MAN = Titulos_df['Nominal'] * TRM

    # Calculo del Yield. Para los bonos no cero cupón se implementa la fórmula.
    Titulos_df['Yield'] = Titulos_df['Facial']
    index_notzc = Titulos_df[Titulos_df['Frecuencia'] != 0].index
    Titulos_df.loc[index_notzc, 'Yield'] = df_IRR(Titulos_df[Titulos_df['Frecuencia'] != 0], fecha_corte,base_map)
    TIR_M = Titulos_df['Yield']


    # El cálculo de la duración se realiza con Macaluay 
    # Para Los títulos cero cupón, luego duración es igual a plazo.
    DURACION = np.zeros(len(Titulos_df))
    index_zerocoupon = Titulos_df[Titulos_df['Frecuencia'] == 0].index
    index_notzc = Titulos_df[Titulos_df['Frecuencia'] != 0].index
    DURACION[index_zerocoupon] = DAYS_MAT[index_zerocoupon]/ 365

    # La función de df_dur_maca recibe un dataframe con las columnas relevantes y calcula la duración de Macaluay.
    # Por eso se filtra para aquellas filas que no corresponden a cero cupón.
    DURACION[index_notzc] = df_dur_maca(Titulos_df[Titulos_df['Frecuencia'] != 0], fecha_corte,base_map)
    DUR_MACA_M = DURACION
    # DM - Duración Modificada
    DM_M = DURACION / (1 + Titulos_df['Yield'])

    # DV01
    # Se diferencian los títulos que serán holdeados hasta el maturity de los que no. 
    # Pues para estos el cálculo del DV01 no tiene sentido.
    index_nohold_titles = Titulos_df.loc[Titulos_df['IF_BMG'] == 1].index.tolist()
    DV01_M = np.array([np.nan] * Titulos_df.shape[0])
    #DV01_M[index_nohold_titles] = VAL_hoy[index_nohold_titles] * DM_M[index_nohold_titles] * 0.0001 # 1pb = 0.0001
    DV01_M = VAL_hoy * DM_M * 0.0001

    # El precio puede ser entendido como (1+r) donde r es la causación ganada hasta el momento. 
    # Note que en el vencimiento, r será igual a la tasa cupón.
    PRICE_hoy = VAL_hoy/ Titulos_df['Nominal'] * 100
    #PRICE_ayer = VAL_ayer/ Titulos_df['Nominal'] * 100

    # Ingreso Valoración
    # El cambio en los intereses acumulados. Debe ser > 0.
    INGRESO_VAL = VAL_hoy - VAL_ayer
    INGRESO_VALOR = pd.Series([np.nan] * titulos_corte_Port)
    INGRESO_VALOR[-len(Titulos_df):] = INGRESO_VAL

    # FREQ CUPON - CODIGOPERIODO
    # Diccionario de Mapeo de strings: Solo simplifica la notación.
    # str_map_m = {0: 'CeroCupon', 1: 'AV', 2:'SV', 4:'TV'}
    FREQ_CUPON_M = Titulos_df['Frecuencia']

    # Fecha de compra (Esta variable solo se incluye para las variables manuales.)
    FECHA_COMPRA = [np.nan] * titulos_corte_Port
    FECHA_COMPRA[-len(Titulos_df):] = Titulos_df['Fecha de compra'].dt.strftime("%d-%m-%Y")
    FECHA_COMPRA = pd.Series(FECHA_COMPRA)

    # Valor Compra Origen
    VALOR_C_ORIGEN = pd.Series(np.array([np.nan] * titulos_corte_Port))
    VALOR_C_ORIGEN[-len(Titulos_df):] = Titulos_df['Valor compra origen']
    #VALOR_C_ORIGEN = pd.Series(VALOR_C_ORIGEN)

    # Se completa la Base Monetaria
    Portafolio = Portafolio.reset_index(drop = True)
    Portafolio.loc[Portafolio.index[(-len(Titulos_df)):],'BASEMONETARIA'] = np.array(Titulos_df['Moneda'])

    # Valor Mercado Ayer
    VALOR_MERCADO_AYER = pd.Series([np.nan] * titulos_corte_Port)
    VALOR_MERCADO_AYER[-len(Titulos_df):] = VAL_ayer

    # Si el título se vende o no
    IF_TRADE_M = Titulos_df['IF_BMG']

    #-----------------------------------------------------------------
    # Incorporación Titulos Manuales con Titulos de Pershing

    # Se incorporan los valores estáticos a los vectores con la información de Holdings.
    ESPECIE = pd.concat([ESPECIE, pd.Series(Titulos_df['Nemo'])], ignore_index= True)
    EMISOR = pd.concat([EMISOR, pd.Series(Titulos_df['Emisor'])], ignore_index= True)
    TITULO = pd.concat([TITULO, pd.Series(Titulos_df['Nombre del título'])], ignore_index= True)
    CUSIP = pd.concat([CUSIP, pd.Series(Titulos_df['ID'])], ignore_index= True)
    ISIN = pd.concat([ISIN, pd.Series(Titulos_df['ISIN'])], ignore_index= True)
    CALIF_CERT = pd.concat([CALIF_CERT, pd.Series(Titulos_df['Calificación'])], ignore_index= True)
    FEC_EMIS =  pd.concat([FEC_EMIS, pd.Series(Titulos_df['Fecha de emisión'])], ignore_index= True)
    FEC_EMIS = pd.to_datetime(FEC_EMIS).dt.strftime('%d-%m-%Y')
    FEC_VENC = pd.concat([FEC_VENC, pd.Series(Titulos_df['Fecha de vencimiento'])], ignore_index= True)
    FEC_VENC = pd.to_datetime(FEC_VENC).dt.strftime('%d-%m-%Y')
    CANTIDAD_VAL_NOM = pd.concat([CANTIDAD_VAL_NOM, pd.Series(Titulos_df['Nominal'])], ignore_index= True)
    FACIAL = pd.concat([FACIAL, pd.Series(FACIAL_MAN * 100)], ignore_index= True)
    FREQ_CUPON = pd.concat([FREQ_CUPON, pd.Series(FREQ_CUPON_M)], ignore_index= True)
    TIR =  pd.concat([TIR, pd.Series(TIR_M)], ignore_index= True)
    MON_ORIGEN = pd.concat([MON_ORIGEN, pd.Series(Titulos_df['Moneda'])], ignore_index= True)
    VAL_MERCADO = pd.concat([VAL_MERCADO, pd.Series(VAL_hoy)], ignore_index= True)
    DIAS = pd.concat([DIAS, pd.Series(DAYS_MAT)], ignore_index= True)
    PRECIO = pd.concat([PRECIO, pd.Series(PRICE_hoy)], ignore_index= True)
    POR = pd.concat([pd.Series(POR), pd.Series(Titulos_df['Por'])], ignore_index= True)
    DUR_MACA = pd.concat([DUR_MACA, pd.Series(DUR_MACA_M)], ignore_index= True)
    DM = pd.concat([DM_H, pd.Series(DM_M)], ignore_index= True)
    DV01 = pd.concat([DV01_H, pd.Series(DV01_M)], ignore_index= True)


    # Nominal COP
    NOMINAL_COP = pd.concat([pd.Series(NOMINAL_COP), pd.Series(NOMINAL_COP_MAN)], ignore_index= True)

    CLAS2 = pd.concat([pd.Series(CLAS2), pd.Series(['YANKEES_USD'] * len(Titulos_df))], ignore_index= True)
    SECTOR = pd.concat([pd.Series(SECTOR), pd.Series(['Financial'] * len(Titulos_df))], ignore_index= True)
    NACIONALIDAD = pd.concat([pd.Series(NACIONALIDAD), pd.Series(['PANAMA'] * len(Titulos_df))], ignore_index= True)
    PORTAFOLIO = pd.concat([pd.Series(PORTAFOLIO), pd.Series(Titulos_df['Portafolio'])], ignore_index= True)

    ASSET_CLASS = pd.concat([pd.Series(ASSET_CLASS), pd.Series(['LEGATRUU'] * len(Titulos_df))], ignore_index= True)
    NIT_P = pd.concat([pd.Series(NIT_P), pd.Series(Titulos_df['NiT Coasa'].astype(int))], ignore_index= True)
    IF_TRADE = pd.concat([pd.Series(IF_TRADE), pd.Series(IF_TRADE_M)], ignore_index= True)

    #-----------------------------------------------------------------
    # Incorporar la Información a Portafolio y exportar

    columnas_din_P = ['ESPECIE','EMISOR', 'TITULO','CUSIP','ISIN', 'CALIF','FECHAEMISION', 'FECHAVENCIMIENTO',
                    'CANTIDADVALORNOMINAL', 'FACIAL', 'CODIGOPERIODO', 'F_COMPRA', 'VALOR_COMPRA_ORIGEN',
                    'MONEDA_ORIGEN','VALOR_TIR','TIR', 'VALORMERCADO', 'DIAS', 'PRECIO', 'POR', 'VR_MDO_AYER', 'INGRESO_VALORACION',
                    'DURACION_MODIFICADA', 'NOMINAL_PESOS', 'CLASIFICACION2', 'SECTOR', 'NACIONALIDAD', 'PORTAFOLIO', 
                    'DV1', 'ASSET_CLASS',
                    'NIT_PORTAFOLIO', 'TIPO_ACTIVO', 'TIPO_DE_PRODUCTO', 'DETALLE_PRODUCTO','CLASIF_RF', 'REGION', 
                    'PERFIL', 'DUR_MACA','IF_TRADE']

    valores_din_P  = [ESPECIE, EMISOR, TITULO, CUSIP, ISIN, CALIF_CERT, FEC_EMIS, FEC_VENC, CANTIDAD_VAL_NOM,
                    FACIAL, FREQ_CUPON, FECHA_COMPRA, VALOR_C_ORIGEN, MON_ORIGEN, TIR, TIR, VAL_MERCADO, DIAS, PRECIO,
                    POR, VALOR_MERCADO_AYER, INGRESO_VALOR, DM, NOMINAL_COP, CLAS2, SECTOR, NACIONALIDAD, 
                    PORTAFOLIO, DV01, ASSET_CLASS, NIT_P, TIPO_ACTIVO, TIPO_PRODUCTO, DETALLE_PRODUCTO, CLASIF_RF,
                    REGION, PERFIL, DUR_MACA, IF_TRADE]

    Portafolio[columnas_din_P] = pd.concat(valores_din_P, axis = 1).set_axis(columnas_din_P, axis = 1)

    # Cargue a Excel de Salida
    #print(Portafolio['NIT_PORTAFOLIO'].tail())
    if not os.path.exists(path_salida):
        with pd.ExcelWriter(path_salida, engine= 'xlsxwriter') as writer:
            Portafolio.to_excel(writer, sheet_name= 'Portafolio', index = False)
            print(f"File created and saved at: {path_salida}")
    else:
        print(f"File already exists at: {path_salida}")

    #-----------------------------------------------------------------
    # Carpintería: Creacion de Output (Hoja de Liquidez)

    cuentas_vigentes = Porta_df.loc[Porta_df['VIGENTE'] == 1, 'CUENTA'].tolist()
    # Vamos a identificar cuales filas dentro de holdings cumplen las condiciones para la hoja de liquidez.
    # Para ello se deben verificar 3 condiciones: La cuenta asociada coincide con una cuenta vigente, sí es Cash y hay Trade.
    # Es correcto usar Trade Date MV, en la medida que para Cash puede no marcar Trade Date Quantity.
    filas_liquidez = holdings[
        holdings['Account'].isin(cuentas_vigentes) & 
        (holdings['Type'] == 'Cash') & 
        (holdings['Trade Date Market Value'] != 0)
    ]

    filas_liquidez = filas_liquidez.reset_index(drop = False)
    # Formateo en caso de ser necesario
    filas_liquidez.loc[:,'CUSIP'] = filas_liquidez['CUSIP'].astype(str)

    # Número de activos a incluir dentro de Portafolio
    # Esto será la suma de los que vienen de Holdings y los que se introducen manualmente.
    titulos_corte_Port_liq = len(filas_liquidez) + len(Saldosc_df)

    Liquidez = pd.DataFrame(columns= ['CLA', 'FTE', 'ESPECIE', 'SERIE', 'EMISOR','TITULO', 'TC', 'COD',
                                    'ISIN','CALIF', 'CODcopy', 'EMISION',	'F_VCTO', 'NOMINAL_ACTUAL',
                                    'FACIAL',	'MODD', 'F_COMPRA',	'VALOR_COMPRA_ORIGEN',	'MONEDA_ORIGEN',
                                    'MONED', 'VALOR_COMPRA', 'T_COMPRA',	'VALOR_TIR', 'TIR',	'VALOR_MERCADO', 
                                    'T_MER', 'DIAS', 'MET',	'PRECIO', 'TASA',	'VR_TASA', 'MARGEN',
                                    'T_DCTO', 'ES', 'FG', 'FV', 'POR', 'CUSIP', 'VR_MDO_AYER', 'VARIA_VR_MERCADO',
                                    'CAUSACION_MERCADO',	'CAUSACION_TIR', 'EFECTO_PATRIMONIAL', 'CUPON', 'FRECUENCIA',
                                    'DURACION_MODIFICADA', 'CONVEXIDAD', 'NOMINAL_PESOS',	'CLASIFICACION_1', 
                                    'CLASIFICACION_CUENTA', 'GRUPO_TASA', 'TIPO_TASA', 'EMISORcopy', 'SECTOR', 'NACIONALIDAD',
                                    'TIPO_EMISOR', 'GRUPO_FINANCIERO',	'PORTAFOLIO', 'EMPRESA', 'OBSERVACIONES', 
                                    'CUPON_REPRECIO', 'DURACION_REPRECIO', 'ASSET_CLASS', 'CBE_NUMERO_CUENTA',
                                    'NIT_EMISOR', 'CLASIFICACION_2', 'FECHAPORTAFOLIO'
                                    ], index = range(titulos_corte_Port_liq))

    # CLA
    Liquidez['CLA'] = "NEG"
    # CLE
    Liquidez['FTE'] = "1CP"
    # EMISOR
    Liquidez['EMISOR'] = "PERSHING"
    # COD y Copia COD
    Liquidez['COD'] = "CUENTAS EXTRANJERAS"
    Liquidez['CODcopy'] = "CUENTAS EXTRANJERAS" 
    # EMPRESA
    Liquidez['EMPRESA'] = "DAVIVIENDA CORREDORES PANAMA"
    # ASSET CLASS
    Liquidez['ASSET_CLASS'] = "VISTA"
    # CLASIFICACIÖN
    Liquidez['CLASIFICACION_2'] = "CUENTAS_AHORRO"
    # Fecha de corte
    Liquidez['FECHAPORTAFOLIO'] = fecha_corte_d

    #-----------------------------------------------------------------
    # Variables dinámicas que se traen directamente de Holdings

    # NOMINAL ACTUAL Y VALOR MERCADO
    VAL_MERCADO_L = filas_liquidez['Trade Date Market Value']

    # MONEDA
    MONEDA_L = filas_liquidez['Symbol']

    #-----------------------------------------------------------------
    # Variables VLOOKUP

    # POR Identifica el número del portafolio. 
    # Mezclamos el número del portafolio con el identificador.

    NUMERO_L = filas_liquidez['Account'].map(Porta_df.set_index('CUENTA')['NUMERO'])
    IDENT_L = np.array(["-A1"] * len(NUMERO_L))
    POR_L = pd.Series(list(map(lambda x, y: f"{x}{y}", NUMERO_L, IDENT_L)))

    # Portafolio
    PORTAFOLIO_L = filas_liquidez['Account'].map(Porta_df.set_index('CUENTA')['PORTAFOLIO'])

    #-----------------------------------------------------------------
    # Variables Manuales

    # El prefijo ML hace referencia a Manual y para la hoja de Liquidez.

    # Especie
    ESPECIE_ML = Saldosc_df['SALDOS CUENTAS']

    # Emisor
    EMISOR_ML = Saldosc_df['SALDOS CUENTAS'].str.split('-').str[0]

    # Portafolio
    PORTAFOLIO_ML = Saldosc_df['SALDOS CUENTAS'].str.split('-').str[1]

    # Nominal Actual y Valor Mercado
    VAL_MERCADO_ML = Saldosc_df['ID_Cuentas'].map(saldos_L.drop_duplicates(subset= 'CUENTA').set_index('CUENTA')['VALOR'])

    # Moneda queda la duda de si es estático.
    MONEDA_ML = np.array(['USD'] * len(Saldosc_df))

    # POR # Queda la duda si es estático
    POR_ML = np.array(['01-A1'] * len(Saldosc_df))

    #-----------------------------------------------------------------
    # Agrupacion entre Manuales y Holdings

    # EMISOR
    Liquidez.loc[Liquidez.index[-Saldosc_df.shape[0]:],'EMISOR'] = np.array(EMISOR_ML)

    # Portafolio L
    PORTAFOLIO_L = pd.concat([PORTAFOLIO_L, pd.Series(PORTAFOLIO_ML)], ignore_index= True)

    # ESPECIE
    Liquidez['ESPECIE'] = np.array(Liquidez['EMISOR'] + '-' + PORTAFOLIO_L)

    # Nominal Actual - Sale de Valor de Mercado (Holdings + Saldos_L)
    VAL_MERCADO_L = pd.concat([VAL_MERCADO_L, pd.Series(VAL_MERCADO_ML)], ignore_index= True)

    # MONEDA
    MONEDA_L =  pd.concat([MONEDA_L, pd.Series(MONEDA_ML)], ignore_index= True)

    # POR
    POR_L =  pd.concat([POR_L, pd.Series(POR_ML)], ignore_index= True)

    # CBE NUMERO CUENTA
    Liquidez.loc[Liquidez.index[-Saldosc_df.shape[0]:], 'CBE_NUMERO_CUENTA'] = np.array(Saldosc_df['ID_Cuentas'])

    #-----------------------------------------------------------------
    # Incorporar la info restante a Liquidez y Exportar

    columnas_din_L = ['NOMINAL_ACTUAL', 'MONED', 'VALOR_MERCADO', 'POR', 'PORTAFOLIO']
    valores_din_L  = [VAL_MERCADO_L,  MONEDA_L, VAL_MERCADO_L, POR_L, PORTAFOLIO_L]

    Liquidez[columnas_din_L] = pd.concat(valores_din_L, axis = 1).set_axis(columnas_din_L, axis = 1)

    # Exportar
    sheet_name = 'Liquidez'

    if os.path.exists(path_salida): 
        # Load the existing Excel file
        with pd.ExcelWriter(path_salida, engine='openpyxl', mode='a') as writer:
            # Load the workbook
            workbook = load_workbook(path_salida)
            # Check if the sheet already exists
            if sheet_name in workbook.sheetnames: 
                print(f"Warning: The sheet '{sheet_name}' already exists. No changes made.")
            else: 
                # If the sheet doesn't exist, append it 
                Liquidez.to_excel(writer, sheet_name=sheet_name, index = False) 
                print(f"Sheet '{sheet_name}' has been successfully appended.") 
    else: 
        print(f"Warning: The file '{path_salida}' does not exist.")

    #-----------------------------------------------------------------
    # Se realiza el append de dos hojas vacias: Divisas y Futuros

    Spot_Divisas = pd.DataFrame(columns= ['DET', 'TRANSACCION',	'FECHA_CUMPLIMIENTO', 'ESPECIE',
                                        'TITULO',	'CONSEC',	'T_FACIAL',	'MOD',	'PLAZO', 'MON',	'VR_NOMINAL',
                                        'TASA_NEG', 'VR_NOMINAL_ACTUAL', '_FECHA_EMISION', 'FECHAVECIMIENTO',
                                        'MNT', 'VR_TRASACCION', 'VR_NETO', 'VR_RECIBIDO', 'NOMBRE_CONTRAPARTE',
                                        'NIT_CONTRAPARTE', 'PORTAFOLIO', 'SCF', 'USUARIO_NEGOCIADOR', 'PLATAFORMA',
                                        'FECHA_PACTO', 'HORA_NEGOCIACION', 'TIPO_PRODUCTO', 'TIPO_OPERACION', 
                                        'NEMOTECNICO', 'CRU/COV',	'NIT_EMISOR', 'ASSET_CLASS', 'VALORACION_AYER',
                                        'VALORACION_HOY',	'UTILIDAD',	'VALOR_OPERACION',	'CANAL_ORIGEN',	'NIT_CLIENTE'
                                    ], index = range(1))

    Futuros = pd.DataFrame(columns= ['FECHA_OPERACION', 'NEMOTECNICO',	'FECHA_CUMPLIMIENTO', 'PUNTA_OPERACION',
                                    'TIPO_PRODUCTO',	'EMPRESA', 'NO_CONTRATOS', 'NOMINAL', 'VALOR_PRESENTE',
                                    'PORTAFOLIO', 'MONEDA',	'PRECIO_VALORACION', 'MER',	'FECHA_PACTO',	'COMERCIAL',
                                    'U_P_DIARIO', 'U_P_ACUMULADO'], index = range(1))

    # Exportar Divisas
    sheet_name = 'Spot Divisas'

    if os.path.exists(path_salida): 
        # Load the existing Excel file
        with pd.ExcelWriter(path_salida, engine='openpyxl', mode='a') as writer:
            # Load the workbook
            workbook = load_workbook(path_salida)
            # Check if the sheet already exists
            if sheet_name in workbook.sheetnames: 
                print(f"Warning: The sheet '{sheet_name}' already exists. No changes made.")
            else: 
                # If the sheet doesn't exist, append it 
                Spot_Divisas.to_excel(writer, sheet_name=sheet_name, index = False) 
                print(f"Sheet '{sheet_name}' has been successfully appended.") 
    else: 
        print(f"Warning: The file '{path_salida}' does not exist.")

    # Exportar Futuros
    sheet_name = 'Futuros'

    if os.path.exists(path_salida): 
        # Load the existing Excel file
        with pd.ExcelWriter(path_salida, engine='openpyxl', mode='a') as writer:
            # Load the workbook
            workbook = load_workbook(path_salida)
            # Check if the sheet already exists
            if sheet_name in workbook.sheetnames: 
                print(f"Warning: The sheet '{sheet_name}' already exists. No changes made.")
            else: 
                # If the sheet doesn't exist, append it 
                Futuros.to_excel(writer, sheet_name=sheet_name, index = False) 
                print(f"Sheet '{sheet_name}' has been successfully appended.") 
    else: 
        print(f"Warning: The file '{path_salida}' does not exist.")

    #-----------------------------------------------------------------
    # Carga Información a GCP

    # El schema define cada uno de los formatos de las columnas que se carga. Portafolio
    schema_port = [
        bigquery.SchemaField("FECHAPORTAFOLIO", "TIMESTAMP", mode="REQUIRED"),
        bigquery.SchemaField("CLA", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("FTE", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("ESPECIE", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("SERIE", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("EMISOR", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("TITULO", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("TC", "FLOAT64", mode="NULLABLE"),
        bigquery.SchemaField("CUSIP", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("ISIN", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CALIF", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("COD", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("FECHAEMISION", "TIMESTAMP", mode="NULLABLE"),
        bigquery.SchemaField("FECHAVENCIMIENTO", "TIMESTAMP", mode="NULLABLE"),
        bigquery.SchemaField("CANTIDADVALORNOMINAL", "FLOAT64", mode="REQUIRED"),
        bigquery.SchemaField("FACIAL", "FLOAT64", mode="NULLABLE"),
        bigquery.SchemaField("CODIGOPERIODO", "INTEGER", mode="NULLABLE"),
        bigquery.SchemaField("F_COMPRA", "TIMESTAMP", mode="NULLABLE"),
        bigquery.SchemaField("VALOR_COMPRA_ORIGEN", "FLOAT64", mode="NULLABLE"),
        bigquery.SchemaField("MONEDA_ORIGEN", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("BASEMONETARIA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("VALOR_COMPRA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("T_COMPRA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("VALOR_TIR", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("TIR", "FLOAT64", mode="NULLABLE"),
        bigquery.SchemaField("VALORMERCADO", "FLOAT64", mode="REQUIRED"),
        bigquery.SchemaField("T_MER", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("DIAS", "INTEGER", mode="NULLABLE"),
        bigquery.SchemaField("MET", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("PRECIO", "FLOAT64", mode="REQUIRED"),
        bigquery.SchemaField("TASA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("VR_TASA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("MARGEN", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("T_DCTO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("ES", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("FG", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("FV", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("POR", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("VR_MDO_AYER", "FLOAT64", mode="NULLABLE"),
        bigquery.SchemaField("INGRESO_VALORACION", "FLOAT64", mode="NULLABLE"),
        bigquery.SchemaField("CAUSACION", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("DOS_MENOS_UNO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("EFECTO_PATRIMONIAL", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("DURACION_MODIFICADA", "FLOAT64", mode="NULLABLE"),
        bigquery.SchemaField("CONVEXIDAD", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("NOMINAL_PESOS", "FLOAT64", mode="NULLABLE"),
        bigquery.SchemaField("CLASIFICACION1", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CLASIFICACION2", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("GRUPO_TASA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("TIPO_TASA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("EMISOR2", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("SECTOR", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("NACIONALIDAD", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("TIPO_EMISOR", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("GRUPO_FINANCIERO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("PORTAFOLIO", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("EMPRESA", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("ID_TITULO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CUPON_REPRECIO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("DV1", "FLOAT64", mode="NULLABLE"),
        bigquery.SchemaField("ASSET_CLASS", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CODIGOTIPOOPERACIONPORTAFOLIO", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("NIT_PORTAFOLIO", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("TIPO_ACTIVO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("TIPO_DE_PRODUCTO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("DETALLE_PRODUCTO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CLASIF_RF", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("REGION_PERFIL", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("REGION", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("PERFIL", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("DUR_MACA", "FLOAT64", mode="NULLABLE"),
        bigquery.SchemaField("IF_TRADE", "FLOAT64", mode="NULLABLE") 
        ] 

    # Formateo de columnas previo a GCP
    cols_to_datetime = ['FECHAPORTAFOLIO','FECHAEMISION','FECHAVENCIMIENTO','F_COMPRA']
    cols_to_float64 = ['TC']
    cols_to_string = ['FTE','SERIE','NIT_PORTAFOLIO','COD','VALOR_COMPRA','T_COMPRA','VALOR_TIR','T_MER','TASA','VR_TASA','MARGEN',
                    'T_DCTO','ES','FG','FV','CAUSACION','DOS_MENOS_UNO','EFECTO_PATRIMONIAL','CONVEXIDAD',
                    'CLASIFICACION1','GRUPO_TASA','TIPO_TASA','EMISOR2','TIPO_EMISOR','GRUPO_FINANCIERO',
                    'ID_TITULO','CUPON_REPRECIO','REGION_PERFIL']

    Portafolio = Formateo_df2GCP(Portafolio, cols_to_datetime, cols_to_float64, cols_to_string, dayfirst=True)
    Portafolio['DIAS'] = pd.to_numeric(Portafolio['DIAS'], errors='coerce').astype('Int64')

    # Se realiza la carga de información vía APPEND a GCP
    client, dataset_ref, tables_ref = charge_serv(where_to_run=where_to_run,project=project, dataset_id = dataset_id, tables_names=[])
    #client, dataset_ref, tables_ref = charge_serv(where_to_run='local',project=project, dataset_id = dataset_id, tables_names=[])
    table_ref = "{}.{}.{}".format(project,dataset_id,'Portafolio_H_Pan')
    nombre_fecha_GCP = "FECHAPORTAFOLIO"
    query = [ideas_querys['cargue_generico'].format(nombre_fecha_GCP,table_ref,nombre_fecha_GCP)]
    fechas_bq = multiple_query(client,query)[0]
    booleano = checker_fechas_cargue(fechas_bq,nombre_fecha_GCP,fecha_corte_ayer)
    if booleano:
        upload_table(client,big_query_table_ref=table_ref,table_to_append=Portafolio,schema=schema_port)

    # Liquidez
    # El schema define cada uno de los formatos de las columnas que se carga. Liquidez
    schema_liq = [
        bigquery.SchemaField("CLA", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("FTE", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("ESPECIE", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("SERIE", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("EMISOR", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("TITULO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("TC", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("COD", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("ISIN", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CALIF", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CODcopy", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("EMISION", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("F_VCTO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("NOMINAL_ACTUAL", "FLOAT64", mode="REQUIRED"),
        bigquery.SchemaField("FACIAL", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("MODD", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("F_COMPRA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("VALOR_COMPRA_ORIGEN", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("MONEDA_ORIGEN", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("MONED", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("VALOR_COMPRA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("T_COMPRA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("VALOR_TIR", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("TIR", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("VALOR_MERCADO", "FLOAT64", mode="REQUIRED"),
        bigquery.SchemaField("T_MER", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("DIAS", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("MET", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("PRECIO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("TASA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("VR_TASA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("MARGEN", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("T_DCTO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("ES", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("FG", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("FV", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("POR", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CUSIP", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("VR_MDO_AYER", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("VARIA_VR_MERCADO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CAUSACION_MERCADO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CAUSACION_TIR", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("EFECTO_PATRIMONIAL", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CUPON", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("FRECUENCIA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("DURACION_MODIFICADA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CONVEXIDAD", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("NOMINAL_PESOS", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CLASIFICACION_1", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CLASIFICACION_CUENTA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("GRUPO_TASA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("TIPO_TASA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("EMISORcopy", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("SECTOR", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("NACIONALIDAD", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("TIPO_EMISOR", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("GRUPO_FINANCIERO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("PORTAFOLIO", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("EMPRESA", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("OBSERVACIONES", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CUPON_REPRECIO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("DURACION_REPRECIO", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("ASSET_CLASS", "STRING", mode="REQUIRED"),
        bigquery.SchemaField("CBE_NUMERO_CUENTA", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("NIT_EMISOR", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("CLASIFICACION_2", "STRING", mode="NULLABLE"),
        bigquery.SchemaField("FECHAPORTAFOLIO", "TIMESTAMP", mode="REQUIRED")
        ]

    # Formateo de columnas previo a GCP
    Liquidez['FECHAPORTAFOLIO'] = pd.to_datetime(Liquidez['FECHAPORTAFOLIO'], dayfirst= True)

    cols_to_string = ['CLA', 'FTE', 'ESPECIE', 'SERIE', 'EMISOR','TITULO', 'TC', 'COD',
                    'ISIN','CALIF', 'CODcopy', 'EMISION',	'F_VCTO', 'FACIAL',	'MODD',
                    'F_COMPRA', 'VALOR_COMPRA_ORIGEN',	'MONEDA_ORIGEN', 'MONED',
                    'VALOR_COMPRA', 'T_COMPRA', 'VALOR_TIR', 'TIR', 'T_MER', 'DIAS',
                    'MET', 'PRECIO', 'TASA', 'VR_TASA', 'MARGEN', 'T_DCTO', 'ES', 'FG', 
                    'FV', 'POR', 'CUSIP', 'VR_MDO_AYER', 'VARIA_VR_MERCADO',
                    'CAUSACION_MERCADO',	'CAUSACION_TIR', 'EFECTO_PATRIMONIAL', 'CUPON',
                    'FRECUENCIA', 'DURACION_MODIFICADA', 'CONVEXIDAD', 'NOMINAL_PESOS',
                    'CLASIFICACION_1', 'CLASIFICACION_CUENTA', 'GRUPO_TASA', 'TIPO_TASA', 
                    'EMISORcopy', 'SECTOR', 'NACIONALIDAD', 'TIPO_EMISOR', 'GRUPO_FINANCIERO',
                    'PORTAFOLIO', 'EMPRESA', 'OBSERVACIONES', 'CUPON_REPRECIO', 'DURACION_REPRECIO',
                    'ASSET_CLASS', 'CBE_NUMERO_CUENTA', 'NIT_EMISOR', 'CLASIFICACION_2']
    Liquidez[cols_to_string] = Liquidez[cols_to_string].astype(str)

    table_ref = "{}.{}.{}".format(project,dataset_id,'Liquidez_H_Pan')
    nombre_fecha_GCP = "FECHAPORTAFOLIO"
    query = [ideas_querys['cargue_generico'].format(nombre_fecha_GCP,table_ref,nombre_fecha_GCP)]
    fechas_bq = multiple_query(client,query)[0]
    booleano = checker_fechas_cargue(fechas_bq,nombre_fecha_GCP,fecha_corte_ayer)
    if booleano:
        upload_table(client,big_query_table_ref=table_ref,table_to_append=Liquidez,schema=schema_liq)
    else:
        print('No se cargó')

    

