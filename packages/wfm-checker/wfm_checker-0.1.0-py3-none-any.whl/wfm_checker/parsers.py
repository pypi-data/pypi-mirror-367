"""
Input parsers for various file formats.

This module contains functions to parse item lists from TXT, CSV, and XLSX files.
"""

import re
import csv
import openpyxl
from typing import List, Tuple, Optional, Dict, Any
from .normalizers import get_ayatan_star_data


def is_ayatan_sculpture(item_name: str) -> bool:
    """Check if an item is an Ayatan sculpture."""
    max_amber, max_cyan = get_ayatan_star_data(item_name)
    return max_amber is not None and max_cyan is not None


def parse_rank(value) -> int:
    """
    Parse rank value, supporting 'MAX' keyword.
    
    Args:
        value: Input value (can be number or 'MAX')
        
    Returns:
        Rank as integer, or -1 for MAX (to be resolved later)
    """
    if not value or not str(value).strip():
        return 0
    
    value_str = str(value).strip().upper()
    
    if value_str == "MAX":
        return -1  # Special value indicating max rank should be found
    else:
        try:
            return int(value_str)
        except ValueError:
            return 0


def parse_ayatan_stars(value: str, item_name: str) -> Tuple[Optional[int], Optional[int]]:
    """
    Parse Ayatan star configuration from input value.
    
    Args:
        value: Input value (can be "FULL", "EMPTY", or numeric)
        item_name: Name of the Ayatan sculpture
        
    Returns:
        Tuple of (amber_stars, cyan_stars)
    """
    if not value or not str(value).strip():
        return None, None
    
    value_str = str(value).strip().upper()
    
    if value_str == "FULL":
        # Get max star configuration
        max_amber, max_cyan = get_ayatan_star_data(item_name)
        return max_amber, max_cyan
    elif value_str == "EMPTY":
        return 0, 0
    else:
        # Try to parse as number
        try:
            return int(value_str), None  # Single number assumes amber stars
        except ValueError:
            return None, None


def parse_input(file_path: str) -> List[Dict[str, Any]]:
    """
    Main input parser that routes to format-specific parsers.
    
    Args:
        file_path: Path to the input file
        
    Returns:
        List of dictionaries containing item data with keys:
        - quantity: int
        - item_name: str
        - rank: int (for mods)
        - amber_stars: Optional[int] (for Ayatan sculptures)
        - cyan_stars: Optional[int] (for Ayatan sculptures)
        
    Raises:
        ValueError: If file format is not supported
    """
    if file_path.endswith('.txt'):
        return parse_txt(file_path)
    elif file_path.endswith('.csv'):
        return parse_csv(file_path)
    elif file_path.endswith('.xlsx'):
        return parse_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file format: {file_path}")


def parse_txt(file_path: str) -> List[Dict[str, Any]]:
    """
    Text file parser with support for optional rank parameter.
    
    Args:
        file_path: Path to the text file
        
    Returns:
        List of dictionaries containing item data
    """
    items = []
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    for line in lines:
        line = line.strip()
        if not line or line.endswith(':'):
            continue
        # Remove category labels
        line = re.sub(r'^[\w\s]+:\s*', '', line)

        # Split into individual entries
        entries = re.split(r',\s*', line)

        for entry in entries:
            entry = entry.strip()
            if not entry:
                continue

            # Match patterns with quantity and optional rank (including MAX)
            match = re.match(r'^(\d+)\s+(?:copies of|copy of|of)?\s*(.+?)(?:\s+r?(max|\d+))?$', entry, re.IGNORECASE)
            if match:
                quantity = int(match.group(1))
                item_name = match.group(2).strip().lower()
                rank_str = match.group(3) or "0"
                rank = parse_rank(rank_str)
                items.append({
                    'quantity': quantity,
                    'item_name': item_name,
                    'rank': rank,
                    'amber_stars': None,
                    'cyan_stars': None
                })
            else:
                # Handle entries without explicit quantity
                tokens = entry.split()
                if tokens and (tokens[-1].isdigit() or tokens[-1].upper() == "MAX"):
                    rank = parse_rank(tokens[-1])
                    item_name = " ".join(tokens[:-1]).lower()
                    items.append({
                        'quantity': 1,
                        'item_name': item_name,
                        'rank': rank,
                        'amber_stars': None,
                        'cyan_stars': None
                    })
                else:
                    items.append({
                        'quantity': 1,
                        'item_name': entry.lower(),
                        'rank': 0,
                        'amber_stars': None,
                        'cyan_stars': None
                    })
    return items


def parse_csv(file_path: str) -> List[Dict[str, Any]]:
    """
    CSV parser with support for rank, amber_stars, and cyan_stars columns.
    
    Args:
        file_path: Path to the CSV file
        
    Returns:
        List of dictionaries containing item data
    """
    items = []
    with open(file_path, 'r', encoding='utf-8') as file:
        reader = csv.reader(file)
        headers = next(reader, None)
        
        # Detect column positions
        quantity_col, item_col, rank_col, amber_col, cyan_col = detect_columns(headers)
        
        for row in reader:
            if len(row) <= max(quantity_col, item_col):
                continue
                
            quantity = parse_quantity(row[quantity_col])
            item_name = row[item_col].strip().lower()
            
            # Initialize defaults
            rank = 0
            amber_stars = None
            cyan_stars = None
            
            # Check if this is an Ayatan sculpture
            if is_ayatan_sculpture(item_name):
                # For Ayatan sculptures, use amber/cyan columns with FULL/EMPTY support
                if amber_col is not None and len(row) > amber_col and row[amber_col]:
                    amber_stars, cyan_from_amber = parse_ayatan_stars(row[amber_col], item_name)
                    if cyan_from_amber is not None:
                        cyan_stars = cyan_from_amber
                
                # If cyan column exists and we didn't get it from amber parsing
                if cyan_col is not None and len(row) > cyan_col and row[cyan_col] and cyan_stars is None:
                    # For separate cyan column, just parse as number or FULL/EMPTY
                    try:
                        cyan_stars = int(row[cyan_col])
                    except (ValueError, TypeError):
                        # Try FULL/EMPTY parsing for cyan column too
                        _, cyan_stars = parse_ayatan_stars(row[cyan_col], item_name)
            else:
                # For non-Ayatan items, use rank column with MAX support
                if rank_col is not None and len(row) > rank_col:
                    rank = parse_rank(row[rank_col])
            
            if item_name:
                items.append({
                    'quantity': quantity,
                    'item_name': item_name,
                    'rank': rank,
                    'amber_stars': amber_stars,
                    'cyan_stars': cyan_stars
                })
    return items


def parse_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """
    Excel parser with support for rank, amber_stars, and cyan_stars columns.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        List of dictionaries containing item data
    """
    items = []
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Detect column positions
    headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
    quantity_col, item_col, rank_col, amber_col, cyan_col = detect_columns(headers)
    
    start_row = 2 if headers else 1
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if len(row) <= max(quantity_col, item_col):
            continue
        if row[item_col] is None:
            continue
        quantity = parse_quantity(row[quantity_col])
        item_name = str(row[item_col]).strip().lower()
        
        # Initialize defaults
        rank = 0
        amber_stars = None
        cyan_stars = None
        
        # Check if this is an Ayatan sculpture
        if is_ayatan_sculpture(item_name):
            # For Ayatan sculptures, use amber/cyan columns with FULL/EMPTY support
            if amber_col is not None and len(row) > amber_col and row[amber_col] is not None:
                amber_stars, cyan_from_amber = parse_ayatan_stars(str(row[amber_col]), item_name)
                if cyan_from_amber is not None:
                    cyan_stars = cyan_from_amber
            
            # If cyan column exists and we didn't get it from amber parsing
            if cyan_col is not None and len(row) > cyan_col and row[cyan_col] is not None and cyan_stars is None:
                # For separate cyan column, just parse as number or FULL/EMPTY
                try:
                    cyan_stars = int(row[cyan_col])
                except (ValueError, TypeError):
                    # Try FULL/EMPTY parsing for cyan column too
                    _, cyan_stars = parse_ayatan_stars(str(row[cyan_col]), item_name)
        else:
            # For non-Ayatan items, use rank column with MAX support
            if rank_col is not None and len(row) > rank_col and row[rank_col] is not None:
                rank = parse_rank(row[rank_col])
            
        if item_name:
            items.append({
                'quantity': quantity,
                'item_name': item_name,
                'rank': rank,
                'amber_stars': amber_stars,
                'cyan_stars': cyan_stars
            })
    return items


def detect_columns(headers: Optional[List]) -> Tuple[int, int, Optional[int], Optional[int], Optional[int]]:
    """
    Detect quantity, item, rank, amber_stars, and cyan_stars columns.
    
    Args:
        headers: List of column headers
        
    Returns:
        Tuple of (quantity_col, item_col, rank_col, amber_col, cyan_col)
    """
    quantity_col = 0
    item_col = 1
    rank_col = None
    amber_col = None
    cyan_col = None
    
    if headers:
        header_lower = [str(h).lower().strip() for h in headers]
        try:
            quantity_col = header_lower.index('quantity')
        except ValueError:
            pass
        try:
            item_col = header_lower.index('item')
        except ValueError:
            pass
        try:
            rank_col = header_lower.index('rank')
        except ValueError:
            pass
        
        # Look for amber stars column variations
        for amber_variant in ['amber_stars', 'amber stars', 'amber', 'a_stars', 'astars']:
            try:
                amber_col = header_lower.index(amber_variant)
                break
            except ValueError:
                continue
        
        # Look for cyan stars column variations
        for cyan_variant in ['cyan_stars', 'cyan stars', 'cyan', 'c_stars', 'cstars']:
            try:
                cyan_col = header_lower.index(cyan_variant)
                break
            except ValueError:
                continue
    
    return quantity_col, item_col, rank_col, amber_col, cyan_col


def parse_quantity(value) -> int:
    """
    Parse quantity from various input types.
    
    Args:
        value: Input value to parse as quantity
        
    Returns:
        Parsed quantity as integer, defaults to 1 if parsing fails
    """
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return 1
