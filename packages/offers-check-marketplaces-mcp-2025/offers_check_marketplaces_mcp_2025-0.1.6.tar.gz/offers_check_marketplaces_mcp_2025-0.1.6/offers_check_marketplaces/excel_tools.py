"""
Модуль для работы с Excel файлами через MCP Tools.
Предоставляет инструменты для парсинга и выгрузки Excel таблиц.
"""

import pandas as pd
import logging
from typing import List, Dict, Any, Optional, Union
from pathlib import Path
import asyncio
from concurrent.futures import ThreadPoolExecutor
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .error_handling import (
    handle_errors,
    ErrorCategory,
    ErrorSeverity,
    DataProcessingError,
    ValidationError,
    handle_validation_errors,
    error_handler,
    log_recovery_attempt
)

logger = logging.getLogger(__name__)


class ExcelTools:
    """Класс для работы с Excel файлами через MCP Tools."""
    
    def __init__(self):
        """Инициализация Excel инструментов."""
        self.executor = ThreadPoolExecutor(max_workers=2)
    
    @handle_errors(
        category=ErrorCategory.PARSING,
        severity=ErrorSeverity.HIGH,
        component="excel_tools",
        recovery_action="Возврат пустого результата при ошибке парсинга",
        return_on_error={"status": "error", "data": [], "message": "Ошибка парсинга Excel файла"}
    )
    async def parse_excel_file(self, file_path: str, sheet_name: Optional[str] = None, 
                              header_row: int = 0, max_rows: Optional[int] = None) -> Dict[str, Any]:
        """
        Парсит Excel файл и возвращает данные в структурированном виде.
        
        Args:
            file_path: Путь к Excel файлу
            sheet_name: Название листа (если None, берется первый лист)
            header_row: Номер строки с заголовками (0-based)
            max_rows: Максимальное количество строк для чтения
            
        Returns:
            Словарь с данными из Excel файла
        """
        try:
            logger.info(f"Парсинг Excel файла: {file_path}")
            
            # Проверяем существование файла
            if not Path(file_path).exists():
                raise FileNotFoundError(f"Excel файл не найден: {file_path}")
            
            # Асинхронно читаем Excel файл
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(
                self.executor,
                self._parse_excel_sync,
                file_path,
                sheet_name,
                header_row,
                max_rows
            )
            
            logger.info(f"Успешно распарсен Excel файл: {len(result['data'])} строк")
            return result
            
        except Exception as e:
            logger.error(f"Ошибка при парсинге Excel файла: {e}")
            raise
    
    def _parse_excel_sync(self, file_path: str, sheet_name: Optional[str], 
                         header_row: int, max_rows: Optional[int]) -> Dict[str, Any]:
        """
        Синхронный парсинг Excel файла.
        
        Args:
            file_path: Путь к Excel файлу
            sheet_name: Название листа
            header_row: Номер строки с заголовками
            max_rows: Максимальное количество строк
            
        Returns:
            Словарь с данными из Excel файла
        """
        # Читаем Excel файл
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=header_row,
            nrows=max_rows,
            engine='openpyxl'
        )
        
        # Получаем информацию о листах
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = workbook.sheetnames
        active_sheet = sheet_name if sheet_name else sheet_names[0]
        workbook.close()
        
        # Преобразуем DataFrame в список словарей
        data = []
        for index, row in df.iterrows():
            row_data = {}
            for column in df.columns:
                value = row[column]
                # Обрабатываем NaN значения
                if pd.isna(value):
                    row_data[str(column)] = None
                elif isinstance(value, (int, float)):
                    row_data[str(column)] = value
                else:
                    row_data[str(column)] = str(value).strip()
            
            row_data['_row_index'] = int(index)
            data.append(row_data)
        
        return {
            "status": "success",
            "file_path": file_path,
            "sheet_name": active_sheet,
            "available_sheets": sheet_names,
            "columns": [str(col) for col in df.columns.tolist()],
            "total_rows": len(data),
            "data": data,
            "parsed_at": datetime.now().isoformat()
        }
    
    @handle_errors(
        category=ErrorCategory.PARSING,
        severity=ErrorSeverity.MEDIUM,
        component="excel_tools",
        recovery_action="Возврат базовой информации о файле",
        return_on_error={"status": "error", "message": "Ошибка получения информации о файле"}
    )
    async def get_excel_info(self, file_path: str) -> Dict[str, Any]:
        """
        Получает информацию о структуре Excel файла без полного чтения данных.
        
        Args:
            file_path: Путь к Excel файлу
            
        Returns:
            Словарь с информацией о структуре файла
        """
        try:
            logger.info(f"Получение информации о Excel файле: {file_path}")
            
            if not Path(file_path).exists():
                raise FileNotFoundError(f"Excel файл не найден: {file_path}")
            
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(
                self.executor,
                self._get_excel_info_sync,
                file_path
            )
            
            return result
            
        except Exception as e:
            logger.error(f"Ошибка при получении информации о Excel файле: {e}")
            raise
    
    def _get_excel_info_sync(self, file_path: str) -> Dict[str, Any]:
        """
        Синхронное получение информации о Excel файле.
        
        Args:
            file_path: Путь к Excel файлу
            
        Returns:
            Словарь с информацией о файле
        """
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        
        sheets_info = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Получаем размеры листа
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # Читаем первые несколько строк для определения заголовков
            headers = []
            if max_row > 0:
                first_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
                headers = [str(cell) if cell is not None else f"Column_{i+1}" 
                          for i, cell in enumerate(first_row) if i < max_col]
            
            sheets_info.append({
                "name": sheet_name,
                "max_row": max_row,
                "max_column": max_col,
                "headers": headers,
                "estimated_data_rows": max_row - 1 if max_row > 0 else 0
            })
        
        workbook.close()
        
        return {
            "status": "success",
            "file_path": file_path,
            "file_size": Path(file_path).stat().st_size,
            "sheets": sheets_info,
            "total_sheets": len(sheets_info),
            "analyzed_at": datetime.now().isoformat()
        }
    
    @handle_errors(
        category=ErrorCategory.PARSING,
        severity=ErrorSeverity.HIGH,
        component="excel_tools",
        recovery_action="Возврат False при ошибке экспорта",
        return_on_error={"status": "error", "success": False, "message": "Ошибка экспорта в Excel"}
    )
    async def export_to_excel(self, data: List[Dict[str, Any]], file_path: str,
                             sheet_name: str = "Data", include_index: bool = False,
                             auto_adjust_columns: bool = True,
                             apply_formatting: bool = True) -> Dict[str, Any]:
        """
        Экспортирует данные в Excel файл с форматированием.
        
        Args:
            data: Список словарей с данными для экспорта
            file_path: Путь для сохранения Excel файла
            sheet_name: Название листа
            include_index: Включать ли индекс строк
            auto_adjust_columns: Автоматически подгонять ширину колонок
            apply_formatting: Применять ли форматирование
            
        Returns:
            Словарь с результатом экспорта
        """
        try:
            logger.info(f"Экспорт данных в Excel файл: {file_path}")
            
            if not data:
                raise ValueError("Нет данных для экспорта")
            
            # Создаем директорию если не существует
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(
                self.executor,
                self._export_to_excel_sync,
                data,
                file_path,
                sheet_name,
                include_index,
                auto_adjust_columns,
                apply_formatting
            )
            
            logger.info(f"Успешно экспортированы данные в {file_path}")
            return result
            
        except Exception as e:
            logger.error(f"Ошибка при экспорте в Excel: {e}")
            raise
    
    def _export_to_excel_sync(self, data: List[Dict[str, Any]], file_path: str,
                             sheet_name: str, include_index: bool,
                             auto_adjust_columns: bool, apply_formatting: bool) -> Dict[str, Any]:
        """
        Синхронный экспорт данных в Excel файл.
        
        Args:
            data: Данные для экспорта
            file_path: Путь к файлу
            sheet_name: Название листа
            include_index: Включать индекс
            auto_adjust_columns: Подгонять колонки
            apply_formatting: Применять форматирование
            
        Returns:
            Результат экспорта
        """
        # Создаем DataFrame из данных
        df = pd.DataFrame(data)
        
        # Создаем Excel writer
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Записываем данные
            df.to_excel(writer, sheet_name=sheet_name, index=include_index)
            
            # Получаем worksheet для форматирования
            worksheet = writer.sheets[sheet_name]
            
            if apply_formatting:
                self._apply_excel_formatting(worksheet, df, include_index)
            
            if auto_adjust_columns:
                self._auto_adjust_columns(worksheet)
        
        return {
            "status": "success",
            "success": True,
            "file_path": file_path,
            "sheet_name": sheet_name,
            "rows_exported": len(data),
            "columns_exported": len(data[0].keys()) if data else 0,
            "exported_at": datetime.now().isoformat()
        }
    
    def _apply_excel_formatting(self, worksheet, df: pd.DataFrame, include_index: bool):
        """
        Применяет форматирование к Excel листу.
        
        Args:
            worksheet: Лист Excel для форматирования
            df: DataFrame с данными
            include_index: Включен ли индекс
        """
        # Стили для заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Стили для границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Применяем форматирование к заголовкам
        header_row = 1
        start_col = 2 if include_index else 1
        
        for col_num in range(start_col, len(df.columns) + start_col):
            cell = worksheet.cell(row=header_row, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Применяем границы к данным
        for row in range(1, len(df) + 2):
            for col in range(1, len(df.columns) + start_col):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
    
    def _auto_adjust_columns(self, worksheet):
        """
        Автоматически подгоняет ширину колонок.
        
        Args:
            worksheet: Лист Excel для подгонки колонок
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Устанавливаем ширину с небольшим запасом, но не более 50
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @handle_errors(
        category=ErrorCategory.PARSING,
        severity=ErrorSeverity.MEDIUM,
        component="excel_tools",
        recovery_action="Возврат пустого результата",
        return_on_error={"status": "error", "data": [], "message": "Ошибка фильтрации данных"}
    )
    async def filter_excel_data(self, data: List[Dict[str, Any]], 
                               filters: Dict[str, Any]) -> Dict[str, Any]:
        """
        Фильтрует данные Excel по заданным критериям.
        
        Args:
            data: Исходные данные для фильтрации
            filters: Словарь с критериями фильтрации
            
        Returns:
            Отфильтрованные данные
        """
        try:
            logger.info(f"Фильтрация данных по критериям: {filters}")
            
            filtered_data = data.copy()
            
            for column, criteria in filters.items():
                if isinstance(criteria, dict):
                    # Сложные критерии фильтрации
                    filtered_data = self._apply_complex_filter(filtered_data, column, criteria)
                else:
                    # Простая фильтрация по значению
                    filtered_data = [
                        row for row in filtered_data 
                        if row.get(column) == criteria
                    ]
            
            logger.info(f"Фильтрация завершена: {len(filtered_data)} строк из {len(data)}")
            
            return {
                "status": "success",
                "original_count": len(data),
                "filtered_count": len(filtered_data),
                "data": filtered_data,
                "filters_applied": filters,
                "filtered_at": datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Ошибка при фильтрации данных: {e}")
            raise
    
    def _apply_complex_filter(self, data: List[Dict[str, Any]], 
                             column: str, criteria: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        Применяет сложные критерии фильтрации.
        
        Args:
            data: Данные для фильтрации
            column: Колонка для фильтрации
            criteria: Критерии фильтрации
            
        Returns:
            Отфильтрованные данные
        """
        filtered_data = []
        
        for row in data:
            value = row.get(column)
            if value is None:
                continue
            
            match = True
            
            # Проверяем различные критерии
            if 'equals' in criteria and value != criteria['equals']:
                match = False
            
            if 'not_equals' in criteria and value == criteria['not_equals']:
                match = False
            
            if 'contains' in criteria and criteria['contains'] not in str(value):
                match = False
            
            if 'not_contains' in criteria and criteria['not_contains'] in str(value):
                match = False
            
            if 'starts_with' in criteria and not str(value).startswith(criteria['starts_with']):
                match = False
            
            if 'ends_with' in criteria and not str(value).endswith(criteria['ends_with']):
                match = False
            
            # Числовые критерии
            if isinstance(value, (int, float)):
                if 'greater_than' in criteria and value <= criteria['greater_than']:
                    match = False
                
                if 'less_than' in criteria and value >= criteria['less_than']:
                    match = False
                
                if 'greater_equal' in criteria and value < criteria['greater_equal']:
                    match = False
                
                if 'less_equal' in criteria and value > criteria['less_equal']:
                    match = False
            
            if match:
                filtered_data.append(row)
        
        return filtered_data
    
    @handle_errors(
        category=ErrorCategory.PARSING,
        severity=ErrorSeverity.MEDIUM,
        component="excel_tools",
        recovery_action="Возврат исходных данных",
        return_on_error={"status": "error", "data": [], "message": "Ошибка трансформации данных"}
    )
    async def transform_excel_data(self, data: List[Dict[str, Any]], 
                                  transformations: Dict[str, Any]) -> Dict[str, Any]:
        """
        Трансформирует данные Excel согласно заданным правилам.
        
        Args:
            data: Исходные данные для трансформации
            transformations: Правила трансформации
            
        Returns:
            Трансформированные данные
        """
        try:
            logger.info(f"Трансформация данных по правилам: {transformations}")
            
            transformed_data = []
            
            for row in data:
                transformed_row = row.copy()
                
                # Применяем трансформации
                for column, rules in transformations.items():
                    if column in transformed_row:
                        transformed_row[column] = self._apply_transformation_rules(
                            transformed_row[column], rules
                        )
                
                transformed_data.append(transformed_row)
            
            logger.info(f"Трансформация завершена: {len(transformed_data)} строк")
            
            return {
                "status": "success",
                "original_count": len(data),
                "transformed_count": len(transformed_data),
                "data": transformed_data,
                "transformations_applied": transformations,
                "transformed_at": datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Ошибка при трансформации данных: {e}")
            raise
    
    def _apply_transformation_rules(self, value: Any, rules: Dict[str, Any]) -> Any:
        """
        Применяет правила трансформации к значению.
        
        Args:
            value: Исходное значение
            rules: Правила трансформации
            
        Returns:
            Трансформированное значение
        """
        if value is None:
            return value
        
        # Строковые трансформации
        if isinstance(value, str):
            if rules.get('to_upper'):
                value = value.upper()
            elif rules.get('to_lower'):
                value = value.lower()
            elif rules.get('strip'):
                value = value.strip()
            
            if 'replace' in rules:
                replace_rules = rules['replace']
                if isinstance(replace_rules, dict):
                    for old, new in replace_rules.items():
                        value = value.replace(old, new)
        
        # Числовые трансформации
        if isinstance(value, (int, float)):
            if 'multiply' in rules:
                value = value * rules['multiply']
            elif 'divide' in rules:
                value = value / rules['divide']
            elif 'add' in rules:
                value = value + rules['add']
            elif 'subtract' in rules:
                value = value - rules['subtract']
        
        # Преобразование типов
        if 'convert_to' in rules:
            target_type = rules['convert_to']
            try:
                if target_type == 'int':
                    value = int(float(value))
                elif target_type == 'float':
                    value = float(value)
                elif target_type == 'str':
                    value = str(value)
            except (ValueError, TypeError):
                logger.warning(f"Не удалось преобразовать {value} в {target_type}")
        
        return value
    
    def __del__(self):
        """Очистка ресурсов при удалении объекта."""
        if hasattr(self, 'executor'):
            self.executor.shutdown(wait=False)