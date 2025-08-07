#!/usr/bin/env python3
"""
XML Comparison Tool - Modified Version
Compares multiple XML files against a main XML file in a single table format
Shows only XML tag names and displays all comparisons in columns
"""

import xml.etree.ElementTree as ET
import os
import sys
from pathlib import Path
import argparse
from tabulate import tabulate
import pandas as pd
from collections import defaultdict, OrderedDict
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel output will not be available.")
    print("Install with: pip3 install openpyxl")

class XMLComparator:
    def __init__(self, main_xml_path):
        self.main_xml_path = main_xml_path
        self.main_data = self._parse_xml_to_dict(main_xml_path)
        self.main_flat = self._flatten_dict(self.main_data)
        self.comparison_data = OrderedDict()
        self.all_xml_paths = set()
    
    def _parse_xml_to_dict(self, xml_path):
        """Parse XML file and convert to nested dictionary"""
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            return self._element_to_dict(root)
        except ET.ParseError as e:
            print(f"Error parsing {xml_path}: {e}")
            return {}
        except FileNotFoundError:
            print(f"File not found: {xml_path}")
            return {}
    
    def _element_to_dict(self, element):
        """Convert XML element to dictionary recursively"""
        result = {}
        
        # Add attributes if any
        if element.attrib:
            result['@attributes'] = element.attrib
        
        # Process children
        children = list(element)
        if children:
            child_dict = defaultdict(list)
            for child in children:
                child_data = self._element_to_dict(child)
                child_dict[child.tag].append(child_data)
            
            # Convert single-item lists to single items
            for key, value in child_dict.items():
                if len(value) == 1:
                    result[key] = value[0]
                else:
                    result[key] = value
        
        # Add text content if present and no children
        elif element.text and element.text.strip():
            return element.text.strip()
        
        return result
    
    def _flatten_dict(self, d, parent_key='', sep='/'):
        """Flatten nested dictionary and return only tag names as keys"""
        items = []
        if isinstance(d, dict):
            for k, v in d.items():
                # Skip attributes for cleaner output
                if k.startswith('@'):
                    continue
                    
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(self._flatten_dict(v, new_key, sep=sep).items())
                elif isinstance(v, list):
                    for i, item in enumerate(v):
                        list_key = f"{new_key}[{i}]"
                        if isinstance(item, dict):
                            items.extend(self._flatten_dict(item, list_key, sep=sep).items())
                        else:
                            items.append((list_key, str(item)))
                else:
                    items.append((new_key, str(v)))
        return dict(items)
    
    def add_comparison_file(self, compare_xml_path):
        """Add a comparison XML file to the analysis"""
        compare_data = self._parse_xml_to_dict(compare_xml_path)
        
        if not compare_data:
            return
        
        compare_flat = self._flatten_dict(compare_data)
        filename = Path(compare_xml_path).stem  # Get filename without extension
        
        # Store the flattened data
        self.comparison_data[filename] = compare_flat
        
        # Collect all unique XML paths
        self.all_xml_paths.update(compare_flat.keys())
        self.all_xml_paths.update(self.main_flat.keys())
    
    def compare_multiple_files(self, xml_files_list):
        """Add multiple XML files for comparison"""
        for xml_file in xml_files_list:
            if xml_file != self.main_xml_path:  # Skip main file
                print(f"Processing {xml_file}...")
                self.add_comparison_file(xml_file)
    
    def generate_comparison_table(self, show_only_differences=False, output_file=None):
        """Generate a single comparison table with all files"""
        if not self.comparison_data:
            print("No comparison data found.")
            return
        
        # Prepare data for the table
        table_data = []
        
        main_filename = Path(self.main_xml_path).stem
        
        # Sort XML paths for consistent ordering
        sorted_paths = sorted(self.all_xml_paths)
        
        for xml_path in sorted_paths:
            main_value = self.main_flat.get(xml_path, "MISSING")
            
            # Create row data
            row = {
                'XML_Label': xml_path,
                main_filename: main_value
            }
            
            # Add values from all comparison files
            has_difference = False
            for compare_filename, compare_flat in self.comparison_data.items():
                compare_value = compare_flat.get(xml_path, "MISSING")
                row[compare_filename] = compare_value
                
                # Check if there's a difference
                if compare_value != main_value:
                    has_difference = True
            
            # Add row to table (filter if needed)
            if not show_only_differences or has_difference:
                table_data.append(row)
        
        if not table_data:
            print("No data to display.")
            return
        
        # Create DataFrame
        df = pd.DataFrame(table_data)
        
        # Display table
        print("\n" + "="*120)
        print("XML COMPARISON TABLE")
        print("="*120)
        print(f"Main File: {main_filename}")
        print(f"Compared Files: {', '.join(self.comparison_data.keys())}")
        print("-"*120)
        
        # Format table for better display
        print(tabulate(df, headers=df.columns, tablefmt='grid', showindex=False))
        
        # Save to file if specified
        if output_file:
            df.to_csv(output_file, index=False)
            print(f"\nComparison table saved to: {output_file}")
        
        # Show summary statistics
        self._show_summary_stats(df, main_filename)
    
    def _show_summary_stats(self, df, main_filename):
        """Show summary statistics"""
        print(f"\n" + "="*60)
        print("SUMMARY STATISTICS")
        print("="*60)
        
        total_labels = len(df)
        print(f"Total XML Labels Compared: {total_labels}")
        
        # Count differences for each file
        comparison_files = [col for col in df.columns if col not in ['XML_Label', main_filename]]
        
        stats_data = []
        for compare_file in comparison_files:
            matches = sum(df[main_filename] == df[compare_file])
            differences = total_labels - matches
            missing_in_compare = sum(df[compare_file] == "MISSING")
            extra_in_compare = sum((df[main_filename] == "MISSING") & (df[compare_file] != "MISSING"))
            
            stats_data.append({
                'File': compare_file,
                'Matches': matches,
                'Differences': differences,
                'Missing_in_File': missing_in_compare,
                'Extra_in_File': extra_in_compare,
                'Match_Percentage': f"{(matches/total_labels)*100:.1f}%"
            })
        
        stats_df = pd.DataFrame(stats_data)
        print(tabulate(stats_df, headers=stats_df.columns, tablefmt='grid', showindex=False))
    
    def generate_difference_report(self, output_file=None):
        """Generate a report showing only differences"""
        print(f"\n" + "="*80)
        print("DIFFERENCES ONLY REPORT")
        print("="*80)
        
        self.generate_comparison_table(show_only_differences=True, output_file=output_file)
    
    def generate_specific_label_report(self, label_pattern, output_file=None):
        """Generate a report showing only XML labels that match the specified pattern"""
        if not self.comparison_data:
            print("No comparison data found.")
            return
        
        # Prepare data for the table
        table_data = []
        main_filename = Path(self.main_xml_path).stem
        
        # Filter XML paths based on pattern (case-insensitive)
        filtered_paths = [path for path in sorted(self.all_xml_paths) 
                         if label_pattern.upper() in path.upper()]
        
        if not filtered_paths:
            print(f"No XML labels found matching pattern: '{label_pattern}'")
            return
        
        print(f"\n" + "="*120)
        print(f"XML COMPARISON TABLE - FILTERED BY: '{label_pattern}'")
        print("="*120)
        print(f"Main File: {main_filename}")
        print(f"Compared Files: {', '.join(self.comparison_data.keys())}")
        print(f"Matching Labels: {len(filtered_paths)} out of {len(self.all_xml_paths)} total")
        print("-"*120)
        
        for xml_path in filtered_paths:
            main_value = self.main_flat.get(xml_path, "MISSING")
            
            # Create row data
            row = {
                'XML_Label': xml_path,
                main_filename: main_value
            }
            
            # Add values from all comparison files
            for compare_filename, compare_flat in self.comparison_data.items():
                compare_value = compare_flat.get(xml_path, "MISSING")
                row[compare_filename] = compare_value
            
            table_data.append(row)
        
        if not table_data:
            print("No matching data to display.")
            return
        
        # Create DataFrame
        df = pd.DataFrame(table_data)
        
        # Format table for better display
        print(tabulate(df, headers=df.columns, tablefmt='grid', showindex=False))
        
        # Save to file if specified
        if output_file:
            df.to_csv(output_file, index=False)
            print(f"\nFiltered comparison table saved to: {output_file}")
        
        # Show summary statistics for filtered data
        self._show_summary_stats(df, main_filename)
    
    def generate_excel_output(self, output_file, show_only_differences=False, include_summary=True):
        """Generate Excel output with colors and proper formatting"""
        if not EXCEL_AVAILABLE:
            print("Error: openpyxl is required for Excel output. Install with: pip3 install openpyxl")
            return
        
        if not self.comparison_data:
            print("No comparison data found.")
            return
        
        # Prepare data
        main_filename = Path(self.main_xml_path).stem
        sorted_paths = sorted(self.all_xml_paths)
        
        table_data = []
        for xml_path in sorted_paths:
            main_value = self.main_flat.get(xml_path, "MISSING")
            
            row = {
                'XML_Label': xml_path,
                main_filename: main_value
            }
            
            has_difference = False
            for compare_filename, compare_flat in self.comparison_data.items():
                compare_value = compare_flat.get(xml_path, "MISSING")
                row[compare_filename] = compare_value
                
                if compare_value != main_value:
                    has_difference = True
            
            if not show_only_differences or has_difference:
                table_data.append(row)
        
        if not table_data:
            print("No data to display.")
            return
        
        df = pd.DataFrame(table_data)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "XML Comparison"
        
        # Define colors
        colors = {
            'header': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'main_column': PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid'),
            'match': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),  # Light green
            'different': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),  # Light red
            'missing': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),  # Light yellow
            'extra': PatternFill(start_color='D7E4BD', end_color='D7E4BD', fill_type='solid')    # Light green-yellow
        }
        
        # Define fonts
        header_font = Font(bold=True, color='FFFFFF', size=12)
        regular_font = Font(size=11)
        label_font = Font(bold=True, size=11)
        
        # Define alignment
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = colors['header']
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Write data and apply formatting
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            main_value = str(row[main_filename])
            
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                cell.border = thin_border
                cell.font = regular_font
                
                # Apply specific formatting based on column and value
                if col_name == 'XML_Label':
                    # First column - XML labels
                    cell.font = label_font
                    cell.alignment = left_alignment
                    cell.fill = colors['main_column']
                elif col_name == main_filename:
                    # Main file column
                    cell.alignment = center_alignment
                    cell.fill = colors['main_column']
                else:
                    # Comparison columns
                    cell.alignment = center_alignment
                    compare_value = str(value)
                    
                    if compare_value == "MISSING":
                        cell.fill = colors['missing']
                    elif main_value == "MISSING" and compare_value != "MISSING":
                        cell.fill = colors['extra']
                    elif compare_value == main_value:
                        cell.fill = colors['match']
                    else:
                        cell.fill = colors['different']
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws, df)
        
        # Add summary sheet if requested
        if include_summary:
            self._add_summary_sheet(wb, df, main_filename)
        
        # Add legend sheet
        self._add_legend_sheet(wb)
        
        # Save the workbook
        wb.save(output_file)
        print(f"\nExcel file saved: {output_file}")
        print(f"Sheets created: XML Comparison, Summary Statistics, Legend")
    
    def _adjust_column_widths(self, ws, df):
        """Auto-adjust column widths based on content"""
        from openpyxl.utils import get_column_letter
        
        for col_idx, column in enumerate(df.columns, 1):
            max_length = len(str(column))  # Header length
            
            # Check data in this column
            for row in df.itertuples():
                try:
                    value_length = len(str(getattr(row, column)))
                    max_length = max(max_length, value_length)
                except:
                    pass
            
            # Set width with some padding (max 50 characters)
            adjusted_width = min(max_length + 3, 80)
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_summary_sheet(self, wb, df, main_filename):
        """Add summary statistics sheet"""
        ws_summary = wb.create_sheet(title="Summary Statistics")
        
        # Colors for summary
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        # Title
        ws_summary.cell(row=1, column=1, value="XML Comparison Summary").font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:F1')
        
        # Statistics headers
        stats_headers = ['File', 'Total Labels', 'Matches', 'Differences', 'Missing in File', 'Extra in File', 'Match %']
        for col_idx, header in enumerate(stats_headers, 1):
            cell = ws_summary.cell(row=3, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Calculate statistics
        comparison_files = [col for col in df.columns if col not in ['XML_Label', main_filename]]
        total_labels = len(df)
        
        row_idx = 4
        for compare_file in comparison_files:
            matches = sum(df[main_filename] == df[compare_file])
            differences = total_labels - matches
            missing_in_compare = sum(df[compare_file] == "MISSING")
            extra_in_compare = sum((df[main_filename] == "MISSING") & (df[compare_file] != "MISSING"))
            match_percentage = (matches / total_labels) * 100
            
            stats_row = [compare_file, total_labels, matches, differences, 
                        missing_in_compare, extra_in_compare, f"{match_percentage:.1f}%"]
            
            for col_idx, value in enumerate(stats_row, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Color coding for match percentage
                if col_idx == 7:  # Match percentage column
                    if match_percentage >= 90:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif match_percentage >= 70:
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            row_idx += 1
        
        # Auto-adjust column widths for summary
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(stats_headers) + 1):
            column_letter = get_column_letter(col_idx)
            ws_summary.column_dimensions[column_letter].width = 25
    
    def _add_legend_sheet(self, wb):
        """Add legend/explanation sheet"""
        ws_legend = wb.create_sheet(title="Legend")
        
        # Title
        ws_legend.cell(row=1, column=1, value="Color Legend & Instructions").font = Font(bold=True, size=14)
        
        # Legend items
        legend_items = [
            ("", ""),  # Empty row
            ("Color", "Meaning"),
            ("Green", "Values match between files"),
            ("Red", "Values are different"),
            ("Yellow", "Value missing in comparison file"),
            ("Light Blue", "Main file column"),
            ("", ""),
            ("Instructions:", ""),
            ("1. Green cells = Perfect matches", ""),
            ("2. Red cells = Different values (need attention)", ""),
            ("3. Yellow cells = Missing configurations", ""),
            ("4. Use filters to focus on specific issues", ""),
            ("5. Check Summary sheet for overall statistics", "")
        ]
        
        colors_legend = {
            'Green': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'Red': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'Yellow': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'Light Blue': PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
        }
        
        for row_idx, (item1, item2) in enumerate(legend_items, 1):
            ws_legend.cell(row=row_idx, column=1, value=item1)
            ws_legend.cell(row=row_idx, column=2, value=item2)
            
            # Apply colors to legend
            if item1 in colors_legend:
                ws_legend.cell(row=row_idx, column=1).fill = colors_legend[item1]
        
        # Set column widths
        ws_legend.column_dimensions['A'].width = 20
        ws_legend.column_dimensions['B'].width = 40

def main():
    parser = argparse.ArgumentParser(description='Compare XML files in a single table format')
    parser.add_argument('main_xml', help='Path to the main XML file (main-update.xml)')
    parser.add_argument('compare_files', nargs='+', help='Paths to XML files to compare')
    parser.add_argument('--differences-only', '-d', action='store_true', 
                       help='Show only rows with differences')
    parser.add_argument('--output', '-o', help='Save comparison table to CSV file')
    parser.add_argument('--excel', '-e', help='Save comparison table to Excel file with colors and formatting')
    parser.add_argument('--filter-label', '-f', help='Filter results by XML label pattern')
    parser.add_argument('--no-summary', action='store_true', help='Skip summary sheet in Excel output')
    
    args = parser.parse_args()
    
    # Verify main XML file exists
    if not os.path.exists(args.main_xml):
        print(f"Error: Main XML file '{args.main_xml}' not found.")
        sys.exit(1)
    
    # Initialize comparator
    comparator = XMLComparator(args.main_xml)
    
    # Add comparison files
    comparator.compare_multiple_files(args.compare_files)
    
    # Generate reports based on options
    if args.excel:
        # Excel output
        include_summary = not args.no_summary
        comparator.generate_excel_output(args.excel, 
                                       show_only_differences=args.differences_only,
                                       include_summary=include_summary)
    elif args.filter_label:
        comparator.generate_specific_label_report(args.filter_label)
    elif args.differences_only:
        comparator.generate_difference_report(output_file=args.output)
    else:
        comparator.generate_comparison_table(output_file=args.output)

if __name__ == "__main__":
    # Example usage when run directly
    if len(sys.argv) == 1:
        print("XML Comparison Tool - Single Table Format with Excel Output")
        print("Usage examples:")
        print("  python3 xml_compare.py main-update.xml file1.xml file2.xml file3.xml")
        print("  python3 xml_compare.py main-update.xml *.xml --differences-only")
        print("  python3 xml_compare.py main-update.xml *.xml --excel results.xlsx")
        print("  python3 xml_compare.py main-update.xml *.xml --excel results.xlsx --differences-only")
        print("  python3 xml_compare.py main-update.xml *.xml --filter-label 'CAMERA'")
        print("\nFor 50+ XML files, use --excel option for better visualization!")
        print("Install Excel support: pip3 install openpyxl")
    else:
        main()

