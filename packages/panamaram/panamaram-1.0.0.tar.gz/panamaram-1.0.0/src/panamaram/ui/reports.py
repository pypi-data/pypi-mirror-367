# ui/reports.py

import sqlite3
import calendar
import csv
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QGroupBox, QHBoxLayout, QLabel, QRadioButton,
    QDateEdit, QComboBox, QPushButton, QTableWidget,
    QTableWidgetItem, QFileDialog, QMessageBox, QSpinBox
)
from PySide6.QtCore import QDate
from fpdf import FPDF
from panamaram.utils import path_utils

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

DB_PATH = path_utils.get_secure_db_path(decrypted=True)

class ReportWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Generate Reports")
        self.setMinimumWidth(700)

        main = QVBoxLayout(self)

        # --- Report type selection ---
        type_group = QGroupBox("Report Data")
        type_layout = QHBoxLayout(type_group)
        self.radio_expense = QRadioButton("Expense")
        self.radio_income = QRadioButton("Income")
        self.radio_combined = QRadioButton("Income & Expense Report")
        self.radio_expense.setChecked(True)
        type_layout.addWidget(self.radio_expense)
        type_layout.addWidget(self.radio_income)
        type_layout.addWidget(self.radio_combined)
        main.addWidget(type_group)

        # --- Period selection ---
        period_group = QGroupBox("Report By")
        period_layout = QHBoxLayout(period_group)
        self.radio_day = QRadioButton("Day")
        self.radio_month = QRadioButton("Month")
        self.radio_year = QRadioButton("Year")
        self.radio_month.setChecked(True)
        period_layout.addWidget(self.radio_day)
        period_layout.addWidget(self.radio_month)
        period_layout.addWidget(self.radio_year)
        main.addWidget(period_group)

        # --- Dynamic date selectors ---
        self.day_selector = QDateEdit()
        self.day_selector.setCalendarPopup(True)
        self.day_selector.setDate(QDate.currentDate())
        self.day_selector.setDisplayFormat("dd-MM-yyyy")

        self.month_selector = QComboBox()
        for i in range(1, 13):
            self.month_selector.addItem(calendar.month_name[i], i)

        self.year_selector = QSpinBox()
        self.year_selector.setRange(2000, QDate.currentDate().year())
        self.year_selector.setValue(QDate.currentDate().year())

        self.yearonly_selector = QSpinBox()
        self.yearonly_selector.setRange(2000, QDate.currentDate().year())
        self.yearonly_selector.setValue(QDate.currentDate().year())

        self.period_box = QHBoxLayout()
        self.period_box.addWidget(QLabel("Day:"))
        self.period_box.addWidget(self.day_selector)
        self.period_box.addWidget(QLabel("Month:"))
        self.period_box.addWidget(self.month_selector)
        self.period_box.addWidget(QLabel("Year:"))
        self.period_box.addWidget(self.year_selector)
        self.period_box.addWidget(QLabel("Year Only:"))
        self.period_box.addWidget(self.yearonly_selector)
        main.addLayout(self.period_box)

        self.update_period_inputs()

        widgets_to_watch = [
            self.radio_expense, self.radio_income, self.radio_combined,
            self.radio_day, self.radio_month, self.radio_year,
            self.day_selector, self.month_selector, self.year_selector, self.yearonly_selector
        ]
        for widget in widgets_to_watch:
            if hasattr(widget, 'toggled'):
                widget.toggled.connect(self.on_options_changed)
            if hasattr(widget, 'valueChanged'):
                widget.valueChanged.connect(self.on_options_changed)
            if hasattr(widget, 'currentIndexChanged'):
                widget.currentIndexChanged.connect(self.on_options_changed)
            if hasattr(widget, 'dateChanged'):
                widget.dateChanged.connect(self.on_options_changed)

        self.btn_generate = QPushButton("Generate Report")
        self.btn_generate.clicked.connect(self.generate_report)
        main.addWidget(self.btn_generate)

        self.results_table = QTableWidget(0, 4, self)
        main.addWidget(self.results_table)

        self.btn_export_pdf = QPushButton("Export to PDF")
        self.btn_export_pdf.clicked.connect(self.export_pdf)
        self.btn_export_pdf.setEnabled(False)
        main.addWidget(self.btn_export_pdf)

        self.btn_export_csv = QPushButton("Export to CSV")
        self.btn_export_csv.clicked.connect(self.export_csv)
        self.btn_export_csv.setEnabled(False)
        main.addWidget(self.btn_export_csv)

        self.btn_export_xlsx = QPushButton("Export to XLSX")
        self.btn_export_xlsx.clicked.connect(self.export_xlsx)
        self.btn_export_xlsx.setEnabled(False)
        main.addWidget(self.btn_export_xlsx)

        self.report_data = []
        self.report_totals = {}

    def update_period_inputs(self):
        self.day_selector.setVisible(self.radio_day.isChecked())
        self.month_selector.setVisible(self.radio_month.isChecked())
        self.year_selector.setVisible(self.radio_month.isChecked())
        self.yearonly_selector.setVisible(self.radio_year.isChecked())

        for i in range(self.period_box.count()):
            widget = self.period_box.itemAt(i).widget()
            if not widget:
                continue
            text = widget.text() if hasattr(widget, 'text') else ""
            if text == "Day:":
                widget.setVisible(self.radio_day.isChecked())
            elif text == "Month:":
                widget.setVisible(self.radio_month.isChecked())
            elif text == "Year:":
                widget.setVisible(self.radio_month.isChecked())
            elif text == "Year Only:":
                widget.setVisible(self.radio_year.isChecked())

    def on_options_changed(self, *args):
        self.update_period_inputs()
        self.clear_report_output()

    def clear_report_output(self):
        self.results_table.clear()
        self.results_table.setRowCount(0)
        self.results_table.setColumnCount(0)
        self.results_table.setHorizontalHeaderLabels([])
        self.btn_export_pdf.setEnabled(False)
        self.btn_export_csv.setEnabled(False)
        self.btn_export_xlsx.setEnabled(False)
        self.report_data = []
        self.report_totals = {}

    def generate_report(self):
        report_type = (
            "expense"
            if self.radio_expense.isChecked()
            else "income"
            if self.radio_income.isChecked()
            else "combined"
        )

        if self.radio_day.isChecked():
            by, value = "day", self.day_selector.date().toString("yyyy-MM-dd")
        elif self.radio_month.isChecked():
            by, value = "month", f"{self.year_selector.value():04d}-{self.month_selector.currentData():02d}"
        else:
            by, value = "year", str(self.yearonly_selector.value())

        rows, totals = self.fetch_data(report_type, by, value)
        self.report_data = rows
        self.report_totals = totals

        if report_type == "expense":
            headers = ["Date", "Amount", "Category"]
        elif report_type == "income":
            headers = ["Date", "Amount", "Source"]
        else:
            headers = ["Type", "Date", "Amount", "Category/Source"]

        if not rows:
            self.results_table.clear()
            self.results_table.setRowCount(0)
            self.results_table.setColumnCount(0)
            self.results_table.setHorizontalHeaderLabels([])
            QMessageBox.information(self, "No Data", "No data available for the selected period.")
            self.btn_export_pdf.setEnabled(False)
            self.btn_export_csv.setEnabled(False)
            self.btn_export_xlsx.setEnabled(False)
            return

        self.results_table.setColumnCount(len(headers))
        self.results_table.setRowCount(len(rows))
        self.results_table.setHorizontalHeaderLabels(headers)
        for row_idx, row in enumerate(rows):
            row_to_show = row[:-1] if len(row) > len(headers) else row
            for col_idx, val in enumerate(row_to_show):
                item = QTableWidgetItem(str(val))
                self.results_table.setItem(row_idx, col_idx, item)
        self.results_table.resizeColumnsToContents()

        if rows:
            self.results_table.setRowCount(len(rows) + 1)
            if report_type == "combined":
                self.results_table.setItem(
                    len(rows),
                    2,
                    QTableWidgetItem(
                        f"Total Expense: {totals.get('expense', 0):,.2f} | Total Income: {totals.get('income', 0):,.2f}"
                    ),
                )
            else:
                self.results_table.setItem(
                    len(rows), 1, QTableWidgetItem(f"Total: {totals.get('amount', 0):,.2f}")
                )

        self.btn_export_pdf.setEnabled(True)
        self.btn_export_csv.setEnabled(True)
        self.btn_export_xlsx.setEnabled(True)

    def fetch_data(self, kind, by, value):
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        if kind == "expense":
            if by == "day":
                cursor.execute(
                    "SELECT date, amount, category FROM expenses WHERE date=? ORDER BY date",
                    (value,),
                )
            elif by == "month":
                cursor.execute(
                    "SELECT date, amount, category FROM expenses WHERE substr(date,1,7)=? ORDER BY date",
                    (value,),
                )
            else:
                cursor.execute(
                    "SELECT date, amount, category FROM expenses WHERE substr(date,1,4)=? ORDER BY date",
                    (value,),
                )
            rows = cursor.fetchall()
            total = sum(r[1] for r in rows)
            totals = {"amount": total}

        elif kind == "income":
            if by == "day":
                cursor.execute(
                    "SELECT date, amount, source FROM income WHERE date=? ORDER BY date",
                    (value,),
                )
            elif by == "month":
                cursor.execute(
                    "SELECT date, amount, source FROM income WHERE substr(date,1,7)=? ORDER BY date",
                    (value,),
                )
            else:
                cursor.execute(
                    "SELECT date, amount, source FROM income WHERE substr(date,1,4)=? ORDER BY date",
                    (value,),
                )
            rows = cursor.fetchall()
            total = sum(r[1] for r in rows)
            totals = {"amount": total}

        else:
            if by == "day":
                cursor.execute(
                    "SELECT 'Expense', date, amount, category FROM expenses WHERE date=?",
                    (value,),
                )
                rows_exp = cursor.fetchall()
                cursor.execute(
                    "SELECT 'Income', date, amount, source FROM income WHERE date=?", (value,)
                )
                rows_inc = cursor.fetchall()
            elif by == "month":
                cursor.execute(
                    "SELECT 'Expense', date, amount, category FROM expenses WHERE substr(date,1,7)=?",
                    (value,),
                )
                rows_exp = cursor.fetchall()
                cursor.execute(
                    "SELECT 'Income', date, amount, source FROM income WHERE substr(date,1,7)=?",
                    (value,),
                )
                rows_inc = cursor.fetchall()
            else:
                cursor.execute(
                    "SELECT 'Expense', date, amount, category FROM expenses WHERE substr(date,1,4)=?",
                    (value,),
                )
                rows_exp = cursor.fetchall()
                cursor.execute(
                    "SELECT 'Income', date, amount, source FROM income WHERE substr(date,1,4)=?",
                    (value,),
                )
                rows_inc = cursor.fetchall()
            rows = sorted(rows_exp + rows_inc, key=lambda x: x[1])
            totals = {"expense": sum(r[2] for r in rows_exp), "income": sum(r[2] for r in rows_inc)}

        conn.close()
        return rows, totals

    def export_pdf(self):
        if not self.report_data:
            QMessageBox.warning(self, "No Data", "Generate a report first.")
            return

        fname, _ = QFileDialog.getSaveFileName(self, "Export Report as PDF", "", "PDF Files (*.pdf)")
        if not fname:
            return

        report_type_text = (
            "Expense Report"
            if self.radio_expense.isChecked()
            else "Income Report"
            if self.radio_income.isChecked()
            else "Income & Expense Report"
        )

        if self.radio_day.isChecked():
            period_desc = self.day_selector.date().toString("d MMMM yyyy")
        elif self.radio_month.isChecked():
            month_name = calendar.month_name[self.month_selector.currentData()]
            year_val = self.year_selector.value()
            period_desc = f"{month_name} {year_val}"
        else:
            period_desc = str(self.yearonly_selector.value())

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 16)
        pdf.cell(0, 10, "Panamaram - Personal Finance Expense Tracker", align='C', ln=1)
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, f"{report_type_text} - {period_desc}", align='C', ln=1)
        pdf.set_font("Arial", size=11)

        if report_type_text == "Expense Report":
            headers = ["Date", "Amount", "Category"]
        elif report_type_text == "Income Report":
            headers = ["Date", "Amount", "Source"]
        else:
            headers = ["Type", "Date", "Amount", "Category/Source"]

        col_width = pdf.epw / len(headers)
        for header in headers:
            pdf.cell(col_width, 8, str(header), border=1)
        pdf.ln()

        for row in self.report_data:
            row_to_show = row[:-1] if len(row) > len(headers) else row
            for item in row_to_show:
                pdf.cell(col_width, 8, str(item), border=1)
            pdf.ln()

        pdf.ln(2)
        pdf.set_font("Arial", "B", 12)
        if report_type_text == "Income & Expense Report":
            t = self.report_totals
            pdf.cell(
                0,
                10,
                f"Total Expense: {t.get('expense', 0):,.2f} | Total Income: {t.get('income', 0):,.2f}",
                ln=1,
            )
        else:
            pdf.cell(0, 10, f"Total: {self.report_totals.get('amount', 0):,.2f}", ln=1)

        try:
            pdf.output(fname)
            QMessageBox.information(self, "Exported", f"PDF successfully saved to:\n{fname}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to save PDF: {e}")

    def export_csv(self):
        if not self.report_data:
            QMessageBox.warning(self, "No Data", "Generate a report first.")
            return

        fname, _ = QFileDialog.getSaveFileName(self, "Export Report as CSV", "", "CSV Files (*.csv)")
        if not fname:
            return

        report_type_text = (
            "Expense Report"
            if self.radio_expense.isChecked()
            else "Income Report"
            if self.radio_income.isChecked()
            else "Income & Expense Report"
        )

        if report_type_text == "Expense Report":
            headers = ["Date", "Amount", "Category"]
        elif report_type_text == "Income Report":
            headers = ["Date", "Amount", "Source"]
        else:
            headers = ["Type", "Date", "Amount", "Category/Source"]

        try:
            with open(fname, mode="w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerow(headers)
                for row in self.report_data:
                    row_to_write = row[:-1] if len(row) > len(headers) else row
                    writer.writerow(row_to_write)
            QMessageBox.information(self, "Exported", f"CSV successfully saved to:\n{fname}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to save CSV: {e}")

    def export_xlsx(self):
        if not self.report_data:
            QMessageBox.warning(self, "No Data", "Generate a report first.")
            return

        fname, _ = QFileDialog.getSaveFileName(self, "Export Report as XLSX", "", "Excel Files (*.xlsx)")
        if not fname:
            return

        report_type_text = (
            "Expense Report"
            if self.radio_expense.isChecked()
            else "Income Report"
            if self.radio_income.isChecked()
            else "Income & Expense Report"
        )

        if report_type_text == "Expense Report":
            headers = ["Date", "Amount", "Category"]
        elif report_type_text == "Income Report":
            headers = ["Date", "Amount", "Source"]
        else:
            headers = ["Type", "Date", "Amount", "Category/Source"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # Write title and subtitle
        ws.merge_cells("A1:{}1".format(chr(ord('A') + len(headers) - 1)))
        ws["A1"] = "Panamaram - Personal Finance Expense Tracker"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal='center')

        ws.merge_cells("A2:{}2".format(chr(ord('A') + len(headers) - 1)))
        if self.radio_day.isChecked():
            period_desc = self.day_selector.date().toString("d MMMM yyyy")
        elif self.radio_month.isChecked():
            month_name = calendar.month_name[self.month_selector.currentData()]
            year_val = self.year_selector.value()
            period_desc = f"{month_name} {year_val}"
        else:
            period_desc = str(self.yearonly_selector.value())
        ws["A2"] = f"{report_type_text} - {period_desc}"
        ws["A2"].font = Font(size=12, bold=True)
        ws["A2"].alignment = Alignment(horizontal='center')

        # Write table headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal='left')

        # Write table data
        for row_idx, row in enumerate(self.report_data, start=4):
            row_to_write = row[:-1] if len(row) > len(headers) else row
            for col_idx, value in enumerate(row_to_write, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = Border(left=Side(border_style="thin"),
                                    right=Side(border_style="thin"),
                                    top=Side(border_style="thin"),
                                    bottom=Side(border_style="thin"))
                cell.alignment = Alignment(horizontal='left')

        # Totals
        if report_type_text == "Income & Expense Report":
            t = self.report_totals
            ws.cell(
                row=4 + len(self.report_data),
                column=3,
                value=f"Total Expense: {t.get('expense', 0):,.2f} | Total Income: {t.get('income', 0):,.2f}"
            )
        else:
            ws.cell(
                row=4 + len(self.report_data),
                column=2,
                value=f"Total: {self.report_totals.get('amount', 0):,.2f}"
            )

        # Autosize columns
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 2

        try:
            wb.save(fname)
            QMessageBox.information(self, "Exported", f"XLSX successfully saved to:\n{fname}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to save XLSX: {e}")
