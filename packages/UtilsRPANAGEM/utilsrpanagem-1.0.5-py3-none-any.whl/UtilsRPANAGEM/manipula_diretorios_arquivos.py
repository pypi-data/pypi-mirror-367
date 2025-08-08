import os
from openpyxl import Workbook
import logging
# import logging.handlers
from logging.handlers import TimedRotatingFileHandler

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

import shutil


class Arquivos():
    """Manipula arquivos."""

    def __init__(self):
        self.module_logger = self.criar_logger_diario("./log", nome_arquivo_log="manipula_diretorio_arquivos.log", logger=logging.getLogger(__name__))

    def arquivos_existem(self, *caminhos):
        """
        Verifica se todos os arquivos fornecidos existem.

        :param *caminhos: Um ou mais caminhos de arquivos a serem verificados.
        :type *caminhos: str

        :returns: True se todos os arquivos existem, False caso contrário.
        :rtype: bool
        """
        return all(os.path.isfile(caminho) for caminho in caminhos)

    def criar_arquivo(self, caminho):
        """
        Cria um arquivo vazio caso ele não exista no caminho especificado.

        Cria o(s) diretorio(s) caso não exista.

        :param caminho: Caminho completo do arquivo a ser criado.
        :type caminho: str

        :returns: True se o arquivo foi criado, False caso contrário.
        :rtype: bool
        """
        if not self.arquivos_existem(caminho):
            # Garante que o diretório do arquivo exista
            os.makedirs(os.path.dirname(caminho), exist_ok=True)
            with open(caminho, 'w') as f:
                pass  # Cria um arquivo vazio
            # print(f"✅ Arquivo criado: {caminho}")
            self.module_logger.info(f"✅ Arquivo criado: {caminho}")
            return True
        else:
            # print(f"ℹ️ Arquivo já existe: {caminho}")
            self.module_logger.warning(f"ℹ️ Arquivo já existe: {caminho}")
            return False

    def criar_xlsx_com_colunas(self, caminho_arquivo, colunas=None):
        """
        Cria um arquivo .xlsx com uma planilha contendo as colunas se forem especificadas.

        Sobrescreve o arquivo se ele já existir.

        :param caminho_arquivo: Caminho e nome do arquivo Excel a ser criado.
        :type caminho_arquivo: str
        :param colunas: Lista com os nomes das colunas. (ex: ["Nome", "Idade", "Email"]).
        :type colunas: list

        :returns: True se o caminho for criado com sucesso, False caso contrario
        :rtype: bool
        """
        if not self.arquivos_existem(caminho_arquivo):
            self.criar_arquivo(caminho_arquivo)
        try:
            wb = Workbook()
            ws = wb.active

            if colunas:
                ws.append(colunas)  # Insere os nomes das colunas como primeira linha

            wb.save(caminho_arquivo)
            # print(f"Arquivo '{caminho_arquivo}' criado com as colunas: {colunas}")
            self.module_logger.info(f"Arquivo '{caminho_arquivo}' criado com as colunas: {colunas}")
            return True
        except Exception as erro:
            # print('❌ OCORREU UM ERRO:', erro)
            self.module_logger.error(f"❌ OCORREU UM ERRO: {erro}")
            return False

    def inserir_tabela_xlsx(self, lista_de_dicionarios, caminho_arquivo, dic_nome_colunas=None):
        """
        Monta tabela e salva em uma planilha xlsx.

        :param lista_de_dicionarios: Lista contendo dicionários com os dados.
        :type lista_de_dicionarios: list
        :param caminho_arquivo: Caminho completo do arquivo Excel (.xlsx) a ser formatado.
        :type caminho_arquivo: str
        :param dic_nome_colunas: dicionario contendo as colunas de lista_de_dicionarios como chaves e o nome da coluna que deseja como valor.
        :type dic_nome_colunas: dict
        """
        try:
            self.module_logger.info("Estruturando dados em tabela...")
            df = pd.DataFrame(lista_de_dicionarios)

            # Converte campos datetime para string, se houver
            for col in df.columns:
                if df[col].dtype == 'datetime64[ns]' or df[col].apply(lambda x: str(type(x))).str.contains("datetime").any():
                    df[col] = df[col].dt.strftime('%d/%m/%Y')

            if dic_nome_colunas is not None:
                #  Apenas as colunas indicadas são renomeadas.
                # inplace=True altera o df diretamente.
                df.rename(columns = dic_nome_colunas,
                          inplace=True)

                # substituir todos os nomes das colunas:
                # df.columns = ['CPF', 'APELIDO', 'DATA HORA CRIAÇÃO', 'DATA HORA MODIFICAÇÃO', 'DATA ADMISSÃO', 'STAATV RM', 'NOME', 'LOG', 'CÓDIGO SAP']

            df.to_excel(caminho_arquivo, index=False)
            self.module_logger.info(f"✅ Relatório salvo com sucesso em: {caminho_arquivo}")

        except Exception as e:
            self.module_logger.error(f"❌ Erro ao estruturar dados em tabela: {e}")

    def formatar_xlsx(self, caminho_arquivo, aba="Sheet1"):
        """
        Aplica estilo visual ao cabeçalho de uma planilha Excel e ajusta automaticamente a largura das colunas.

        :param caminho_arquivo: Caminho completo do arquivo Excel (.xlsx) a ser formatado.
        :type caminho_arquivo: str
        :param aba: Nome da aba onde aplicar o estilo. Padrão é 'Sheet1'.
        :type aba: str

        Estilos aplicados:
        - Cabeçalho com fundo azul (cor: #0070C0)
        - Texto do cabeçalho em branco, negrito e centralizado
        - Largura das colunas ajustada com base no maior conteúdo da coluna
        """
        try:
            self.module_logger.info("Formatando o estilo do relatorio...")

            wb = load_workbook(caminho_arquivo)

            try:
                ws = wb[aba]
            except:
                ws = wb["Sheet"]

            # Definição do estilo azul
            cabecalho_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            fonte_branca = Font(color="FFFFFF", bold=True)
            alinhamento_central = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:  # Primeira linha = cabeçalho
                cell.fill = cabecalho_fill
                cell.font = fonte_branca
                cell.alignment = alinhamento_central

            wb.save(caminho_arquivo)
            self.module_logger.info("✅ Formatação aplicada")

        except Exception as e:
            self.module_logger.error(f"❌ Erro ao formatar o estilo do relatorio: {e}")


        try:
            self.module_logger.info("Ajustando a largura das colunas do relatorio...")

            wb = load_workbook(caminho_arquivo)
            try:
                ws = wb[aba]
            except:
                ws = wb["Sheet"]

            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 4  # margem extra

            wb.save(caminho_arquivo)

            self.module_logger.info("✅ Estilo e largura aplicados")
        except Exception as e:
            self.module_logger.error(f"❌ Erro ao aplicar largura: {e}")

        self.module_logger.info(f"✅ Relatório formatado e salvo com sucesso em: {caminho_arquivo}")

    def criar_logger_diario(self, diretorio_log, nome_arquivo_log="log", logger=logging.getLogger()):
        """
        Cria e configura um logger que grava logs com rotação diária, criando um novo arquivo a cada dia.

        :param diretorio_log: Caminho para o diretório onde os logs serão salvos.
        :type diretorio_log: str
        :param nome_arquivo_log: Nome base do arquivo de log (ex: "log" -> log_2025-06-20.log).
        :type nome_arquivo_log: str
        :param logger: Logger nomeado para __name__.
        :type logger: logging.Logger

        :returns: Logger configurado com rotação diária.
        :rtype: logging.Logger
        """
        # Caminho base para logs rotacionados: log/log_
        dir_arq_log = os.path.join(diretorio_log, nome_arquivo_log)

        try:
            # self.criar_arquivo(dir_arq_log)
            os.makedirs(diretorio_log, exist_ok=True)

            logger.setLevel(logging.DEBUG)
        except Exception as e:
            raise Exception(f"❌ Rootlogger não pode ser iniciado.: {e}")

        try:
            if not logger.handlers:
                formatter = logging.Formatter(
                    "%(name)s - %(levelname)s - %(asctime)s - %(message)s"
                )

                # Handler com rotação diária
                handler = TimedRotatingFileHandler(
                    filename=dir_arq_log,
                    when='midnight',             # rotação diária
                    interval=1,                  # a cada 1 dia
                    backupCount=30,              # mantém últimos 30 arquivos
                    encoding='utf-8',
                    utc=False
                )
                handler.suffix = "%Y-%m-%d.log"  # formato do nome final do arquivo
                handler.setFormatter(formatter)

                # Também loga no console
                console_handler = logging.StreamHandler()
                console_handler.setFormatter(formatter)

                logger.addHandler(handler)
                logger.addHandler(console_handler)

            return logger
        except Exception as e:
            raise Exception(f"O handler do arquivo de log não pode ser configurado. {e}")

    def mover_arquivo(self, origem, destino_pasta):
        """Move um arquivo para outro diretorio.

        :param origem: Caminho completo do arquivo.
        :type origem: str
        :param destino_pasta: Caminho para onde se quer mover o arquivo
        :type destino_pasta: str
        """
        nome_arquivo = os.path.basename(origem)
        destino_arquivo = os.path.join(destino_pasta, nome_arquivo)

        if self.arquivos_existem(origem) is True:
            if os.path.isdir(destino_pasta) is True:

                shutil.move(origem, destino_arquivo)
                self.module_logger.info("✅ Arquivo movido com sucesso!")
            else:
                self.module_logger.warning("❌ Diretorio de destino não encontrado, verifique se caminho esta correto.")
                raise FileNotFoundError("❌ Diretorio de destino não encontrado, verifique se caminho esta correto.")
        else:
            self.module_logger.warning("❌ Arquivo não encontrado, verifique se o caminho esta correto.")
            raise FileNotFoundError("❌ Arquivo não encontrado, verifique se o caminho esta correto.")


class Diretorios():
    """Manipula diretórios."""

    def __init__(self):
        obj_arq = Arquivos()
        self.module_logger = obj_arq.criar_logger_diario("./log", nome_arquivo_log="manipula_diretorio_arquivos.log", logger=logging.getLogger(__name__))

    def diretorio_existe(self, caminho):
        """
        Verifica se um diretório existe no caminho especificado.

        :param caminho: Caminho do diretório a ser verificado.
        :type caminho: str

        :returns: True se o diretório existe, False caso contrário.
        :rtype: bool
        """
        return os.path.isdir(caminho)

    def criar_diretorio(self, caminho):
        """
        Cria o diretório especificado caso ele não exista.

        :param caminho: Caminho do diretório a ser verificado ou criado.
        :type caminho: str

        :returns: True se o diretório foi criado, False caso contrário.
        :rtype: bool
        """
        if not self.diretorio_existe(caminho):
            os.makedirs(caminho)
            # print(f"✅ Diretório criado: {caminho}")
            self.module_logger.info(f"✅ Diretório criado: {caminho}")
            return True
        else:
            # print(f"ℹ️ Diretório já existe: {caminho}")
            self.module_logger.warning(f"ℹ️ Diretório já existe: {caminho}")
            return False
